from .models import Usuario, Restaurante, UserRoles, Item, Area, Fornecedor, FornecedorItemCodigo, Estoque, Cotacao, CotacaoStatus, CotacaoItem, Pedido, PedidoStatus, Lista, ListaMaeItem, ListaItemRef, Submissao, SubmissaoStatus, SugestaoItem, SugestaoStatus, ListaRapida, ListaRapidaItem, StatusListaRapida, PrioridadeItem, ConviteToken, ConviteRestaurante, Checklist, ChecklistStatus, ChecklistItem, POPConfiguracao, POPCategoria, POPTemplate, POPLista, POPListaTarefa, POPExecucao, POPExecucaoItem, TipoVerificacao, CriticidadeTarefa, RecorrenciaLista, StatusExecucao, Notificacao, TipoNotificacao, AppLog, ConviteFornecedor, ItemPrecoHistorico, brasilia_now
from .extensions import db
from . import repositories
from werkzeug.security import generate_password_hash, check_password_hash
from flask_jwt_extended import create_access_token, get_jwt
from uuid import uuid4
from datetime import datetime, timedelta, timezone
import time
from io import StringIO
import csv
from sqlalchemy import func, or_
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import IntegrityError
from flask import current_app, request
import re
import unicodedata

_SUPER_DASHBOARD_CACHE = {}
_SUPER_DASHBOARD_CACHE_TTL_SECONDS = 60

def _is_admin_or_super_admin(usuario):
    """Verifica se o usuário possui privilégios de administrador."""
    return usuario and usuario.role in (UserRoles.ADMIN, UserRoles.SUPER_ADMIN)

def _get_super_dashboard_cache_key(restaurante_id, period):
    return f"{restaurante_id or 'all'}:{period}"

def _get_restaurante_padrao_id():
    restaurante = Restaurante.query.filter_by(slug='kzn', deletado=False).first()
    if restaurante:
        return restaurante.id
    restaurante = Restaurante.query.filter_by(deletado=False).order_by(Restaurante.id).first()
    return restaurante.id if restaurante else None

def _get_restaurante_regional_id():
    restaurante = Restaurante.query.filter_by(
        slug='fornecedores-regionais',
        deletado=False
    ).first()
    if not restaurante:
        restaurante = Restaurante.query.filter_by(
            nome='Fornecedores Regionais',
            deletado=False
        ).first()
    return restaurante.id if restaurante else None

def _get_usuario_restaurante_id(user_id):
    usuario = db.session.get(Usuario, user_id)
    return usuario.restaurante_id if usuario else None

def _get_fornecedor_por_id(fornecedor_id, restaurante_id=None):
    query = Fornecedor.query.filter_by(id=fornecedor_id)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    return query.first()

def _get_item_catalogo_por_id(item_id, restaurante_id=None):
    query = ListaMaeItem.query.filter_by(id=item_id)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    return query.first()

def _parse_iso_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    value_str = str(value)
    if value_str.endswith("Z"):
        value_str = value_str[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(value_str)
    except (ValueError, TypeError):
        return None

def _normalize_datetime(value):
    if value is None:
        return None
    if value.tzinfo is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)

def log_event(acao, entidade=None, entidade_id=None, mensagem=None, usuario_id=None, restaurante_id=None, metadata=None, impersonator_id=None):
    try:
        claims = {}
        try:
            claims = get_jwt()
        except Exception:
            claims = {}

        if usuario_id is None:
            usuario_id = claims.get('sub')
            try:
                usuario_id = int(usuario_id) if usuario_id is not None else None
            except (TypeError, ValueError):
                usuario_id = None

        if restaurante_id is None:
            restaurante_id = claims.get('restaurante_id')

        if impersonator_id is None:
            impersonator_id = claims.get('impersonated_by')
            try:
                impersonator_id = int(impersonator_id) if impersonator_id is not None else None
            except (TypeError, ValueError):
                impersonator_id = None

        ip_address = None
        user_agent = None
        try:
            ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
            user_agent = request.headers.get('User-Agent')
        except Exception:
            pass

        log = AppLog(
            acao=acao,
            entidade=entidade,
            entidade_id=entidade_id,
            mensagem=mensagem,
            meta=metadata or {},
            restaurante_id=restaurante_id,
            usuario_id=usuario_id,
            impersonator_id=impersonator_id,
            ip_address=ip_address,
            user_agent=user_agent
        )
        db.session.add(log)
        return log
    except Exception as e:
        current_app.logger.warning(f"[LOG] Falha ao registrar log: {str(e)}")
        return None

def listar_logs(filters):
    query = AppLog.query

    restaurante_id = filters.get('restaurante_id')
    if restaurante_id is not None:
        query = query.filter(AppLog.restaurante_id == restaurante_id)

    usuario_id = filters.get('usuario_id')
    if usuario_id is not None:
        query = query.filter(AppLog.usuario_id == usuario_id)

    acao = filters.get('acao')
    if acao:
        query = query.filter(AppLog.acao == acao)

    entidade = filters.get('entidade')
    if entidade:
        query = query.filter(AppLog.entidade == entidade)

    start_date = _parse_iso_datetime(filters.get('start_date'))
    if start_date:
        query = query.filter(AppLog.criado_em >= start_date)

    end_date = _parse_iso_datetime(filters.get('end_date'))
    if end_date:
        query = query.filter(AppLog.criado_em <= end_date)

    total = query.count()

    limit = filters.get('limit', 200)
    offset = filters.get('offset', 0)
    query = query.order_by(AppLog.criado_em.desc()).offset(offset).limit(limit)

    logs = []
    for log in query.all():
        log_dict = log.to_dict()
        if log.usuario:
            log_dict['usuario'] = {
                "id": log.usuario.id,
                "nome": log.usuario.nome,
                "email": log.usuario.email,
                "role": log.usuario.role.value if hasattr(log.usuario.role, 'value') else log.usuario.role,
            }
        if log.restaurante:
            log_dict['restaurante'] = {
                "id": log.restaurante.id,
                "nome": log.restaurante.nome,
                "slug": log.restaurante.slug,
            }
        if log.impersonator:
            log_dict['impersonator'] = {
                "id": log.impersonator.id,
                "nome": log.impersonator.nome,
                "email": log.impersonator.email,
            }
        logs.append(log_dict)

    return {"logs": logs, "total": total}, 200

def register_user(data):
    """Cria um novo usuário no sistema."""
    if Usuario.query.filter_by(email=data['email']).first():
        return {"error": "E-mail já cadastrado."}, 409

    # Verifica se username já existe (se fornecido)
    if data.get('username') and Usuario.query.filter_by(username=data['username']).first():
        return {"error": "Nome de usuário já cadastrado."}, 409

    hashed_password = generate_password_hash(data['senha'])

    # Verifica se foi fornecido token de admin
    TOKEN_ADMIN = "Kaiser@210891"
    token_fornecido = data.get('token_admin')

    # Se token fornecido e correto, cria como ADMIN aprovado
    if token_fornecido and token_fornecido == TOKEN_ADMIN:
        role = UserRoles.ADMIN
        aprovado = True
        mensagem = "Administrador criado e aprovado automaticamente!"
    else:
        # Comportamento padrão: COLLABORATOR não aprovado
        role = UserRoles.COLLABORATOR
        aprovado = False
        mensagem = "Solicitação de cadastro enviada com sucesso. Aguardando aprovação do administrador."

    new_user = Usuario(
        nome=data['nome'],
        username=data.get('username'),
        email=data['email'],
        senha_hash=hashed_password,
        role=role,
        restaurante_id=_get_restaurante_padrao_id(),
        aprovado=aprovado
    )

    db.session.add(new_user)
    db.session.flush()
    log_event(
        acao="create",
        entidade="usuario",
        entidade_id=new_user.id,
        mensagem="Cadastro de usuario criado",
        restaurante_id=new_user.restaurante_id,
        usuario_id=new_user.id
    )
    db.session.commit()

    return {"message": mensagem}, 201

def authenticate_user(data):
    """Autentica um usuário e retorna um token JWT."""
    # Aceita login com email ou username
    login_field = data.get('email') or data.get('username')

    # Busca por email ou username
    user = Usuario.query.filter(
        (Usuario.email == login_field) | (Usuario.username == login_field)
    ).first()

    if not user or not check_password_hash(user.senha_hash, data['senha']):
        return {"error": "Credenciais inválidas."}, 401

    if not user.aprovado:
        return {"error": "Usuário pendente de aprovação."}, 403

    if not user.ativo:
        return {"error": "Usuário desativado. Entre em contato com o administrador."}, 403

    # Atualiza token de sessão para garantir apenas um login ativo
    session_token = str(uuid4())
    user.session_token = session_token
    user.session_updated_at = brasilia_now()
    log_event(
        acao="login",
        entidade="usuario",
        entidade_id=user.id,
        mensagem="Login realizado",
        restaurante_id=user.restaurante_id,
        usuario_id=user.id
    )
    db.session.commit()

    # O identity deve ser uma STRING (Flask-JWT-Extended espera string no campo 'sub')
    # Dados adicionais vão em additional_claims
    additional_claims = {
        "role": user.role.value,
        "nome": user.nome,
        "email": user.email,
        "restaurante_id": user.restaurante_id,
        "wizard_status": user.wizard_status or {},
        "session_token": session_token,
    }
    expires = timedelta(days=1)
    access_token = create_access_token(
        identity=str(user.id),  # Converte ID para string
        additional_claims=additional_claims,  # Role, nome e outros dados extras
        expires_delta=expires
    )

    return {"access_token": access_token}, 200

def iniciar_impersonacao(super_admin_id, usuario_id):
    """Gera token para SUPER_ADMIN atuar como outro usuario."""
    super_admin = db.session.get(Usuario, super_admin_id)
    if not super_admin or super_admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode impersonar."}, 403

    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    additional_claims = {
        "role": usuario.role.value,
        "nome": usuario.nome,
        "email": usuario.email,
        "restaurante_id": usuario.restaurante_id,
        "wizard_status": usuario.wizard_status or {},
        "session_token": usuario.session_token,
        "impersonated_by": super_admin.id,
        "impersonated_by_nome": super_admin.nome,
        "impersonated_by_email": super_admin.email,
    }

    access_token = create_access_token(
        identity=str(usuario.id),
        additional_claims=additional_claims,
        expires_delta=timedelta(days=1)
    )

    log_event(
        acao="impersonate_start",
        entidade="usuario",
        entidade_id=usuario.id,
        mensagem=f"Impersonacao iniciada por {super_admin.nome}",
        restaurante_id=usuario.restaurante_id,
        usuario_id=usuario.id,
        metadata={"impersonated_by": super_admin.id},
        impersonator_id=super_admin.id
    )
    db.session.commit()

    return {"access_token": access_token}, 200

def encerrar_impersonacao(super_admin_id):
    """Gera novo token para SUPER_ADMIN encerrando impersonacao."""
    super_admin = db.session.get(Usuario, super_admin_id)
    if not super_admin or super_admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode encerrar impersonação."}, 403

    additional_claims = {
        "role": super_admin.role.value,
        "nome": super_admin.nome,
        "email": super_admin.email,
        "restaurante_id": super_admin.restaurante_id,
        "wizard_status": super_admin.wizard_status or {},
        "session_token": super_admin.session_token,
    }

    access_token = create_access_token(
        identity=str(super_admin.id),
        additional_claims=additional_claims,
        expires_delta=timedelta(days=1)
    )

    log_event(
        acao="impersonate_end",
        entidade="usuario",
        entidade_id=super_admin.id,
        mensagem="Impersonacao encerrada",
        restaurante_id=super_admin.restaurante_id,
        usuario_id=super_admin.id,
        impersonator_id=super_admin.id
    )
    db.session.commit()

    return {"access_token": access_token}, 200

def approve_user(user_id, current_role=None, current_restaurante_id=None):
    """Aprova o cadastro de um usuário."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404
    if current_role != UserRoles.SUPER_ADMIN.value:
        if user.role == UserRoles.SUPER_ADMIN:
            return {"error": "Acesso negado."}, 403
        if user.restaurante_id != current_restaurante_id:
            return {"error": "Acesso negado."}, 403

    user.aprovado = True
    log_event(
        acao="approve",
        entidade="usuario",
        entidade_id=user.id,
        mensagem="Usuario aprovado",
        restaurante_id=user.restaurante_id,
        usuario_id=user.id
    )
    db.session.commit()
    return {"message": f"Usuário {user.nome} aprovado com sucesso."}, 200

def update_user_by_admin(user_id, data, current_role=None, current_restaurante_id=None):
    """Atualiza os dados de um usuário a pedido de um admin."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404
    if current_role != UserRoles.SUPER_ADMIN.value:
        if user.role == UserRoles.SUPER_ADMIN:
            return {"error": "Acesso negado."}, 403
        if user.restaurante_id != current_restaurante_id:
            return {"error": "Acesso negado."}, 403

    # Validação de email duplicado
    if 'email' in data and data['email'] != user.email:
        if Usuario.query.filter_by(email=data['email']).first():
            return {"error": "E-mail já cadastrado."}, 409

    # Atualiza os campos
    user.nome = data.get('nome', user.nome)
    user.email = data.get('email', user.email)
    
    # Atualiza o role, se fornecido e válido
    if 'role' in data and data['role'] in [r.value for r in UserRoles]:
        if data['role'] == UserRoles.SUPER_ADMIN.value:
            return {"error": "Não é permitido definir SUPER_ADMIN via API."}, 403
        if current_role != UserRoles.SUPER_ADMIN.value and data['role'] != UserRoles.COLLABORATOR.value:
            return {"error": "Apenas SUPER_ADMIN pode definir ADMIN."}, 403
        user.role = UserRoles(data['role'])

    if 'restaurante_id' in data and current_role == UserRoles.SUPER_ADMIN.value:
        restaurante = Restaurante.query.filter_by(id=data['restaurante_id'], deletado=False).first()
        if not restaurante:
            return {"error": "Restaurante não encontrado."}, 404
        user.restaurante_id = restaurante.id

    log_event(
        acao="update",
        entidade="usuario",
        entidade_id=user.id,
        mensagem="Usuario atualizado por admin",
        restaurante_id=user.restaurante_id,
        usuario_id=user.id
    )
    db.session.commit()
    return {"message": "Usuário atualizado com sucesso.", "user": user.to_dict()}, 200

def create_user_by_admin(data, current_role=None, current_restaurante_id=None):
    """Cria um novo usuário (colaborador ou admin) com controle multi-tenant."""
    try:
        print("[SERVICE] Iniciando criacao de usuario...")
        print(f"[SERVICE] Email: {data.get('email')}")
        print(f"[SERVICE] Nome: {data.get('nome')}")
        print(f"[SERVICE] Username: {data.get('username')}")
        print(f"[SERVICE] Role: {data.get('role')}")

        # Validação de email duplicado
        if Usuario.query.filter_by(email=data['email']).first():
            print(f"[SERVICE] E-mail ja existe: {data['email']}")
            return {"error": "E-mail já cadastrado."}, 409

        # Verifica se username já existe (se fornecido)
        if data.get('username'):
            existing_username = Usuario.query.filter_by(username=data['username']).first()
            if existing_username:
                print(f"[SERVICE] Username ja existe: {data['username']}")
                return {"error": "Nome de usuário já cadastrado."}, 409

        role_value = data.get('role', UserRoles.COLLABORATOR.value)
        if role_value not in [r.value for r in UserRoles]:
            return {"error": "Perfil inválido."}, 400
        if role_value == UserRoles.SUPER_ADMIN.value:
            return {"error": "Não é permitido criar SUPER_ADMIN via API."}, 403
        if current_role and current_role != UserRoles.SUPER_ADMIN.value and role_value != UserRoles.COLLABORATOR.value:
            return {"error": "Apenas SUPER_ADMIN pode criar ADMIN."}, 403

        target_restaurante_id = None
        if current_role == UserRoles.SUPER_ADMIN.value:
            target_restaurante_id = data.get('restaurante_id')
        else:
            target_restaurante_id = current_restaurante_id

        if role_value != UserRoles.SUPER_ADMIN.value and not target_restaurante_id:
            return {"error": "Restaurante é obrigatório."}, 400

        restaurante = None
        if target_restaurante_id:
            restaurante = Restaurante.query.filter_by(id=target_restaurante_id, deletado=False).first()
            if not restaurante:
                return {"error": "Restaurante não encontrado."}, 404

        # Hash da senha
        senha_original = data['senha']
        hashed_password = generate_password_hash(senha_original)
        print("[SERVICE] Senha hashada com sucesso")

        role = UserRoles(role_value)
        print(f"[SERVICE] Role definida: {role}")

        new_user = Usuario(
            nome=data['nome'],
            username=data.get('username') if data.get('username') else None,
            email=data['email'],
            senha_hash=hashed_password,
            senha_texto_puro=senha_original,  # Armazena senha em texto puro
            role=role,
            restaurante_id=restaurante.id if restaurante else None,
            aprovado=True
        )

        print("[SERVICE] Objeto Usuario criado")

        db.session.add(new_user)
        print("[SERVICE] Usuario adicionado a sessao")
        db.session.flush()
        log_event(
            acao="create",
            entidade="usuario",
            entidade_id=new_user.id,
            mensagem="Usuario criado por admin",
            restaurante_id=new_user.restaurante_id,
            usuario_id=new_user.id
        )

        db.session.commit()
        print(f"[SERVICE] Usuario salvo no banco! ID: {new_user.id}")

        return {"message": f"Usuário {new_user.nome} criado com sucesso como {role.value}."}, 201

    except KeyError as e:
        print(f"[SERVICE] Campo obrigatorio ausente: {e}")
        return {"error": f"Campo obrigatório ausente: {str(e)}"}, 400

    except Exception as e:
        print(f"[SERVICE] Erro inesperado: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return {"error": f"Erro ao criar usuário: {str(e)}"}, 500


def get_all_users(restaurante_id=None, current_role=None, status=None):
    """Retorna usuários cadastrados com filtro opcional por restaurante."""
    query = Usuario.query

    if status == 'pending':
        query = query.filter_by(aprovado=False)

    if current_role != UserRoles.SUPER_ADMIN.value:
        query = query.filter(Usuario.role != UserRoles.SUPER_ADMIN)
        if restaurante_id is not None:
            query = query.filter_by(restaurante_id=restaurante_id)
    else:
        if restaurante_id is not None:
            query = query.filter_by(restaurante_id=restaurante_id)

    return query.order_by(Usuario.nome).all(), 200

def compartilhar_usuario_whatsapp(user_id: int, current_role=None, current_restaurante_id=None):
    """Gera texto formatado para compartilhar credenciais do usuário via WhatsApp."""
    from kaizen_app.models import Usuario

    user = Usuario.query.get(user_id)
    if not user:
        raise ValueError("Usuário não encontrado")
    if current_role != UserRoles.SUPER_ADMIN.value:
        if user.role == UserRoles.SUPER_ADMIN:
            raise ValueError("Acesso negado.")
        if user.restaurante_id != current_restaurante_id:
            raise ValueError("Acesso negado.")

    senha_atual = (user.senha_texto_puro or '').strip()
    if not senha_atual:
        raise ValueError("Senha atual não disponível. Redefina a senha do usuário antes de compartilhar.")

    # Tradução de roles
    role_label = "Administrador" if user.role.value == "ADMIN" else "Colaborador"
    status_aprovacao = "✅ Aprovado" if user.aprovado else "⏳ Pendente de Aprovação"
    status_ativo = "✅ Ativo" if user.ativo else "🔴 Inativo"

    # Formatar data de criação
    data_criacao = user.criado_em.strftime("%d/%m/%Y às %H:%M") if user.criado_em else "Não disponível"

    # Gerar texto formatado para WhatsApp
    linhas = [
        "*🔐 Credenciais de Acesso - Kaizen Listas*",
        "",
        f"*Nome:* {user.nome}",
        f"*E-mail:* {user.email}",
        f"*Senha:* `{senha_atual}`",
        f"*Perfil:* {role_label}",
        f"*Status:* {status_aprovacao}",
        f"*Situação:* {status_ativo}",
        f"*Criado em:* {data_criacao}",
        "",
        "*📱 Como Acessar:*",
        "1. Acesse o sistema pelo navegador",
        "2. Use o e-mail acima como login",
        "3. Use a senha fornecida acima",
        "",
        "⚠️ *Importante:* Por segurança, *altere sua senha* após o primeiro acesso em *Perfil > Mudar Senha*.",
        "",
        "---",
        f"_Mensagem gerada em {brasilia_now().strftime('%d/%m/%Y às %H:%M')}_"
    ]

    texto = "\n".join(linhas)

    return {"texto": texto}

def change_password(user_id, data):
    """Altera a senha de um usuário."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404

    # Verifica se a senha atual está correta
    if not check_password_hash(user.senha_hash, data.get('senha_atual')):
        return {"error": "Senha atual incorreta."}, 401

    # Verifica se nova senha e confirmação são iguais
    if data.get('nova_senha') != data.get('confirmar_senha'):
        return {"error": "A nova senha e a confirmação não coincidem."}, 400

    # Atualiza a senha
    nova_senha = data.get('nova_senha')
    user.senha_hash = generate_password_hash(nova_senha)
    user.senha_texto_puro = nova_senha
    db.session.commit()

    return {"message": "Senha alterada com sucesso."}, 200

def update_user_profile(user_id, data):
    """Atualiza o perfil de um usuário."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404

    # Verifica se o email já está em uso por outro usuário
    if 'email' in data and data['email'] != user.email:
        existing_user = Usuario.query.filter_by(email=data['email']).first()
        if existing_user:
            return {"error": "E-mail já está em uso."}, 409

    # Verifica se o username já está em uso por outro usuário
    if 'username' in data and data['username'] != user.username:
        existing_user = Usuario.query.filter_by(username=data['username']).first()
        if existing_user:
            return {"error": "Nome de usuário já está em uso."}, 409

    # Atualiza os campos permitidos
    if 'nome' in data:
        user.nome = data['nome']
    if 'username' in data:
        user.username = data['username']
    if 'email' in data:
        user.email = data['email']

    db.session.commit()

    return {"message": "Perfil atualizado com sucesso.", "user": user.to_dict()}, 200

def update_wizard_status(user_id, wizard_status):
    """Atualiza o progresso do wizard do usuário logado."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404

    if not isinstance(wizard_status, dict):
        return {"error": "wizardStatus deve ser um objeto válido."}, 400

    current_status = user.wizard_status or {}
    user.wizard_status = {**current_status, **wizard_status}
    db.session.commit()

    return {"wizard_status": user.wizard_status}, 200

def get_user_profile(user_id):
    """Retorna o perfil de um usuário."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404

    return user.to_dict(), 200

def delete_user(user_id, current_role=None, current_restaurante_id=None):
    """Deleta um usuário e todos os registros relacionados."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404
    if current_role != UserRoles.SUPER_ADMIN.value:
        if user.role == UserRoles.SUPER_ADMIN:
            return {"error": "Acesso negado."}, 403
        if user.restaurante_id != current_restaurante_id:
            return {"error": "Acesso negado."}, 403

    # Remove relacionamentos many-to-many com listas
    user.listas_atribuidas = []
    db.session.flush()

    # SQLAlchemy vai lidar com cascade delete automático para registros relacionados
    log_event(
        acao="delete",
        entidade="usuario",
        entidade_id=user.id,
        mensagem="Usuario deletado",
        restaurante_id=user.restaurante_id,
        usuario_id=user.id
    )
    db.session.delete(user)
    db.session.commit()

    return {"message": f"Usuário {user.nome} deletado com sucesso."}, 200

def deactivate_user(user_id, current_role=None, current_restaurante_id=None):
    """Desativa um usuário (não conseguirá fazer login)."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404
    if current_role != UserRoles.SUPER_ADMIN.value:
        if user.role == UserRoles.SUPER_ADMIN:
            return {"error": "Acesso negado."}, 403
        if user.restaurante_id != current_restaurante_id:
            return {"error": "Acesso negado."}, 403

    if not user.ativo:
        return {"error": "Usuário já está desativado."}, 400

    user.ativo = False
    log_event(
        acao="deactivate",
        entidade="usuario",
        entidade_id=user.id,
        mensagem="Usuario desativado",
        restaurante_id=user.restaurante_id,
        usuario_id=user.id
    )
    db.session.commit()

    return {"message": f"Usuário {user.nome} desativado com sucesso."}, 200

def reactivate_user(user_id, current_role=None, current_restaurante_id=None):
    """Reativa um usuário desativado."""
    user = db.session.get(Usuario, user_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404
    if current_role != UserRoles.SUPER_ADMIN.value:
        if user.role == UserRoles.SUPER_ADMIN:
            return {"error": "Acesso negado."}, 403
        if user.restaurante_id != current_restaurante_id:
            return {"error": "Acesso negado."}, 403

    if user.ativo:
        return {"error": "Usuário já está ativo."}, 400

    user.ativo = True
    log_event(
        acao="reactivate",
        entidade="usuario",
        entidade_id=user.id,
        mensagem="Usuario reativado",
        restaurante_id=user.restaurante_id,
        usuario_id=user.id
    )
    db.session.commit()

    return {"message": f"Usuário {user.nome} reativado com sucesso."}, 200


# --- Serviços de Inventário ---

def create_item(data, restaurante_id=None):
    fornecedor_id = data.get('fornecedor_id')
    fornecedor = _get_fornecedor_por_id(fornecedor_id, restaurante_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404
    new_item = Item(
        nome=data['nome'],
        unidade_medida=data['unidade_medida'],
        fornecedor_id=fornecedor.id
    )
    repositories.add_instance(new_item)
    log_event(
        acao="create",
        entidade="item",
        entidade_id=new_item.id,
        mensagem="Item criado",
        restaurante_id=fornecedor.restaurante_id
    )
    db.session.commit()
    return {
        "id": new_item.id,
        "nome": new_item.nome,
        "unidade_medida": new_item.unidade_medida,
        "fornecedor_id": new_item.fornecedor_id
    }, 201

def get_all_items(restaurante_id=None):
    query = Item.query
    if restaurante_id is not None:
        query = query.join(Fornecedor).filter(Fornecedor.restaurante_id == restaurante_id)
    return query.all(), 200

def get_item_by_id(item_id, restaurante_id=None):
    query = Item.query.filter(Item.id == item_id)
    if restaurante_id is not None:
        query = query.join(Fornecedor).filter(Fornecedor.restaurante_id == restaurante_id)
    return query.first(), 200

def update_item(item_id, data, restaurante_id=None):
    item_query = Item.query.filter(Item.id == item_id)
    if restaurante_id is not None:
        item_query = item_query.join(Fornecedor).filter(Fornecedor.restaurante_id == restaurante_id)
    item = item_query.first()
    if not item:
        return {"error": "Item não encontrado"}, 404

    if 'fornecedor_id' in data:
        fornecedor = _get_fornecedor_por_id(data.get('fornecedor_id'), restaurante_id)
        if not fornecedor:
            return {"error": "Fornecedor não encontrado."}, 404
        item.fornecedor_id = fornecedor.id

    for key, value in data.items():
        if key == 'fornecedor_id':
            continue
        if hasattr(item, key):
            setattr(item, key, value)

    log_event(
        acao="update",
        entidade="item",
        entidade_id=item.id,
        mensagem="Item atualizado",
        restaurante_id=item.fornecedor.restaurante_id if item.fornecedor else None
    )
    db.session.commit()
    return item.to_dict(), 200

def delete_item(item_id, restaurante_id=None):
    item, _ = get_item_by_id(item_id, restaurante_id)
    if not item:
        return {"error": "Item não encontrado"}, 404
    log_event(
        acao="delete",
        entidade="item",
        entidade_id=item.id,
        mensagem="Item deletado",
        restaurante_id=item.fornecedor.restaurante_id if item.fornecedor else None
    )
    db.session.delete(item)
    db.session.commit()
    return {}, 204


# ==============================
# SERVIÇO: Itens de Fornecedor
# ==============================

def create_item_fornecedor(fornecedor_id, data, restaurante_id=None):
    fornecedor = Fornecedor.query.filter_by(id=fornecedor_id).first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404
    if restaurante_id is not None and fornecedor.restaurante_id != restaurante_id:
        return {"error": "Você não tem permissão para acessar este fornecedor."}, 403

    nome = (data.get('nome') or '').strip()
    unidade_medida = (data.get('unidade_medida') or '').strip()
    if not nome or not unidade_medida:
        return {"error": "Nome e unidade de medida são obrigatórios."}, 400

    item_existente = Item.query.filter(
        Item.fornecedor_id == fornecedor_id,
        func.lower(Item.nome) == func.lower(nome)
    ).first()
    if item_existente:
        return {"error": f"Item '{nome}' já existe no catálogo deste fornecedor."}, 409

    new_item = Item(
        nome=nome,
        unidade_medida=unidade_medida,
        fornecedor_id=fornecedor_id
    )
    db.session.add(new_item)
    try:
        db.session.flush()
        log_event(
            acao="create",
            entidade="item",
            entidade_id=new_item.id,
            mensagem="Item criado via fornecedor",
            restaurante_id=fornecedor.restaurante_id
        )
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": f"Item '{nome}' já existe no sistema."}, 409

    return {
        "id": new_item.id,
        "nome": new_item.nome,
        "unidade_medida": new_item.unidade_medida,
        "fornecedor_id": new_item.fornecedor_id
    }, 201


def get_itens_by_fornecedor(fornecedor_id, restaurante_id=None):
    fornecedor = Fornecedor.query.filter_by(id=fornecedor_id).first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404
    if restaurante_id is not None and fornecedor.restaurante_id != restaurante_id:
        return {"error": "Você não tem permissão para acessar este fornecedor."}, 403

    itens = Item.query.filter_by(fornecedor_id=fornecedor_id).order_by(Item.nome).all()
    return [
        {
            "id": item.id,
            "nome": item.nome,
            "unidade_medida": item.unidade_medida,
            "fornecedor_id": item.fornecedor_id
        }
        for item in itens
    ], 200


def update_item_fornecedor(item_id, data, restaurante_id=None):
    item = db.session.query(Item).join(Fornecedor).filter(Item.id == item_id).first()
    if not item:
        return {"error": "Item não encontrado."}, 404
    if restaurante_id is not None and item.fornecedor.restaurante_id != restaurante_id:
        return {"error": "Você não tem permissão para editar este item."}, 403

    if 'nome' in data and data['nome']:
        nome = data['nome'].strip()
        duplicado = Item.query.filter(
            Item.fornecedor_id == item.fornecedor_id,
            func.lower(Item.nome) == func.lower(nome),
            Item.id != item_id
        ).first()
        if duplicado:
            return {"error": f"Já existe um item com o nome '{nome}' neste fornecedor."}, 409
        item.nome = nome

    if 'unidade_medida' in data and data['unidade_medida']:
        item.unidade_medida = data['unidade_medida'].strip()

    try:
        log_event(
            acao="update",
            entidade="item",
            entidade_id=item.id,
            mensagem="Item atualizado via fornecedor",
            restaurante_id=item.fornecedor.restaurante_id if item.fornecedor else None
        )
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Já existe um item com este nome no sistema."}, 409

    return {
        "id": item.id,
        "nome": item.nome,
        "unidade_medida": item.unidade_medida,
        "fornecedor_id": item.fornecedor_id
    }, 200


def delete_item_fornecedor(item_id, restaurante_id=None):
    item = db.session.query(Item).join(Fornecedor).filter(Item.id == item_id).first()
    if not item:
        return {"error": "Item não encontrado."}, 404
    if restaurante_id is not None and item.fornecedor.restaurante_id != restaurante_id:
        return {"error": "Você não tem permissão para deletar este item."}, 403

    em_uso_estoque = Estoque.query.filter_by(item_id=item_id).count() > 0
    em_uso_cotacao = CotacaoItem.query.filter_by(item_id=item_id).count() > 0

    if em_uso_estoque or em_uso_cotacao:
        return {
            "error": "Não é possível deletar este item pois ele está sendo usado em estoques ou cotações."
        }, 409

    db.session.delete(item)
    try:
        log_event(
            acao="delete",
            entidade="item",
            entidade_id=item.id,
            mensagem="Item deletado via fornecedor",
            restaurante_id=item.fornecedor.restaurante_id if item.fornecedor else None
        )
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Não foi possível deletar o item."}, 409

    return {"message": "Item deletado com sucesso."}, 200


def _normalize_unidade_csv(value):
    if value is None:
        return 'un'
    raw = value.strip()
    if not raw:
        return 'un'
    upper = raw.upper()
    mapping = {
        'KG': 'kg',
        'G': 'g',
        'L': 'L',
        'ML': 'mL',
        'UN': 'unidade',
        'UND': 'unidade',
        'UNID': 'unidade',
        'UNIDADE': 'unidade',
        'CX': 'caixa',
        'CAIXA': 'caixa',
        'PCT': 'pacote',
        'PACOTE': 'pacote',
        'FD': 'fardo',
        'FARDO': 'fardo',
        'SC': 'saco',
        'SACO': 'saco',
    }
    return mapping.get(upper, raw)


def importar_itens_fornecedor_csv(fornecedor_id, csv_content, restaurante_id=None):
    fornecedor = Fornecedor.query.filter_by(id=fornecedor_id).first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404
    if restaurante_id is not None and fornecedor.restaurante_id != restaurante_id:
        return {"error": "Você não tem permissão para acessar este fornecedor."}, 403

    if not csv_content or not csv_content.strip():
        return {"error": "Conteúdo CSV vazio."}, 400

    reader = csv.reader(StringIO(csv_content), delimiter=';')
    itens_criados = 0
    itens_atualizados = 0
    codigos_criados = 0
    codigos_atualizados = 0
    linhas_ignoradas = 0
    linhas_sem_codigo = 0
    erros = []

    for idx, row in enumerate(reader, start=1):
        if not row:
            linhas_ignoradas += 1
            continue

        codigo = row[0].strip() if len(row) > 0 else ''
        descricao = row[1].strip() if len(row) > 1 else ''
        unidade = row[2].strip() if len(row) > 2 else ''

        if not descricao:
            linhas_ignoradas += 1
            continue

        if codigo.lower() == 'codigo' or descricao.lower() == 'descricao':
            continue

        unidade_normalizada = _normalize_unidade_csv(unidade)

        item = Item.query.filter(func.lower(Item.nome) == func.lower(descricao)).first()
        if item:
            if item.fornecedor_id != fornecedor_id:
                erros.append({
                    "linha": idx,
                    "descricao": descricao,
                    "error": "Item já pertence a outro fornecedor."
                })
                continue
            if unidade_normalizada and item.unidade_medida != unidade_normalizada:
                item.unidade_medida = unidade_normalizada
                itens_atualizados += 1
        else:
            item = Item(
                nome=descricao,
                unidade_medida=unidade_normalizada,
                fornecedor_id=fornecedor_id
            )
            db.session.add(item)
            db.session.flush()
            itens_criados += 1

        if not codigo:
            linhas_sem_codigo += 1
            continue

        codigo_existente = FornecedorItemCodigo.query.filter_by(
            fornecedor_id=fornecedor_id,
            item_id=item.id
        ).first()

        if codigo_existente:
            if codigo_existente.codigo != codigo:
                codigo_conflito = FornecedorItemCodigo.query.filter_by(
                    fornecedor_id=fornecedor_id,
                    codigo=codigo
                ).first()
                if codigo_conflito and codigo_conflito.item_id != item.id:
                    erros.append({
                        "linha": idx,
                        "descricao": descricao,
                        "error": f"Código {codigo} já está associado a outro item."
                    })
                    continue
                codigo_existente.codigo = codigo
                codigos_atualizados += 1
        else:
            codigo_conflito = FornecedorItemCodigo.query.filter_by(
                fornecedor_id=fornecedor_id,
                codigo=codigo
            ).first()
            if codigo_conflito and codigo_conflito.item_id != item.id:
                erros.append({
                    "linha": idx,
                    "descricao": descricao,
                    "error": f"Código {codigo} já está associado a outro item."
                })
                continue
            novo_codigo = FornecedorItemCodigo(
                fornecedor_id=fornecedor_id,
                item_id=item.id,
                codigo=codigo
            )
            db.session.add(novo_codigo)
            codigos_criados += 1

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Erro ao importar itens. Verifique duplicidades."}, 409

    return {
        "message": "Importação concluída.",
        "itens_criados": itens_criados,
        "itens_atualizados": itens_atualizados,
        "codigos_criados": codigos_criados,
        "codigos_atualizados": codigos_atualizados,
        "linhas_ignoradas": linhas_ignoradas,
        "linhas_sem_codigo": linhas_sem_codigo,
        "erros": erros
    }, 200


def importar_itens_fornecedor_texto(fornecedor_id, texto, restaurante_id=None):
    fornecedor = Fornecedor.query.filter_by(id=fornecedor_id).first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404
    if restaurante_id is not None and fornecedor.restaurante_id != restaurante_id:
        return {"error": "Você não tem permissão para acessar este fornecedor."}, 403

    if not texto or not texto.strip():
        return {"error": "Conteúdo de texto vazio."}, 400

    itens_criados = 0
    itens_ignorados = 0
    erros = []

    for idx, linha in enumerate(texto.splitlines(), start=1):
        nome = linha.strip()
        if not nome:
            itens_ignorados += 1
            continue

        if normalize_item_nome(nome) == 'produto':
            continue

        item = Item.query.filter(func.lower(Item.nome) == func.lower(nome)).first()
        if item:
            if item.fornecedor_id != fornecedor_id:
                erros.append({
                    "linha": idx,
                    "descricao": nome,
                    "error": "Item já pertence a outro fornecedor."
                })
            else:
                itens_ignorados += 1
            continue

        novo_item = Item(
            nome=nome,
            unidade_medida='un',
            fornecedor_id=fornecedor_id
        )
        db.session.add(novo_item)
        itens_criados += 1

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Erro ao importar itens. Verifique duplicidades."}, 409

    return {
        "message": "Importação concluída.",
        "itens_criados": itens_criados,
        "itens_ignorados": itens_ignorados,
        "erros": erros
    }, 200


def create_area(data):
    new_area = Area(nome=data['nome'])
    repositories.add_instance(new_area)
    return {"id": new_area.id, "nome": new_area.nome}, 201

def get_all_areas():
    return repositories.get_all(Area), 200

def get_area_by_id(area_id):
    return repositories.get_by_id(Area, area_id), 200

def update_area(area_id, data):
    updated_area = repositories.update_instance(Area, area_id, data)
    if not updated_area:
        return {"error": "Área não encontrada"}, 404
    return updated_area.to_dict(), 200

def delete_area(area_id):

    if not repositories.delete_instance(Area, area_id):

        return {"error": "Área não encontrada"}, 404

    return {}, 204


def assign_colaboradores_to_area(area_id, data, restaurante_id=None):
    """Atribui colaboradores a uma área (substitui a lista atual)."""
    area = repositories.get_by_id(Area, area_id)
    if not area:
        return {"error": "Área não encontrada."}, 404

    colaborador_ids = data.get('colaborador_ids', [])
    if not colaborador_ids:
        area.colaboradores = []
    else:
        colaboradores_novos = []
        for user_id in colaborador_ids:
            user = repositories.get_by_id(Usuario, user_id)
            if (
                user
                and user.role == UserRoles.COLLABORATOR
                and (restaurante_id is None or user.restaurante_id == restaurante_id)
            ):
                colaboradores_novos.append(user)
        area.colaboradores = colaboradores_novos

    db.session.commit()

    return {
        "id": area.id,
        "nome": area.nome,
        "colaboradores": [{"id": c.id, "nome": c.nome, "email": c.email} for c in area.colaboradores]
    }, 200


def get_area_colaboradores(area_id, restaurante_id=None):
    """Retorna os colaboradores de uma área."""
    area = repositories.get_by_id(Area, area_id)
    if not area:
        return {"error": "Área não encontrada."}, 404

    return {
        "id": area.id,
        "nome": area.nome,
        "colaboradores": [{"id": c.id, "nome": c.nome, "email": c.email} for c in area.colaboradores]
    }, 200


def get_listas_da_area(area_id, restaurante_id=None):
    """Retorna todas as listas não deletadas vinculadas a uma área."""
    area = repositories.get_by_id(Area, area_id)
    if not area:
        return {"error": "Área não encontrada."}, 404

    query = Lista.query.filter_by(area_id=area_id, deletado=False)
    if restaurante_id:
        query = query.filter_by(restaurante_id=restaurante_id)
    listas = query.all()

    return [{"id": l.id, "nome": l.nome, "area_id": l.area_id} for l in listas], 200


def assign_listas_to_area(area_id, data, restaurante_id=None):
    """Vincula um conjunto de listas a uma área (replace: desmarca as anteriores)."""
    area = repositories.get_by_id(Area, area_id)
    if not area:
        return {"error": "Área não encontrada."}, 404

    # Desvincula todas as listas que atualmente pertencem a esta área
    query_atuais = Lista.query.filter_by(area_id=area_id, deletado=False)
    if restaurante_id:
        query_atuais = query_atuais.filter_by(restaurante_id=restaurante_id)
    for lista in query_atuais.all():
        lista.area_id = None

    # Vincula as novas listas
    lista_ids = data.get('lista_ids', [])
    for lid in lista_ids:
        lista = _get_lista_por_restaurante(lid, restaurante_id)
        if lista:
            lista.area_id = area_id

    db.session.commit()

    listas_atualizadas = Lista.query.filter_by(area_id=area_id, deletado=False).all()
    return {
        "id": area.id,
        "nome": area.nome,
        "listas": [{"id": l.id, "nome": l.nome} for l in listas_atualizadas]
    }, 200


def get_area_status(area_id):

    """Verifica se uma área possui itens com quantidade atual abaixo da mínima."""

    has_pending_items = Estoque.query.filter_by(area_id=area_id).filter(Estoque.quantidade_atual < Estoque.quantidade_minima).first() is not None

    return {"has_pending_items": has_pending_items}, 200



def create_fornecedor(data, restaurante_id=None):
    # Extrai lista_ids do data se existir
    lista_ids = data.pop('lista_ids', [])
    target_restaurante_id = restaurante_id or data.get('restaurante_id')
    if not target_restaurante_id:
        return {"error": "Restaurante é obrigatório."}, 400

    restaurante = Restaurante.query.filter_by(id=target_restaurante_id, deletado=False).first()
    if not restaurante:
        return {"error": "Restaurante não encontrado."}, 404

    # Cria novo fornecedor com campos básicos
    new_fornecedor = Fornecedor(
        nome=data['nome'],
        contato=data.get('contato'),
        meio_envio=data.get('meio_envio'),
        responsavel=data.get('responsavel'),
        observacao=data.get('observacao'),
        compartilhado_regiao=data.get('compartilhado_regiao', False),
        restaurante_id=restaurante.id
    )
    repositories.add_instance(new_fornecedor)
    log_event(
        acao="create",
        entidade="fornecedor",
        entidade_id=new_fornecedor.id,
        mensagem="Fornecedor criado",
        restaurante_id=new_fornecedor.restaurante_id
    )
    db.session.commit()

    # Adiciona as listas ao fornecedor se houver
    if lista_ids:
        from .models import Lista
        for lista_id in lista_ids:
            lista = repositories.get_by_id(Lista, lista_id)
            if lista:
                new_fornecedor.listas.append(lista)
        repositories.add_instance(new_fornecedor)

    return {"id": new_fornecedor.id, "nome": new_fornecedor.nome}, 201

def get_all_fornecedores(restaurante_id=None, somente_compartilhados=False):
    query = Fornecedor.query
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    if somente_compartilhados:
        query = query.filter_by(compartilhado_regiao=True)
    return query.all(), 200

def get_fornecedor_by_id(fornecedor_id, restaurante_id=None):
    query = Fornecedor.query.filter_by(id=fornecedor_id)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    return query.first(), 200

def get_fornecedor_detalhes_estoque(fornecedor_id, restaurante_id=None):
    query = Fornecedor.query.filter_by(id=fornecedor_id)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    fornecedor = query.first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado"}, 404

    itens = Item.query.filter_by(fornecedor_id=fornecedor_id).all()

    fornecedor_dict = fornecedor.to_dict()
    fornecedor_dict['listas'] = [
        {'id': lista.id, 'nome': lista.nome}
        for lista in fornecedor.listas
    ]

    return {
        "fornecedor": fornecedor_dict,
        "itens": [item.to_dict() for item in itens],
        "estoques": []
    }, 200

def update_fornecedor(fornecedor_id, data, restaurante_id=None):
    from .extensions import db
    from .models import Lista

    try:
        # Extrai lista_ids do data se existir
        lista_ids = data.pop('lista_ids', None)
        allowed_fields = {
            'nome',
            'telefone',
            'cnpj',
            'endereco',
            'cidade',
            'estado',
            'cep',
            'site'
        }

        # Busca o fornecedor
        query = Fornecedor.query.filter_by(id=fornecedor_id)
        if restaurante_id is not None:
            query = query.filter_by(restaurante_id=restaurante_id)
        updated_fornecedor = query.first()
        if not updated_fornecedor:
            return {"error": "Fornecedor não encontrado"}, 404

        # Atualiza campos do fornecedor
        for key, value in data.items():
            if key not in allowed_fields:
                continue
            if key == 'estado' and value:
                estado_norm = str(value).strip().upper()
                if len(estado_norm) != 2:
                    return {"error": "Estado deve ter 2 caracteres (UF)."}, 400
                value = estado_norm
            setattr(updated_fornecedor, key, value)

        # Atualiza as listas se foi fornecido lista_ids
        if lista_ids is not None:
            # Limpa listas anteriores e adiciona novas
            updated_fornecedor.listas = []
            db.session.flush()
            
            for lista_id in lista_ids:
                lista = repositories.get_by_id(Lista, lista_id)
                if lista:
                    updated_fornecedor.listas.append(lista)

        # Commit das mudanças
        log_event(
            acao="update",
            entidade="fornecedor",
            entidade_id=updated_fornecedor.id,
            mensagem="Fornecedor atualizado",
            restaurante_id=updated_fornecedor.restaurante_id
        )
        db.session.commit()

        return updated_fornecedor.to_dict(), 200
    
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao atualizar fornecedor {fornecedor_id}: {str(e)}")
        return {"error": f"Erro ao atualizar fornecedor: {str(e)}"}, 500

def delete_fornecedor(fornecedor_id, restaurante_id=None):
    query = Fornecedor.query.filter_by(id=fornecedor_id)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    fornecedor = query.first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado"}, 404
    log_event(
        acao="delete",
        entidade="fornecedor",
        entidade_id=fornecedor.id,
        mensagem="Fornecedor deletado",
        restaurante_id=fornecedor.restaurante_id
    )
    db.session.delete(fornecedor)
    db.session.commit()
    return {}, 204


def get_pedidos_fornecedor_por_lista(fornecedor_id, restaurante_id=None):
    """
    Retorna os pedidos de um fornecedor agrupados por lista.
    Formato: {lista_id: {lista_nome: str, pedidos: [...]}}
    """
    fornecedor = Fornecedor.query.filter_by(id=fornecedor_id)
    if restaurante_id is not None:
        fornecedor = fornecedor.filter_by(restaurante_id=restaurante_id)
    fornecedor = fornecedor.first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado"}, 404

    # Agrupa pedidos por lista
    resultado = {}

    # Para cada lista atribuída ao fornecedor
    for lista in fornecedor.listas:
        # Busca itens que pertencem a essa lista
        itens_lista = Item.query.filter_by(fornecedor_id=fornecedor_id).all()

        pedidos_lista = []
        for item in itens_lista:
            # Busca pedidos pendentes para esse item
            pedidos_item = Pedido.query.filter_by(
                item_id=item.id,
                fornecedor_id=fornecedor_id,
                status='PENDENTE'
            ).all()

            for pedido in pedidos_item:
                pedidos_lista.append({
                    'item_nome': item.nome,
                    'quantidade': float(pedido.quantidade_solicitada),
                    'unidade': item.unidade_medida,
                    'data_pedido': pedido.data_pedido.isoformat(),
                    'usuario': pedido.usuario.nome if pedido.usuario else 'N/A'
                })

        if pedidos_lista:
            resultado[lista.id] = {
                'lista_nome': lista.nome,
                'pedidos': pedidos_lista,
                'total_itens': len(pedidos_lista)
            }

    return resultado, 200


def get_pedidos_fornecedor_consolidado(fornecedor_id, restaurante_id=None):
    """
    Retorna todos os pedidos de um fornecedor consolidados (sem separação por lista).
    """
    fornecedor = Fornecedor.query.filter_by(id=fornecedor_id)
    if restaurante_id is not None:
        fornecedor = fornecedor.filter_by(restaurante_id=restaurante_id)
    fornecedor = fornecedor.first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado"}, 404

    # Busca todos os itens desse fornecedor
    itens_fornecedor = Item.query.filter_by(fornecedor_id=fornecedor_id).all()

    pedidos_consolidados = []
    for item in itens_fornecedor:
        # Busca todos os pedidos pendentes para esse item
        pedidos_item = Pedido.query.filter_by(
            item_id=item.id,
            fornecedor_id=fornecedor_id,
            status='PENDENTE'
        ).all()

        for pedido in pedidos_item:
            pedidos_consolidados.append({
                'item_nome': item.nome,
                'quantidade': float(pedido.quantidade_solicitada),
                'unidade': item.unidade_medida,
                'data_pedido': pedido.data_pedido.isoformat(),
                'usuario': pedido.usuario.nome if pedido.usuario else 'N/A'
            })

    return {
        'fornecedor_nome': fornecedor.nome,
        'fornecedor_contato': fornecedor.contato,
        'fornecedor_meio_envio': fornecedor.meio_envio,
        'total_pedidos': len(pedidos_consolidados),
        'pedidos': pedidos_consolidados
    }, 200


# --- Serviços de Cotação ---

def create_quotation_from_stock(fornecedor_id, restaurante_id=None):
    """Cria uma nova Cotação com base na necessidade de estoque para um fornecedor."""
    fornecedor = Fornecedor.query.filter_by(id=fornecedor_id)
    if restaurante_id is not None:
        fornecedor = fornecedor.filter_by(restaurante_id=restaurante_id)
    fornecedor = fornecedor.first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado"}, 404

    pedidos = {}
    for item in fornecedor.itens:
        total_a_pedir = 0
        estoques_do_item = Estoque.query.filter_by(item_id=item.id).all()
        for estoque in estoques_do_item:
            if estoque.quantidade_atual < estoque.quantidade_minima:
                total_a_pedir += (estoque.quantidade_minima - estoque.quantidade_atual)
        
        if total_a_pedir > 0:
            pedidos[item] = total_a_pedir

    if not pedidos:
        return {"message": f"Nenhum item precisa de reposição para o fornecedor {fornecedor.nome}."}, 200

    # Cria a Cotação
    nova_cotacao = Cotacao(fornecedor_id=fornecedor_id)
    repositories.add_instance(nova_cotacao)

    # Adiciona os itens na cotação
    for item, quantidade in pedidos.items():
        cotacao_item = CotacaoItem(
            cotacao_id=nova_cotacao.id,
            item_id=item.id,
            quantidade=quantidade,
            preco_unitario=0  # Preço a ser preenchido pelo admin
        )
        db.session.add(cotacao_item)
    
    db.session.commit()

    return nova_cotacao.to_dict(), 201


def get_all_cotacoes():
    """Retorna todas as cotações."""
    return repositories.get_all(Cotacao), 200

def get_cotacao_by_id(cotacao_id):
    """Retorna uma cotação específica com seus itens."""
    cotacao = repositories.get_by_id(Cotacao, cotacao_id)
    if not cotacao:
        return {"error": "Cotação não encontrada"}, 404
    # Aqui precisamos de uma serialização mais inteligente que inclua os itens
    # Vou modificar o to_dict da Cotação no models.py depois
    return cotacao, 200

def update_cotacao_item_price(item_id, data):
    """Atualiza o preço de um item em uma cotação."""
    cotacao_item = repositories.get_by_id(CotacaoItem, item_id)
    if not cotacao_item:
        return {"error": "Item de cotação não encontrado"}, 404
    
    if 'preco_unitario' in data:
        updated_item = repositories.update_instance(CotacaoItem, item_id, {'preco_unitario': data['preco_unitario']})
        return updated_item.to_dict(), 200
    return {"error": "Preço unitário não fornecido"}, 400


def get_estoque_by_area(area_id):
    estoque_list = Estoque.query.filter_by(area_id=area_id).all()
    return estoque_list, 200

def update_estoque_item(estoque_id, data):
    estoque_item = repositories.get_by_id(Estoque, estoque_id)
    if not estoque_item:
        return {"error": "Item de estoque não encontrado."}, 404

    if not data:
        return {"error": "Dados inválidos."}, 400

    last_known_update = data.get("last_known_update")
    if last_known_update:
        client_dt = _parse_iso_datetime(last_known_update)
        if client_dt is None:
            return {"error": "Timestamp invalido."}, 400
        server_dt = _normalize_datetime(estoque_item.atualizado_em)
        client_dt = _normalize_datetime(client_dt)
        if server_dt and client_dt and server_dt > client_dt:
            return {
                "error": "Conflito de versao",
                "server_data": estoque_item.to_dict()
            }, 409
    
    # Apenas a quantidade_atual pode ser atualizada por esta rota
    if 'quantidade_atual' in data:
        estoque_item.quantidade_atual = data['quantidade_atual']
        estoque_item.atualizado_em = brasilia_now()
        db.session.commit()
        return {
            "message": "Estoque atualizado",
            "estoque": estoque_item.to_dict()
        }, 200
    return {"error": "Dados inválidos. Forneça quantidade_atual."}, 400

def save_estoque_draft(data):
    """Salva um rascunho do estoque de uma área."""
    area_id = data.get('area_id')
    items_data = data.get('items', [])

    if not area_id:
        return {"error": "ID da área não fornecido."}, 400

    for item_data in items_data:
        estoque_id = item_data.get('id')
        quantidade_atual = item_data.get('quantidade_atual')

        if estoque_id is None or quantidade_atual is None:
            continue # Pula itens mal formatados

        estoque_item = repositories.get_by_id(Estoque, estoque_id)
        if estoque_item and estoque_item.area_id == area_id:
            estoque_item.quantidade_atual = quantidade_atual
            db.session.add(estoque_item)
    
    db.session.commit()
    return {"message": "Rascunho do estoque salvo com sucesso!"}, 200

def get_pedidos_by_user(user_id):
    """Retorna todos os pedidos feitos por um usuário."""
    pedidos = Pedido.query.filter_by(usuario_id=user_id).order_by(Pedido.data_pedido.desc()).all()
    return pedidos, 200


def get_submissoes_by_user(user_id):
    """
    Retorna todas as submissões do usuário com pedidos agrupados.
    Formato otimizado para exibição em tela.
    """
    submissoes = Submissao.query.options(
        db.joinedload(Submissao.lista),
        db.joinedload(Submissao.pedidos).joinedload(Pedido.item)
    ).filter_by(usuario_id=user_id).order_by(Submissao.data_submissao.desc()).all()
    
    resultado = []
    for sub in submissoes:
        sub_dict = {
            "id": sub.id,
            "lista_id": sub.lista_id,
            "lista_nome": sub.lista.nome if sub.lista else "N/A",
            "data_submissao": sub.data_submissao.isoformat(),
            "status": sub.status.value,
            "total_pedidos": sub.total_pedidos,
            "pedidos": [
                {
                    "id": p.id,
                    "item_nome": p.item.nome if p.item else "N/A",
                    "quantidade_solicitada": float(p.quantidade_solicitada),
                    "status": p.status.value,
                    "unidade": p.item.unidade if p.item else ""
                }
                for p in sub.pedidos
            ]
        }
        resultado.append(sub_dict)
    
    return resultado, 200


# ========================================
# SUPPLIER: Cadastro e Convites
# ========================================

def _validar_dados_cadastro_fornecedor(data):
    nome = (data.get('nome') or '').strip()
    email = (data.get('email') or '').strip().lower()
    senha = data.get('senha') or ''
    nome_empresa = (data.get('nome_empresa') or data.get('empresa') or '').strip()
    telefone = (data.get('telefone') or '').strip()

    if not nome or not email or not senha or not nome_empresa or not telefone:
        return None, {"error": "Nome, email, senha, nome da empresa e telefone são obrigatórios."}, 400

    if len(senha) < 6:
        return None, {"error": "Senha deve ter pelo menos 6 caracteres."}, 400

    if Usuario.query.filter_by(email=email).first():
        return None, {"error": "E-mail já cadastrado."}, 409

    cnpj = (data.get('cnpj') or '').strip()
    cnpj_numeros = ''.join([c for c in cnpj if c.isdigit()])
    if cnpj and len(cnpj_numeros) != 14:
        return None, {"error": "CNPJ inválido."}, 400

    payload = {
        "nome": nome,
        "email": email,
        "senha": senha,
        "nome_empresa": nome_empresa,
        "telefone": telefone,
        "cnpj": cnpj_numeros or None,
        "endereco": (data.get('endereco') or '').strip() or None,
        "cidade": (data.get('cidade') or '').strip() or None,
        "estado": (data.get('estado') or '').strip().upper() or None,
        "cep": (data.get('cep') or '').strip() or None,
        "site": (data.get('site') or '').strip() or None,
    }
    return payload, None, None


def criar_convite_fornecedor(admin_id, limite_usos=1):
    import os

    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode criar convites."}, 403

    if not isinstance(limite_usos, int) or limite_usos < 1 or limite_usos > 100:
        return {"error": "Limite de usos deve ser um número entre 1 e 100."}, 400

    token = uuid4().hex
    expira_em = brasilia_now() + timedelta(hours=72)
    convite = ConviteFornecedor(
        token=token,
        criado_por_id=admin_id,
        expira_em=expira_em,
        limite_usos=limite_usos,
        quantidade_usos=0
    )
    db.session.add(convite)
    db.session.commit()

    base_url = os.environ.get('FRONTEND_URL', 'https://kaizen-compras.up.railway.app')
    link_convite = f"{base_url}/supplier/register-convite?token={token}"

    return {
        **convite.to_dict(),
        "link": link_convite,
        "message": "Convite de fornecedor criado com sucesso!"
    }, 201


def listar_convites_fornecedor(admin_id):
    import os

    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode listar convites."}, 403

    convites = ConviteFornecedor.query.order_by(ConviteFornecedor.criado_em.desc()).all()
    base_url = os.environ.get('FRONTEND_URL', 'https://kaizen-compras.up.railway.app')
    return {
        "convites": [
            {
                **convite.to_dict(),
                "link": f"{base_url}/supplier/register-convite?token={convite.token}"
            }
            for convite in convites
        ],
        "total": len(convites)
    }, 200


def validar_convite_fornecedor(token):
    convite = ConviteFornecedor.query.filter_by(token=token).first()
    if not convite:
        return {"valido": False, "erro": "Token não encontrado"}, 400
    if not convite.esta_valido():
        if convite.quantidade_usos >= convite.limite_usos:
            return {"valido": False, "erro": "Token já foi utilizado"}, 400
        return {"valido": False, "erro": "Token expirado"}, 400
    return {
        "valido": True,
        "criado_em": convite.criado_em.strftime('%d/%m/%Y'),
        "expira_em": convite.expira_em.strftime('%d/%m/%Y %H:%M') if convite.expira_em else None
    }, 200


def excluir_convite_fornecedor(admin_id, convite_id):
    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode excluir convites."}, 403

    convite = ConviteFornecedor.query.get(convite_id)
    if not convite:
        return {"error": "Convite não encontrado."}, 404
    db.session.delete(convite)
    db.session.commit()
    return {"message": "Convite excluído com sucesso!"}, 200


def editar_convite_fornecedor(admin_id, convite_id, data):
    import os

    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode editar convites."}, 403

    convite = ConviteFornecedor.query.get(convite_id)
    if not convite:
        return {"error": "Convite não encontrado."}, 404

    if 'limite_usos' in data:
        novo_limite = data['limite_usos']
        try:
            novo_limite = int(novo_limite)
        except (ValueError, TypeError):
            return {"error": "Limite de usos deve ser um número."}, 400
        if novo_limite < 1 or novo_limite > 100:
            return {"error": "Limite de usos deve ser entre 1 e 100."}, 400
        if novo_limite < convite.quantidade_usos:
            return {"error": f"Limite não pode ser menor que a quantidade já usada ({convite.quantidade_usos})."}, 400
        convite.limite_usos = novo_limite

    if 'ativo' in data:
        convite.ativo = bool(data['ativo'])

    db.session.commit()

    base_url = os.environ.get('FRONTEND_URL', 'https://kaizen-compras.up.railway.app')
    return {
        "message": "Convite atualizado com sucesso!",
        "convite": {
            **convite.to_dict(),
            "link": f"{base_url}/supplier/register-convite?token={convite.token}"
        }
    }, 200


def register_fornecedor_com_convite(data):
    token = (data.get('token') or '').strip()
    convite = ConviteFornecedor.query.filter_by(token=token).first()
    if not convite or not convite.esta_valido():
        return {"error": "Convite inválido ou expirado."}, 400

    payload, error, status = _validar_dados_cadastro_fornecedor(data)
    if error:
        return error, status

    restaurante_id = _get_restaurante_regional_id()
    if not restaurante_id:
        return {"error": "Restaurante regional não configurado."}, 500

    senha_hash = generate_password_hash(payload["senha"])
    session_token = str(uuid4())
    usuario = Usuario(
        nome=payload["nome"],
        email=payload["email"],
        senha_hash=senha_hash,
        role=UserRoles.SUPPLIER,
        aprovado=True,
        ativo=True,
        restaurante_id=restaurante_id,
        session_token=session_token,
        session_updated_at=brasilia_now()
    )
    db.session.add(usuario)
    db.session.flush()

    fornecedor = Fornecedor(
        nome=payload["nome_empresa"],
        usuario_id=usuario.id,
        cnpj=payload["cnpj"],
        telefone=payload["telefone"],
        endereco=payload["endereco"],
        cidade=payload["cidade"],
        estado=payload["estado"],
        cep=payload["cep"],
        site=payload["site"],
        aprovado=True,
        restaurante_id=restaurante_id,
        compartilhado_regiao=True
    )
    db.session.add(fornecedor)
    db.session.flush()

    convite.quantidade_usos += 1
    if convite.quantidade_usos >= convite.limite_usos:
        convite.usado = True
    convite.usado_em = brasilia_now()
    convite.usado_por_id = usuario.id
    convite.fornecedor_id = fornecedor.id

    log_event(
        acao="create",
        entidade="fornecedor",
        entidade_id=fornecedor.id,
        mensagem="Fornecedor criado via convite",
        restaurante_id=restaurante_id,
        usuario_id=usuario.id,
        metadata={
            "usuario_id": usuario.id,
            "convite_id": convite.id
        }
    )
    db.session.commit()

    additional_claims = {
        "role": usuario.role.value,
        "nome": usuario.nome,
        "email": usuario.email,
        "restaurante_id": usuario.restaurante_id,
        "wizard_status": usuario.wizard_status or {},
        "session_token": usuario.session_token,
    }
    access_token = create_access_token(
        identity=str(usuario.id),
        additional_claims=additional_claims,
        expires_delta=timedelta(days=1)
    )
    return {"access_token": access_token}, 201


def register_fornecedor_auto(data):
    payload, error, status = _validar_dados_cadastro_fornecedor(data)
    if error:
        return error, status

    restaurante_id = _get_restaurante_regional_id()
    if not restaurante_id:
        return {"error": "Restaurante regional não configurado."}, 500

    senha_hash = generate_password_hash(payload["senha"])
    usuario = Usuario(
        nome=payload["nome"],
        email=payload["email"],
        senha_hash=senha_hash,
        role=UserRoles.SUPPLIER,
        aprovado=False,
        ativo=True,
        restaurante_id=restaurante_id
    )
    db.session.add(usuario)
    db.session.flush()

    fornecedor = Fornecedor(
        nome=payload["nome_empresa"],
        usuario_id=usuario.id,
        cnpj=payload["cnpj"],
        telefone=payload["telefone"],
        endereco=payload["endereco"],
        cidade=payload["cidade"],
        estado=payload["estado"],
        cep=payload["cep"],
        site=payload["site"],
        aprovado=False,
        restaurante_id=restaurante_id,
        compartilhado_regiao=False
    )
    db.session.add(fornecedor)
    db.session.flush()

    super_admins = Usuario.query.filter_by(role=UserRoles.SUPER_ADMIN, ativo=True).all()
    for admin in super_admins:
        criar_notificacao(
            admin.id,
            "Novo fornecedor aguardando aprovação",
            TipoNotificacao.SOLICITACAO_RESTAURANTE,
            mensagem=f"Fornecedor {fornecedor.nome} aguardando aprovação.",
            restaurante_id=None
        )

    log_event(
        acao="create",
        entidade="fornecedor",
        entidade_id=fornecedor.id,
        mensagem="Fornecedor criado aguardando aprovação",
        restaurante_id=restaurante_id,
        usuario_id=usuario.id,
        metadata={"usuario_id": usuario.id}
    )
    db.session.commit()
    return {"message": "Cadastro realizado. Aguarde aprovação."}, 201


# ========================================
# SUPPLIER: Perfil e Itens
# ========================================

def _obter_fornecedor_por_usuario(user_id):
    return Fornecedor.query.filter_by(usuario_id=user_id).first()


def obter_perfil_fornecedor(user_id):
    fornecedor = _obter_fornecedor_por_usuario(user_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404

    usuario = db.session.get(Usuario, user_id)
    perfil = fornecedor.to_dict()
    perfil.update({
        "usuario": {
            "id": usuario.id,
            "nome": usuario.nome,
            "email": usuario.email,
            "aprovado": usuario.aprovado
        }
    })
    return perfil, 200


def atualizar_perfil_fornecedor(user_id, data):
    fornecedor = _obter_fornecedor_por_usuario(user_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404

    usuario = db.session.get(Usuario, user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    mudancas = {}
    if "email" in data and data["email"] != usuario.email:
        mudancas["email"] = {"anterior": usuario.email, "novo": data["email"]}
    if "email" in data and data["email"] != usuario.email:
        if Usuario.query.filter_by(email=data["email"]).first():
            return {"error": "E-mail já cadastrado."}, 409
        usuario.email = data["email"]
    if "nome" in data and data["nome"]:
        mudancas["nome"] = {"anterior": usuario.nome, "novo": data["nome"]}
        usuario.nome = data["nome"]

    if "nome_empresa" in data:
        mudancas["nome_empresa"] = {"anterior": fornecedor.nome, "novo": data["nome_empresa"]}
        fornecedor.nome = data["nome_empresa"] or None
    elif "nome" in data and "nome_empresa" not in data:
        mudancas["nome_empresa"] = {"anterior": fornecedor.nome, "novo": data["nome"]}
        fornecedor.nome = data["nome"] or None

    for field in [
        "cnpj", "telefone", "endereco", "cidade",
        "estado", "cep", "site", "contato", "meio_envio",
        "responsavel", "observacao"
    ]:
        if field in data:
            mudancas[field] = {"anterior": getattr(fornecedor, field), "novo": data[field]}
            setattr(fornecedor, field, data[field] or None)

    log_event(
        acao="update",
        entidade="fornecedor",
        entidade_id=fornecedor.id,
        mensagem="Perfil do fornecedor atualizado",
        restaurante_id=fornecedor.restaurante_id,
        usuario_id=user_id,
        metadata={"mudancas": mudancas}
    )
    db.session.commit()
    return fornecedor.to_dict(), 200


def _validar_dados_item(data, fornecedor_id, item_id=None):
    nome = (data.get('nome') or '').strip()
    unidade = (data.get('unidade_medida') or '').strip()
    if not nome or not unidade:
        return None, {"error": "Nome e unidade de medida são obrigatórios."}, 400

    duplicado = Item.query.filter(func.lower(Item.nome) == func.lower(nome)).first()
    if duplicado and duplicado.id != item_id:
        return None, {"error": f"Item '{nome}' já existe no sistema."}, 409

    codigo = (data.get('codigo_fornecedor') or '').strip()
    if codigo:
        codigo_duplicado = Item.query.filter(
            Item.fornecedor_id == fornecedor_id,
            Item.codigo_fornecedor == codigo,
            Item.id != item_id
        ).first()
        if codigo_duplicado:
            return None, {"error": "Código do fornecedor já utilizado."}, 409

    payload = {
        "nome": nome,
        "unidade_medida": unidade,
        "codigo_fornecedor": codigo or None,
        "descricao": (data.get('descricao') or '').strip() or None,
        "marca": (data.get('marca') or '').strip() or None,
        "preco_atual": data.get('preco_atual')
    }
    return payload, None, None


def _registrar_alteracao_preco(item_id, preco_anterior, preco_novo, user_id, observacao=None):
    historico = ItemPrecoHistorico(
        item_id=item_id,
        preco_anterior=preco_anterior,
        preco_novo=preco_novo,
        alterado_por_id=user_id,
        observacao=observacao
    )
    db.session.add(historico)


def listar_itens_fornecedor(user_id):
    fornecedor = _obter_fornecedor_por_usuario(user_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404

    itens = Item.query.filter_by(fornecedor_id=fornecedor.id).order_by(Item.nome).all()
    return [item.to_dict() for item in itens], 200


def criar_item_fornecedor(user_id, data):
    fornecedor = _obter_fornecedor_por_usuario(user_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404

    payload, error, status = _validar_dados_item(data, fornecedor.id)
    if error:
        return error, status

    preco = payload.get("preco_atual")
    if preco is not None:
        try:
            preco = float(preco)
            if preco < 0:
                return {"error": "Preço inválido."}, 400
        except (TypeError, ValueError):
            return {"error": "Preço inválido."}, 400

    item = Item(
        nome=payload["nome"],
        unidade_medida=payload["unidade_medida"],
        codigo_fornecedor=payload["codigo_fornecedor"],
        descricao=payload["descricao"],
        marca=payload["marca"],
        preco_atual=preco,
        fornecedor_id=fornecedor.id
    )
    db.session.add(item)
    db.session.flush()

    if preco is not None:
        _registrar_alteracao_preco(item.id, None, preco, user_id, "Cadastro inicial")

    log_event(
        acao="create",
        entidade="fornecedor",
        entidade_id=fornecedor.id,
        mensagem="Item criado pelo fornecedor",
        restaurante_id=fornecedor.restaurante_id,
        usuario_id=user_id,
        metadata={
            "item_id": item.id,
            "item_nome": item.nome
        }
    )
    db.session.commit()
    return item.to_dict(), 201


def _validar_propriedade_item(user_id, item_id):
    fornecedor = _obter_fornecedor_por_usuario(user_id)
    if not fornecedor:
        return None, {"error": "Fornecedor não encontrado."}, 404
    item = Item.query.filter_by(id=item_id, fornecedor_id=fornecedor.id).first()
    if not item:
        return None, {"error": "Item não pertence ao fornecedor."}, 403
    return item, None, None


def obter_item_fornecedor(user_id, item_id):
    item, error, status = _validar_propriedade_item(user_id, item_id)
    if error:
        return error, status
    return item.to_dict(), 200


def atualizar_item_fornecedor(user_id, item_id, data):
    item, error, status = _validar_propriedade_item(user_id, item_id)
    if error:
        return error, status

    nome_anterior = item.nome
    unidade_anterior = item.unidade_medida
    preco_anterior = float(item.preco_atual) if item.preco_atual is not None else None
    payload, error, status = _validar_dados_item(data, item.fornecedor_id, item_id=item.id)
    if error:
        return error, status

    preco_novo = payload.get("preco_atual")
    if preco_novo is not None:
        try:
            preco_novo = float(preco_novo)
            if preco_novo < 0:
                return {"error": "Preço inválido."}, 400
        except (TypeError, ValueError):
            return {"error": "Preço inválido."}, 400

    if preco_novo is not None and preco_novo != preco_anterior:
        _registrar_alteracao_preco(item.id, preco_anterior, preco_novo, user_id, "Atualização de preço")
        item.preco_atual = preco_novo

    item.nome = payload["nome"]
    item.unidade_medida = payload["unidade_medida"]
    item.codigo_fornecedor = payload["codigo_fornecedor"]
    item.descricao = payload["descricao"]
    item.marca = payload["marca"]

    log_event(
        acao="update",
        entidade="fornecedor",
        entidade_id=item.fornecedor_id,
        mensagem="Item atualizado pelo fornecedor",
        restaurante_id=item.fornecedor.restaurante_id if item.fornecedor else None,
        usuario_id=user_id,
        metadata={
            "item_id": item.id,
            "nome_anterior": nome_anterior,
            "unidade_anterior": unidade_anterior,
            "preco_anterior": preco_anterior,
            "nome": item.nome,
            "unidade": item.unidade_medida,
            "preco": item.preco_atual
        }
    )
    db.session.commit()
    return item.to_dict(), 200


def deletar_item_fornecedor(user_id, item_id):
    item, error, status = _validar_propriedade_item(user_id, item_id)
    if error:
        return error, status

    if Estoque.query.filter_by(item_id=item.id).first():
        return {"error": "Item em uso em estoques."}, 409
    if CotacaoItem.query.filter_by(item_id=item.id).first():
        return {"error": "Item em uso em cotações."}, 409
    log_event(
        acao="delete",
        entidade="fornecedor",
        entidade_id=item.fornecedor_id,
        mensagem="Item removido pelo fornecedor",
        restaurante_id=item.fornecedor.restaurante_id if item.fornecedor else None,
        usuario_id=user_id,
        metadata={
            "item_id": item.id,
            "item_nome": item.nome
        }
    )
    db.session.delete(item)
    db.session.commit()
    return {"message": "Item removido."}, 200


def obter_historico_precos_item(user_id, item_id):
    item, error, status = _validar_propriedade_item(user_id, item_id)
    if error:
        return error, status

    historico = ItemPrecoHistorico.query.filter_by(item_id=item.id).order_by(
        ItemPrecoHistorico.alterado_em.desc()
    ).all()
    return [h.to_dict() for h in historico], 200


# ========================================
# ADMIN: Fornecedores
# ========================================

def listar_fornecedores_cadastrados(admin_id, filtro_aprovado=None):
    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role not in (UserRoles.ADMIN, UserRoles.SUPER_ADMIN):
        return {"error": "Acesso negado."}, 403

    query = Fornecedor.query
    if filtro_aprovado is not None:
        query = query.filter_by(aprovado=filtro_aprovado)
    fornecedores = query.order_by(Fornecedor.criado_em.desc()).all()
    resultado = []
    for fornecedor in fornecedores:
        fornecedor_dict = fornecedor.to_dict()
        if fornecedor.usuario_id:
            usuario = db.session.get(Usuario, fornecedor.usuario_id)
            if usuario:
                fornecedor_dict['usuario_email'] = usuario.email
                fornecedor_dict['usuario_senha'] = usuario.senha_texto_puro or '(Definida no cadastro - não armazenada)'
                fornecedor_dict['usuario_id'] = usuario.id
        resultado.append(fornecedor_dict)
    return resultado, 200


def aprovar_fornecedor(admin_id, fornecedor_id, aprovado=True):
    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode aprovar fornecedores."}, 403

    fornecedor = Fornecedor.query.get(fornecedor_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404

    fornecedor.aprovado = bool(aprovado)
    fornecedor.compartilhado_regiao = bool(aprovado)
    if fornecedor.usuario_id:
        usuario = db.session.get(Usuario, fornecedor.usuario_id)
        if usuario:
            usuario.aprovado = bool(aprovado)
    log_event(
        acao="approve" if aprovado else "update",
        entidade="fornecedor",
        entidade_id=fornecedor.id,
        mensagem="Fornecedor aprovado" if aprovado else "Fornecedor reprovado",
        restaurante_id=fornecedor.restaurante_id,
        usuario_id=admin_id
    )
    db.session.commit()
    return fornecedor.to_dict(), 200


def obter_detalhes_fornecedor_admin(admin_id, fornecedor_id):
    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role not in (UserRoles.ADMIN, UserRoles.SUPER_ADMIN):
        return {"error": "Acesso negado."}, 403

    fornecedor = Fornecedor.query.get(fornecedor_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404
    return fornecedor.to_dict(), 200


def listar_itens_de_fornecedor_admin(admin_id, fornecedor_id):
    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role not in (UserRoles.ADMIN, UserRoles.SUPER_ADMIN):
        return {"error": "Acesso negado."}, 403

    itens = Item.query.filter_by(fornecedor_id=fornecedor_id).order_by(Item.nome).all()
    return [item.to_dict() for item in itens], 200


def listar_usuarios_fornecedor(admin_id, fornecedor_id):
    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role not in (UserRoles.ADMIN, UserRoles.SUPER_ADMIN):
        return {"error": "Acesso negado."}, 403

    fornecedor = Fornecedor.query.get(fornecedor_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404

    usuarios = []
    if fornecedor.usuario_id:
        usuario = db.session.get(Usuario, fornecedor.usuario_id)
        if usuario:
            usuarios.append(usuario.to_dict_with_password())

    return {"usuarios": usuarios}, 200


def criar_login_fornecedor(admin_id, fornecedor_id, data):
    admin = db.session.get(Usuario, admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode criar login de fornecedor."}, 403

    fornecedor = Fornecedor.query.get(fornecedor_id)
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404

    if fornecedor.usuario_id:
        return {"error": "Fornecedor já possui login associado."}, 409

    email = (data.get('email') or '').strip().lower()
    senha = (data.get('senha') or '').strip()
    nome = (data.get('nome') or fornecedor.nome or 'Fornecedor').strip()

    if not email or not senha:
        return {"error": "Email e senha são obrigatórios."}, 400

    email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    if not re.match(email_regex, email):
        return {"error": "E-mail inválido."}, 400

    if len(senha) < 6:
        return {"error": "Senha deve ter pelo menos 6 caracteres."}, 400

    if Usuario.query.filter_by(email=email).first():
        return {"error": "E-mail já cadastrado."}, 409

    restaurante_id = fornecedor.restaurante_id
    senha_hash = generate_password_hash(senha)
    usuario = Usuario(
        nome=nome,
        email=email,
        senha_hash=senha_hash,
        senha_texto_puro=senha,
        role=UserRoles.SUPPLIER,
        aprovado=True,
        ativo=True,
        restaurante_id=restaurante_id
    )
    db.session.add(usuario)
    db.session.flush()

    fornecedor.usuario_id = usuario.id
    fornecedor.aprovado = True
    fornecedor.compartilhado_regiao = True
    log_event(
        acao="create",
        entidade="fornecedor",
        entidade_id=fornecedor.id,
        mensagem="Login de fornecedor criado",
        restaurante_id=restaurante_id,
        usuario_id=admin_id,
        metadata={"usuario_id": usuario.id}
    )
    db.session.commit()

    return {
        "message": "Login de fornecedor criado com sucesso.",
        "usuario": usuario.to_dict_with_password()
    }, 201


def get_all_submissoes(status_filter=None, arquivadas=False):
    """
    Retorna todas as submissões (admin) com pedidos agrupados.
    Inclui tanto listas tradicionais quanto listas rápidas.

    Args:
        status_filter: PENDENTE, APROVADO, REJEITADO, PARCIALMENTE_APROVADO ou None
        arquivadas: quando True, retorna apenas submissões arquivadas
    """
    from .models import ListaRapida, StatusListaRapida, ListaRapidaItem

    resultado = []

    # 1. BUSCAR SUBMISSÕES TRADICIONAIS
    query = Submissao.query.options(
        db.joinedload(Submissao.lista),
        db.joinedload(Submissao.usuario),
        db.joinedload(Submissao.pedidos).joinedload(Pedido.item)
    )

    if arquivadas:
        query = query.filter_by(arquivada=True)
    else:
        query = query.filter_by(arquivada=False)

    if status_filter:
        query = query.filter_by(status=status_filter)

    submissoes = query.all()

    for sub in submissoes:
        sub_dict = {
            "id": sub.id,
            "lista_id": sub.lista_id,
            "lista_nome": sub.lista.nome if sub.lista else "N/A",
            "usuario_id": sub.usuario_id,
            "usuario_nome": sub.usuario.nome if sub.usuario else "N/A",
            "data_submissao": sub.data_submissao.isoformat(),
            "status": sub.status.value,
            "arquivada": sub.arquivada,
            "total_pedidos": sub.total_pedidos,
            "tipo_lista": "LISTA_TRADICIONAL",
            "pedidos": [
                {
                    "id": p.id,
                    "item_id": p.lista_mae_item_id,
                    "item_nome": p.item.nome if p.item else "N/A",
                    "quantidade_solicitada": float(p.quantidade_solicitada),
                    "status": p.status.value,
                    "unidade": p.item.unidade if p.item else ""
                }
                for p in sub.pedidos
            ]
        }
        resultado.append(sub_dict)

    # 2. BUSCAR LISTAS RÁPIDAS (exceto RASCUNHO)
    query_rapidas = ListaRapida.query.filter(
        ListaRapida.status != StatusListaRapida.RASCUNHO,
        ListaRapida.deletado == False
    ).options(
        db.joinedload(ListaRapida.usuario)
        # Nota: itens usa lazy='dynamic', não pode usar joinedload
    )

    if arquivadas:
        query_rapidas = query_rapidas.filter(ListaRapida.arquivada == True)
    else:
        query_rapidas = query_rapidas.filter(ListaRapida.arquivada == False)

    # Aplicar filtro de status se fornecido
    if status_filter:
        # Mapear status de submissão para status de lista rápida
        status_map = {
            "PENDENTE": StatusListaRapida.PENDENTE,
            "APROVADO": StatusListaRapida.APROVADA,
            "REJEITADO": StatusListaRapida.REJEITADA
        }
        if status_filter in status_map:
            query_rapidas = query_rapidas.filter_by(status=status_map[status_filter])

    listas_rapidas = query_rapidas.all()

    # Mapear listas rápidas para formato de submissão
    for lista in listas_rapidas:
        # Mapear status de lista rápida para formato de submissão
        status_normalizado = {
            StatusListaRapida.PENDENTE: "PENDENTE",
            StatusListaRapida.APROVADA: "APROVADO",
            StatusListaRapida.REJEITADA: "REJEITADO"
        }.get(lista.status, "PENDENTE")

        # Buscar itens da lista rápida
        itens_lista = lista.itens.all()

        lista_dict = {
            "id": lista.id,  # ID numérico da lista rápida
            "lista_id": lista.id,
            "lista_nome": lista.nome,
            "usuario_id": lista.usuario_id,
            "usuario_nome": lista.usuario.nome if lista.usuario else "N/A",
            "data_submissao": lista.submetido_em.isoformat() if lista.submetido_em else lista.criado_em.isoformat(),
            "status": status_normalizado,
            "arquivada": lista.arquivada,
            "total_pedidos": len(itens_lista),
            "tipo_lista": "LISTA_RAPIDA",
            "pedidos": [
                {
                    "id": item.id,
                    "item_id": item.item_global_id,
                    "item_nome": item.item_global.nome if item.item_global else "N/A",
                    "quantidade_solicitada": 1.0,  # Listas rápidas não têm quantidade explícita
                    "status": status_normalizado,
                    "unidade": item.item_global.unidade if item.item_global else "",
                    "prioridade": item.prioridade.value if item.prioridade else "precisa_comprar",
                    "observacao": item.observacao
                }
                for item in itens_lista
            ]
        }
        resultado.append(lista_dict)

    # 3. ORDENAR POR DATA DE SUBMISSÃO (mais recentes primeiro)
    resultado.sort(key=lambda x: x["data_submissao"], reverse=True)

    return resultado, 200


def get_submissao_by_id(submissao_id):
    """Retorna uma submissão tradicional específica por ID."""
    submissao = Submissao.query.options(
        db.joinedload(Submissao.lista),
        db.joinedload(Submissao.usuario),
        db.joinedload(Submissao.pedidos).joinedload(Pedido.item)
    ).filter_by(id=submissao_id).first()

    if not submissao:
        return {"error": "Submissão não encontrada."}, 404

    sub_dict = {
        "id": submissao.id,
        "lista_id": submissao.lista_id,
        "lista_nome": submissao.lista.nome if submissao.lista else "N/A",
        "usuario_id": submissao.usuario_id,
        "usuario_nome": submissao.usuario.nome if submissao.usuario else "N/A",
        "data_submissao": submissao.data_submissao.isoformat(),
        "status": submissao.status.value,
        "arquivada": submissao.arquivada,
        "total_pedidos": submissao.total_pedidos,
        "tipo_lista": "LISTA_TRADICIONAL",
        "pedidos": [
            {
                "id": p.id,
                "item_id": p.lista_mae_item_id,
                "item_nome": p.item.nome if p.item else "N/A",
                "quantidade_solicitada": float(p.quantidade_solicitada),
                "status": p.status.value,
                "unidade": p.item.unidade if p.item else ""
            }
            for p in submissao.pedidos
        ]
    }

    return sub_dict, 200


def aprovar_submissao(submissao_id):
    """Aprova todos os pedidos de uma submissão."""
    submissao = repositories.get_by_id(Submissao, submissao_id)
    if not submissao:
        return {"error": "Submissão não encontrada."}, 404
    
    # Aprovar todos os pedidos
    for pedido in submissao.pedidos:
        pedido.status = PedidoStatus.APROVADO
    
    submissao.status = SubmissaoStatus.APROVADO
    db.session.commit()
    
    return {"message": f"Submissão #{submissao_id} aprovada com sucesso!"}, 200


def rejeitar_submissao(submissao_id):
    """Rejeita todos os pedidos de uma submissão."""
    submissao = repositories.get_by_id(Submissao, submissao_id)
    if not submissao:
        return {"error": "Submissão não encontrada."}, 404
    
    # Rejeitar todos os pedidos
    for pedido in submissao.pedidos:
        pedido.status = PedidoStatus.REJEITADO
    
    submissao.status = SubmissaoStatus.REJEITADO
    db.session.commit()
    
    return {"message": f"Submissão #{submissao_id} rejeitada."}, 200


def reverter_submissao_para_pendente(submissao_id):
    """
    Reverte o status de uma submissão APROVADA ou REJEITADA para PENDENTE.
    Permite que admin reconsidere a decisão.
    """
    submissao = repositories.get_by_id(Submissao, submissao_id)
    if not submissao:
        return {"error": "Submissão não encontrada."}, 404
    
    # Validar que submissão não está PENDENTE
    if submissao.status == SubmissaoStatus.PENDENTE:
        return {"error": "Submissão já está PENDENTE."}, 400
    
    # Reverter todos os pedidos para PENDENTE
    for pedido in submissao.pedidos:
        pedido.status = PedidoStatus.PENDENTE
    
    submissao.status = SubmissaoStatus.PENDENTE
    db.session.commit()
    
    return {"message": f"Submissão #{submissao_id} revertida para PENDENTE."}, 200


def arquivar_submissao(submissao_id):
    """Arquiva uma submissão de qualquer status."""
    submissao = repositories.get_by_id(Submissao, submissao_id)
    if not submissao:
        return {"error": "Submissão não encontrada."}, 404

    if submissao.arquivada:
        return {"error": "Submissão já está arquivada."}, 400

    submissao.arquivada = True
    submissao.arquivada_em = brasilia_now()
    db.session.commit()

    return {"message": f"Submissão #{submissao_id} arquivada com sucesso."}, 200


def desarquivar_submissao(submissao_id):
    """Remove o status de arquivada de uma submissão."""
    submissao = repositories.get_by_id(Submissao, submissao_id)
    if not submissao:
        return {"error": "Submissão não encontrada."}, 404

    if not submissao.arquivada:
        return {"error": "Submissão não está arquivada."}, 400

    submissao.arquivada = False
    submissao.arquivada_em = None
    db.session.commit()

    return {"message": f"Submissão #{submissao_id} desarquivada com sucesso."}, 200


def deletar_submissao_permanente(submissao_id):
    """Exclui permanentemente uma submissão arquivada e todos os seus pedidos."""
    submissao = repositories.get_by_id(Submissao, submissao_id)
    if not submissao:
        return {"error": "Submissão não encontrada."}, 404

    if not submissao.arquivada:
        return {"error": "Apenas submissões arquivadas podem ser excluídas permanentemente."}, 400

    Pedido.query.filter_by(submissao_id=submissao_id).delete()
    db.session.delete(submissao)
    db.session.commit()

    return {"message": f"Submissão #{submissao_id} excluída permanentemente."}, 200


def editar_quantidades_submissao(submissao_id, pedidos_data):
    """
    Permite que admin edite as quantidades dos pedidos de uma submissão.
    NOTA: Agora recebe quantidades ATUAIS do estoque, não quantidades dos pedidos.
    Comportamento similar ao colaborador.
    
    Args:
        submissao_id: ID da submissão
        pedidos_data: Lista de dict com formato [{item_id: int, quantidade_atual: float}, ...]
    
    Returns:
        tuple: (response_dict, status_code)
    """
    try:
        print(f"[editar_quantidades_submissao] Iniciando edição da submissão #{submissao_id}")
        print(f"[editar_quantidades_submissao] Dados recebidos: {len(pedidos_data) if pedidos_data else 0} itens")
        
        submissao = repositories.get_by_id(Submissao, submissao_id)
        if not submissao:
            print(f"[editar_quantidades_submissao] ERRO: Submissão não encontrada")
            return {"error": "Submissão não encontrada."}, 404
        
        print(f"[editar_quantidades_submissao] Status da submissão: {submissao.status}")
        
        # Validar que submissão está PENDENTE
        if submissao.status != SubmissaoStatus.PENDENTE:
            print(f"[editar_quantidades_submissao] ERRO: Submissão não está PENDENTE")
            return {"error": "Apenas submissões PENDENTES podem ser editadas."}, 400
        
        # Validar dados recebidos
        if not pedidos_data or not isinstance(pedidos_data, list):
            print(f"[editar_quantidades_submissao] ERRO: Dados inválidos")
            return {"error": "Dados inválidos. Esperado: lista de itens."}, 400
        
        # Buscar todos os itens da lista
        refs = ListaItemRef.query.filter_by(lista_id=submissao.lista_id).all()
        refs_map = {ref.item_id: ref for ref in refs}
        
        print(f"[editar_quantidades_submissao] Lista tem {len(refs)} itens")
        
        # Atualizar quantidade_atual de cada item
        itens_atualizados = 0
        for item in pedidos_data:
            item_id = item.get('item_id')
            nova_quantidade_atual = item.get('quantidade_atual')
            
            # Validações
            if not item_id or nova_quantidade_atual is None:
                continue
            
            if item_id not in refs_map:
                print(f"[editar_quantidades_submissao] ERRO: Item #{item_id} não pertence à lista")
                return {"error": f"Item #{item_id} não pertence a esta lista."}, 400
            
            if nova_quantidade_atual < 0:
                print(f"[editar_quantidades_submissao] ERRO: Quantidade negativa para item #{item_id}")
                return {"error": f"Quantidade não pode ser negativa (Item #{item_id})."}, 400
            
            # Atualizar quantidade_atual no ListaItemRef
            ref = refs_map[item_id]
            ref.quantidade_atual = nova_quantidade_atual
            itens_atualizados += 1
        
        print(f"[editar_quantidades_submissao] {itens_atualizados} itens atualizados")
        
        # Deletar pedidos antigos desta submissão
        Pedido.query.filter_by(submissao_id=submissao_id).delete()
        
        # Recriar pedidos com base nas novas quantidades
        pedidos_criados = 0
        for ref in refs:
            # Usa a mesma regra de get_pedido() — respeita usa_threshold/fardo
            pedido_qtd = float(ref.get_pedido())

            if pedido_qtd > 0:
                novo_pedido = Pedido(
                    submissao_id=submissao_id,
                    lista_mae_item_id=ref.item_id,
                    quantidade_solicitada=pedido_qtd,
                    usuario_id=submissao.usuario_id,  # CRÍTICO: pega da submissão
                    status=PedidoStatus.PENDENTE
                )
                db.session.add(novo_pedido)
                pedidos_criados += 1
        
        print(f"[editar_quantidades_submissao] {pedidos_criados} pedidos criados")
        
        # Atualizar total de pedidos da submissão
        submissao.total_pedidos = pedidos_criados

        log_event(
            acao="update",
            entidade="lista",
            entidade_id=submissao.lista_id,
            mensagem="Quantidades da submissão atualizadas",
            restaurante_id=submissao.lista.restaurante_id if submissao.lista else None,
            metadata={
                "submissao_id": submissao_id,
                "itens_atualizados": itens_atualizados,
                "pedidos_criados": pedidos_criados
            }
        )
        db.session.commit()
        
        print(f"[editar_quantidades_submissao] Edição concluída com sucesso")
        
        return {
            "message": f"{itens_atualizados} item(ns) atualizado(s), {pedidos_criados} pedido(s) gerado(s)!",
            "submissao_id": submissao_id,
            "pedidos_criados": pedidos_criados
        }, 200
        
    except Exception as e:
        print(f"[editar_quantidades_submissao] EXCEÇÃO: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return {"error": f"Erro ao editar submissão: {str(e)}"}, 500


def submit_pedidos(user_id):
    """Cria registros de Pedido para todos os itens que estão abaixo do estoque mínimo."""
    itens_a_pedir = db.session.query(Estoque, Item).join(Item).filter(Estoque.quantidade_atual < Estoque.quantidade_minima).all()

    if not itens_a_pedir:
        return {"message": "Nenhum item precisa de reposição." }, 200

    novos_pedidos = []
    for estoque, item in itens_a_pedir:
        quantidade_a_pedir = estoque.quantidade_minima - estoque.quantidade_atual
        novo_pedido = Pedido(
            item_id=item.id,
            fornecedor_id=item.fornecedor_id,
            quantidade_solicitada=quantidade_a_pedir,
            usuario_id=user_id
        )
        novos_pedidos.append(novo_pedido)
        db.session.add(novo_pedido)

    log_event(
        acao="create",
        entidade="pedido",
        entidade_id=None,
        mensagem="Pedidos gerados automaticamente",
        usuario_id=user_id,
        restaurante_id=_get_usuario_restaurante_id(user_id),
        metadata={"pedidos_criados": len(novos_pedidos)}
    )
    db.session.commit()
    return {"message": f"{len(novos_pedidos)} pedidos foram gerados com sucesso." }, 201


def get_all_pedidos(status_filter=None):
    """
    Retorna todos os pedidos do sistema.

    Args:
        status_filter: Filtra por status (PENDENTE, APROVADO, REJEITADO). Se None, retorna todos.
    """
    query = Pedido.query

    if status_filter:
        query = query.filter_by(status=status_filter)

    pedidos = query.order_by(Pedido.data_pedido.desc()).all()
    return pedidos, 200


def _recalcular_status_submissao(submissao):
    if not submissao:
        return
    pedidos = submissao.pedidos or []
    if not pedidos:
        return

    has_pendente = any(p.status == PedidoStatus.PENDENTE for p in pedidos)
    if has_pendente:
        submissao.status = SubmissaoStatus.PENDENTE
        return

    has_aprovado = any(p.status == PedidoStatus.APROVADO for p in pedidos)
    has_rejeitado = any(p.status == PedidoStatus.REJEITADO for p in pedidos)

    if has_aprovado and has_rejeitado:
        submissao.status = SubmissaoStatus.PARCIALMENTE_APROVADO
    elif has_aprovado:
        submissao.status = SubmissaoStatus.APROVADO
    elif has_rejeitado:
        submissao.status = SubmissaoStatus.REJEITADO


def aprovar_pedido(pedido_id):
    """Aprova um pedido pendente."""
    from .models import TipoNotificacao

    pedido = repositories.get_by_id(Pedido, pedido_id)

    if not pedido:
        return {"error": "Pedido não encontrado."}, 404

    if pedido.status != PedidoStatus.PENDENTE:
        return {"error": f"Apenas pedidos pendentes podem ser aprovados. Status atual: {pedido.status.value}"}, 400

    pedido.status = PedidoStatus.APROVADO
    _recalcular_status_submissao(pedido.submissao)
    db.session.commit()

    # Notificar o usuário que criou o pedido
    try:
        # Buscar restaurante_id do usuário do pedido
        usuario = Usuario.query.get(pedido.usuario_id)
        restaurante_id = usuario.restaurante_id if usuario else None

        criar_notificacao(
            usuario_id=pedido.usuario_id,
            titulo="Pedido aprovado",
            tipo=TipoNotificacao.PEDIDO_APROVADO,
            mensagem=f"Seu pedido foi aprovado.",
            pedido_id=pedido.id,
            restaurante_id=restaurante_id
        )
    except Exception as e:
        print(f"[AVISO] Erro ao criar notificação de aprovação para pedido {pedido.id}: {str(e)}")

    return {"message": "Pedido aprovado com sucesso.", "pedido": pedido.to_dict()}, 200


def rejeitar_pedido(pedido_id):
    """Rejeita um pedido pendente."""
    from .models import TipoNotificacao

    pedido = repositories.get_by_id(Pedido, pedido_id)

    if not pedido:
        return {"error": "Pedido não encontrado."}, 404

    if pedido.status != PedidoStatus.PENDENTE:
        return {"error": f"Apenas pedidos pendentes podem ser rejeitados. Status atual: {pedido.status.value}"}, 400

    pedido.status = PedidoStatus.REJEITADO
    _recalcular_status_submissao(pedido.submissao)
    db.session.commit()

    # Notificar o usuário que criou o pedido
    try:
        # Buscar restaurante_id do usuário do pedido
        usuario = Usuario.query.get(pedido.usuario_id)
        restaurante_id = usuario.restaurante_id if usuario else None

        criar_notificacao(
            usuario_id=pedido.usuario_id,
            titulo="Pedido rejeitado",
            tipo=TipoNotificacao.PEDIDO_REJEITADO,
            mensagem=f"Seu pedido foi rejeitado.",
            pedido_id=pedido.id,
            restaurante_id=restaurante_id
        )
    except Exception as e:
        print(f"[AVISO] Erro ao criar notificação de rejeição para pedido {pedido.id}: {str(e)}")

    return {"message": "Pedido rejeitado com sucesso.", "pedido": pedido.to_dict()}, 200


def aprovar_pedidos_lote(pedido_ids):
    """Aprova múltiplos pedidos em lote."""
    if not pedido_ids or not isinstance(pedido_ids, list):
        return {"error": "Lista de IDs inválida."}, 400

    aprovados = 0
    erros = []
    submissao_ids = set()

    for pedido_id in pedido_ids:
        pedido = repositories.get_by_id(Pedido, pedido_id)

        if not pedido:
            erros.append(f"Pedido {pedido_id} não encontrado")
            continue

        if pedido.status != PedidoStatus.PENDENTE:
            erros.append(f"Pedido {pedido_id} não está pendente")
            continue

        pedido.status = PedidoStatus.APROVADO
        aprovados += 1
        if pedido.submissao_id:
            submissao_ids.add(pedido.submissao_id)

    if submissao_ids:
        submissoes = Submissao.query.filter(Submissao.id.in_(submissao_ids)).all()
        for submissao in submissoes:
            _recalcular_status_submissao(submissao)

    db.session.commit()

    return {
        "message": f"{aprovados} pedido(s) aprovado(s) com sucesso.",
        "aprovados": aprovados,
        "erros": erros if erros else None
    }, 200


def rejeitar_pedidos_lote(pedido_ids):
    """Rejeita múltiplos pedidos em lote."""
    if not pedido_ids or not isinstance(pedido_ids, list):
        return {"error": "Lista de IDs inválida."}, 400

    rejeitados = 0
    erros = []
    submissao_ids = set()

    for pedido_id in pedido_ids:
        pedido = repositories.get_by_id(Pedido, pedido_id)

        if not pedido:
            erros.append(f"Pedido {pedido_id} não encontrado")
            continue

        if pedido.status != PedidoStatus.PENDENTE:
            erros.append(f"Pedido {pedido_id} não está pendente")
            continue

        pedido.status = PedidoStatus.REJEITADO
        rejeitados += 1
        if pedido.submissao_id:
            submissao_ids.add(pedido.submissao_id)

    if submissao_ids:
        submissoes = Submissao.query.filter(Submissao.id.in_(submissao_ids)).all()
        for submissao in submissoes:
            _recalcular_status_submissao(submissao)

    db.session.commit()

    return {
        "message": f"{rejeitados} pedido(s) rejeitado(s) com sucesso.",
        "rejeitados": rejeitados,
        "erros": erros if erros else None
    }, 200


def editar_pedido(pedido_id, data):
    """
    Edita um pedido (permite alterar quantidade solicitada).
    Apenas pedidos pendentes podem ser editados.
    """
    pedido = repositories.get_by_id(Pedido, pedido_id)

    if not pedido:
        return {"error": "Pedido não encontrado."}, 404

    if pedido.status != PedidoStatus.PENDENTE:
        return {"error": f"Apenas pedidos pendentes podem ser editados. Status atual: {pedido.status.value}"}, 400

    # Atualizar quantidade
    if 'quantidade_solicitada' in data:
        quantidade = data['quantidade_solicitada']

        if not isinstance(quantidade, (int, float)) or quantidade <= 0:
            return {"error": "Quantidade deve ser um número positivo."}, 400

        quantidade_anterior = float(pedido.quantidade_solicitada) if pedido.quantidade_solicitada is not None else None
        pedido.quantidade_solicitada = quantidade

        log_event(
            acao="update",
            entidade="pedido",
            entidade_id=pedido.id,
            mensagem="Quantidade do pedido atualizada",
            restaurante_id=pedido.usuario.restaurante_id if pedido.usuario else None,
            metadata={
                "quantidade_anterior": quantidade_anterior,
                "quantidade_solicitada": quantidade
            }
        )
    db.session.commit()

    return {"message": "Pedido editado com sucesso.", "pedido": pedido.to_dict()}, 200

def atualizar_estoque_e_calcular_pedido(estoque_id, quantidade_atual, usuario_id):
    """
    Atualiza a quantidade atual de um estoque e calcula o pedido automaticamente.
    Também registra auditoria (usuário e data da submissão).
    """
    estoque = repositories.get_by_id(Estoque, estoque_id)
    if not estoque:
        return {"error": "Item de estoque não encontrado."}, 404

    # Atualiza quantidade atual
    quantidade_anterior = float(estoque.quantidade_atual) if estoque.quantidade_atual is not None else None
    estoque.quantidade_atual = quantidade_atual

    # Calcula pedido automaticamente
    estoque.pedido = estoque.calcular_pedido()

    # Registra auditoria
    estoque.data_ultima_submissao = brasilia_now()
    estoque.usuario_ultima_submissao_id = usuario_id

    log_event(
        acao="update",
        entidade="lista",
        entidade_id=estoque.lista_id,
        mensagem="Quantidade de estoque atualizada",
        restaurante_id=estoque.lista.restaurante_id if estoque.lista else None,
        metadata={
            "estoque_id": estoque.id,
            "item_id": estoque.item_id,
            "quantidade_anterior": quantidade_anterior,
            "quantidade_atual": quantidade_atual
        }
    )
    db.session.commit()
    return estoque.to_dict(), 200

def update_submissao(submissao_id, usuario_id, items_data):
    """
    Atualiza uma submissão existente PENDENTE.
    - Atualiza quantidades atuais dos itens
    - Recalcula pedidos
    - Atualiza ou cria novos pedidos
    - Marca submissao como atualizada
    """
    submissao = repositories.get_by_id(Submissao, submissao_id)
    if not submissao:
        return {"error": "Submissão não encontrada."}, 404
    
    # Validar que é do usuário
    if submissao.usuario_id != usuario_id:
        return {"error": "Você não pode editar esta submissão."}, 403
    
    # Validar que está PENDENTE
    if submissao.status != SubmissaoStatus.PENDENTE:
        return {"error": "Só é possível editar submissões PENDENTES."}, 400
    
    lista_id = submissao.lista_id
    lista = repositories.get_by_id(Lista, lista_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    # Buscar refs atuais
    item_ids = [item.get('estoque_id') for item in items_data if item.get('estoque_id')]
    refs_map = {}
    if item_ids:
        refs = ListaItemRef.query.options(
            db.joinedload(ListaItemRef.item)
        ).filter(
            ListaItemRef.lista_id == lista_id,
            ListaItemRef.item_id.in_(item_ids)
        ).all()
        refs_map = {ref.item_id: ref for ref in refs}

    # Deletar pedidos antigos desta submissão
    Pedido.query.filter_by(submissao_id=submissao_id).delete()

    pedidos_criados = []
    refs_atualizados = []
    agora = brasilia_now()

    for item_data in items_data:
        estoque_id = item_data.get('estoque_id')  # item_id
        quantidade_atual = item_data.get('quantidade_atual')

        if not estoque_id or quantidade_atual is None:
            continue

        ref = refs_map.get(estoque_id)
        if not ref:
            continue

        # Atualiza quantidade
        ref.quantidade_atual = quantidade_atual
        ref.atualizado_em = agora
        refs_atualizados.append(ref)

        # Recria Pedido se necessário
        if float(quantidade_atual) < float(ref.quantidade_minima):
            quantidade_a_pedir = float(ref.quantidade_minima) - float(quantidade_atual)
            
            novo_pedido = Pedido(
                submissao_id=submissao.id,
                lista_mae_item_id=ref.item_id,
                fornecedor_id=None,
                quantidade_solicitada=quantidade_a_pedir,
                usuario_id=usuario_id,
                status=PedidoStatus.PENDENTE
            )
            pedidos_criados.append(novo_pedido)

    # Atualizar contador e data
    submissao.total_pedidos = len(pedidos_criados)
    submissao.data_submissao = agora  # Atualiza data da submissão

    # Commit
    if pedidos_criados:
        db.session.add_all(pedidos_criados)
    log_event(
        acao="update",
        entidade="lista",
        entidade_id=lista_id,
        mensagem="Submissão atualizada pelo colaborador",
        restaurante_id=lista.restaurante_id if lista else None,
        metadata={
            "submissao_id": submissao.id,
            "estoques_atualizados": len(refs_atualizados),
            "pedidos_criados": len(pedidos_criados)
        }
    )
    db.session.commit()

    return {
        "message": f"Submissão atualizada! {len(pedidos_criados)} pedido(s) criado(s).",
        "submissao_id": submissao.id,
        "estoques_atualizados": len(refs_atualizados),
        "pedidos_criados": len(pedidos_criados)
    }, 200


def submit_estoque_lista(lista_id, usuario_id, items_data):
    """
    Submete múltiplos itens de estoque de uma lista.
    REFATORADO: Usa ListaItemRef em vez de Estoque.
    Calcula pedidos automaticamente e cria registros de Pedido se necessário.

    items_data: [{"estoque_id": 1, "quantidade_atual": 5}, ...]
    NOTA: estoque_id é interpretado como item_id (compatibilidade)
    
    OTIMIZADO: Busca TODOS os refs de uma vez (1 query ao invés de N queries).
    Cria Submissao para agrupar pedidos.
    """
    lista = repositories.get_by_id(Lista, lista_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    usuario = repositories.get_by_id(Usuario, usuario_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    if lista.restaurante_id != usuario.restaurante_id and usuario.role != UserRoles.SUPER_ADMIN:
        return {"error": "Acesso negado. Esta lista não pertence ao seu restaurante."}, 403

    # Valida se usuário está atribuído à lista (Admin/SuperAdmin têm acesso automático)
    if not _is_admin_or_super_admin(usuario):
        if usuario_id not in [u.id for u in lista.colaboradores]:
            return {"error": "Você não tem acesso a esta lista."}, 403

    # 🚀 OTIMIZAÇÃO: Buscar TODOS os refs de uma vez com IN (1 query ao invés de 32)
    item_ids = [item.get('estoque_id') for item in items_data if item.get('estoque_id')]
    refs_map = {}
    if item_ids:
        refs = ListaItemRef.query.options(
            db.joinedload(ListaItemRef.item)
        ).filter(
            ListaItemRef.lista_id == lista_id,
            ListaItemRef.item_id.in_(item_ids)
        ).all()
        refs_map = {ref.item_id: ref for ref in refs}

    # Criar Submissao primeiro
    submissao = Submissao(
        lista_id=lista_id,
        usuario_id=usuario_id,
        status=SubmissaoStatus.PENDENTE
    )
    db.session.add(submissao)
    db.session.flush()  # Gera ID da submissao

    pedidos_criados = []
    refs_atualizados = []
    agora = brasilia_now()

    for item_data in items_data:
        estoque_id = item_data.get('estoque_id')  # Na verdade é item_id
        quantidade_atual = item_data.get('quantidade_atual')

        if not estoque_id or quantidade_atual is None:
            continue

        ref = refs_map.get(estoque_id)
        if not ref:
            continue

        # Atualiza quantidade
        ref.quantidade_atual = quantidade_atual
        ref.atualizado_em = agora
        refs_atualizados.append(ref)

        # Cria Pedido usando a mesma regra de get_pedido() — respeita usa_threshold/fardo
        # ref.quantidade_atual já foi atualizado acima, então get_pedido() usa o valor correto
        quantidade_a_pedir = float(ref.get_pedido())
        if quantidade_a_pedir > 0:
            novo_pedido = Pedido(
                submissao_id=submissao.id,
                lista_mae_item_id=ref.item_id,
                fornecedor_id=None,
                quantidade_solicitada=quantidade_a_pedir,
                usuario_id=usuario_id,
                status=PedidoStatus.PENDENTE
            )
            pedidos_criados.append(novo_pedido)
            
            current_app.logger.info(
                f"[SUBMIT] Pedido criado: item {ref.item.nome}, "
                f"qtd={quantidade_a_pedir}, usuario={usuario_id}"
            )

    # Atualizar contador de pedidos na submissao
    submissao.total_pedidos = len(pedidos_criados)

    # Commit único ao final
    if pedidos_criados:
        db.session.add_all(pedidos_criados)
    log_event(
        acao="create",
        entidade="lista",
        entidade_id=lista_id,
        mensagem="Lista submetida com atualização de quantidades",
        restaurante_id=lista.restaurante_id,
        usuario_id=usuario_id,
        metadata={
            "submissao_id": submissao.id,
            "estoques_atualizados": len(refs_atualizados),
            "pedidos_criados": len(pedidos_criados)
        }
    )
    db.session.commit()

    return {
        "message": f"Lista submetida com sucesso! {len(pedidos_criados)} pedido(s) criado(s).",
        "submissao_id": submissao.id,
        "estoques_atualizados": len(refs_atualizados),
        "pedidos_criados": len(pedidos_criados)
    }, 201

# --- Serviços de Lista ---

def _get_lista_por_restaurante(lista_id, restaurante_id, include_deleted=False):
    query = Lista.query.filter_by(id=lista_id)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    if not include_deleted:
        query = query.filter_by(deletado=False)
    return query.first()

def create_lista(data, restaurante_id=None):
    """Cria uma nova lista de compras, opcionalmente com itens do catálogo global."""
    if not data or not data.get('nome'):
        return {"error": "O nome da lista é obrigatório."}, 400

    if restaurante_id is None:
        restaurante_id = data.get('restaurante_id')
    if restaurante_id is None:
        return {"error": "Restaurante é obrigatório."}, 400

    # Validar se já existe uma lista com esse nome
    nome = data['nome'].strip()
    lista_query = Lista.query.filter(func.lower(Lista.nome) == func.lower(nome))
    if restaurante_id is not None:
        lista_query = lista_query.filter_by(restaurante_id=restaurante_id)
    lista_existente = lista_query.first()
    if lista_existente:
        return {"error": f"Já existe uma lista com o nome '{nome}'."}, 409

    area_id = data.get('area_id') or None
    if area_id is not None:
        try:
            area_id = int(area_id)
        except (ValueError, TypeError):
            area_id = None

    nova_lista = Lista(
        nome=nome,
        descricao=data.get('descricao', '').strip() if data.get('descricao') else None,
        restaurante_id=restaurante_id,
        area_id=area_id
    )
    db.session.add(nova_lista)
    db.session.flush()  # Garantir que nova_lista.id está disponível
    
    # Adicionar itens do catálogo global, se fornecidos
    itens_data = data.get('itens', [])
    if itens_data and isinstance(itens_data, list):
        for item_data in itens_data:
            if 'id' in item_data:
                resultado, status = adicionar_item_na_lista_com_copia(
                    nova_lista.id,
                    item_data,
                    restaurante_id
                )
                if status in (200, 201):
                    continue
                continue

            item_id = item_data.get('item_id')
            if not item_id:
                continue

            # Validar que o item existe no catálogo global
            item = _get_item_catalogo_por_id(item_id, restaurante_id)
            if not item:
                continue

            # Criar referência lista-item
            lista_item_ref = ListaItemRef(
                lista_id=nova_lista.id,
                item_id=item_id,
                quantidade_atual=float(item_data.get('quantidade_atual', 0)),
                quantidade_minima=float(item_data.get('quantidade_minima', 1.0))
            )
            db.session.add(lista_item_ref)

    # Vincular colaboradores: payload explícito tem prioridade sobre auto-assign por área
    colaborador_ids = data.get('colaborador_ids')
    if colaborador_ids is not None:
        colaboradores = []
        for uid in colaborador_ids:
            user = repositories.get_by_id(Usuario, uid)
            if user and user.role == UserRoles.COLLABORATOR:
                colaboradores.append(user)
        nova_lista.colaboradores = colaboradores
    elif area_id is not None:
        area = repositories.get_by_id(Area, area_id)
        if area and area.colaboradores:
            nova_lista.colaboradores = list(area.colaboradores)

    log_event(
        acao="create",
        entidade="lista",
        entidade_id=nova_lista.id,
        mensagem="Lista criada",
        restaurante_id=restaurante_id
    )
    db.session.commit()

    lista_dict = nova_lista.to_dict()
    lista_dict['area_nome'] = nova_lista.area.nome if nova_lista.area else None
    return lista_dict, 201


# ===============================================
# SERVIÇO: Busca Unificada de Itens
# ===============================================

def buscar_itens_para_lista(restaurante_id, query=None):
    """
    Busca itens de duas fontes para criação de listas:
    1. Lista Global (ListaMaeItem) do próprio restaurante
    2. Catálogo de Fornecedores (Item) da região
    """
    if not restaurante_id:
        return {"error": "Restaurante é obrigatório."}, 400

    resultado = []

    query_global = ListaMaeItem.query.filter_by(restaurante_id=restaurante_id)
    if query:
        query_global = query_global.filter(ListaMaeItem.nome.ilike(f'%{query}%'))
    itens_globais = query_global.order_by(ListaMaeItem.nome).all()

    nomes_globais = {normalize_item_nome(item.nome) for item in itens_globais}

    for item in itens_globais:
        resultado.append({
            "id": f"global_{item.id}",
            "nome": item.nome,
            "unidade": item.unidade,
            "origem": "lista_global",
            "origem_display": "Lista do Restaurante",
            "fornecedor_id": None,
            "fornecedor_nome": None,
            "ja_na_lista_global": True,
            "item_fornecedor_id": None
        })

    fornecedores = Fornecedor.query.filter_by(restaurante_id=restaurante_id).all()
    fornecedor_ids = [f.id for f in fornecedores]

    if fornecedor_ids:
        query_fornecedores = db.session.query(Item, Fornecedor).join(
            Fornecedor, Item.fornecedor_id == Fornecedor.id
        ).filter(Item.fornecedor_id.in_(fornecedor_ids))

        if query:
            query_fornecedores = query_fornecedores.filter(Item.nome.ilike(f'%{query}%'))

        itens_fornecedores = query_fornecedores.order_by(Item.nome).all()

        for item, fornecedor in itens_fornecedores:
            ja_na_global = normalize_item_nome(item.nome) in nomes_globais
            resultado.append({
                "id": f"fornecedor_{item.id}",
                "nome": item.nome,
                "unidade": item.unidade_medida,
                "origem": "fornecedor",
                "origem_display": fornecedor.nome,
                "fornecedor_id": fornecedor.id,
                "fornecedor_nome": fornecedor.nome,
                "ja_na_lista_global": ja_na_global,
                "item_fornecedor_id": item.id
            })

    resultado_sorted = sorted(resultado, key=lambda x: x['nome'].lower())
    return {"itens": resultado_sorted}, 200


def get_itens_regionais(restaurante_id, query=None):
    """Retorna itens de fornecedores regionais (compartilhados) do restaurante."""
    if not restaurante_id:
        return {"error": "Restaurante é obrigatório."}, 400

    itens_query = db.session.query(Item, Fornecedor).join(
        Fornecedor, Item.fornecedor_id == Fornecedor.id
    ).filter(
        Fornecedor.restaurante_id == restaurante_id,
        Fornecedor.compartilhado_regiao.is_(True)
    )

    if query:
        termo = f"%{query}%"
        itens_query = itens_query.filter(
            or_(Item.nome.ilike(termo), Fornecedor.nome.ilike(termo))
        )

    itens_query = itens_query.order_by(Item.nome)

    itens = [
        {
            "id": item.id,
            "nome": item.nome,
            "unidade_medida": item.unidade_medida,
            "fornecedor_id": fornecedor.id,
            "fornecedor_nome": fornecedor.nome
        }
        for item, fornecedor in itens_query.all()
    ]

    return {"itens": itens, "total": len(itens)}, 200


# ===============================================
# SERVIÇO: Cópia Automática para Lista Global
# ===============================================

def copiar_item_fornecedor_para_global(item_fornecedor_id, restaurante_id):
    """Cria cópia de um item de fornecedor na lista global do restaurante."""
    if not restaurante_id:
        return {"error": "Restaurante é obrigatório."}, 400

    item_fornecedor = Item.query.filter_by(id=item_fornecedor_id).first()
    if not item_fornecedor:
        return {"error": "Item de fornecedor não encontrado."}, 404

    fornecedor = Fornecedor.query.filter_by(id=item_fornecedor.fornecedor_id).first()
    if not fornecedor:
        return {"error": "Fornecedor não encontrado."}, 404

    if fornecedor.restaurante_id != restaurante_id:
        return {"error": "Item não pertence a um fornecedor desta região."}, 403

    item_existente = ListaMaeItem.query.filter(
        ListaMaeItem.restaurante_id == restaurante_id,
        func.lower(ListaMaeItem.nome) == func.lower(item_fornecedor.nome)
    ).first()

    if item_existente:
        return {
            "item_global_id": item_existente.id,
            "created": False,
            "message": "Item já existe na lista global."
        }, 200

    novo_item_global = ListaMaeItem(
        nome=item_fornecedor.nome,
        unidade=item_fornecedor.unidade_medida,
        restaurante_id=restaurante_id
    )
    db.session.add(novo_item_global)
    db.session.commit()

    return {
        "item_global_id": novo_item_global.id,
        "created": True,
        "message": "Item copiado para a lista global com sucesso."
    }, 201


def adicionar_item_na_lista_com_copia(lista_id, item_data, restaurante_id, skip_if_exists=False):
    """
    Adiciona item à lista, copiando automaticamente de fornecedor se necessário.
    """
    lista = db.session.get(Lista, lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Lista não pertence ao seu restaurante."}, 403

    item_id_str = item_data.get('id', '')
    origem, item_id = item_id_str.split('_', 1) if '_' in item_id_str else (None, None)

    if not origem or not item_id:
        return {"error": "ID de item inválido."}, 400

    try:
        item_id = int(item_id)
    except ValueError:
        return {"error": "ID de item inválido."}, 400

    if origem == 'global':
        item_global = _get_item_catalogo_por_id(item_id, restaurante_id)
        if not item_global:
            return {"error": "Item global não encontrado ou não pertence ao seu restaurante."}, 404
        item_global_id = item_global.id
    elif origem == 'fornecedor':
        resultado, status = copiar_item_fornecedor_para_global(item_id, restaurante_id)
        if status not in (200, 201):
            return resultado, status
        item_global_id = resultado['item_global_id']
    else:
        return {"error": "Origem de item inválida."}, 400

    ref_existente = ListaItemRef.query.filter_by(
        lista_id=lista_id,
        item_id=item_global_id
    ).first()

    quantidade_minima = float(item_data.get('quantidade_minima', 1.0))

    if ref_existente:
        if skip_if_exists:
            return {
                "message": "Item já estava na lista.",
                "lista_id": lista_id,
                "item_id": item_global_id,
                "skipped": True
            }, 200

        ref_existente.quantidade_minima = quantidade_minima
        db.session.commit()
        return {
            "message": "Item já estava na lista, quantidade mínima atualizada.",
            "lista_id": lista_id,
            "item_id": item_global_id,
            "updated": True
        }, 200

    nova_ref = ListaItemRef(
        lista_id=lista_id,
        item_id=item_global_id,
        quantidade_atual=0,
        quantidade_minima=quantidade_minima
    )
    db.session.add(nova_ref)
    db.session.commit()

    return {
        "message": "Item adicionado à lista com sucesso.",
        "lista_id": lista_id,
        "item_id": item_global_id
    }, 201

def get_all_listas(restaurante_id=None):
    """Retorna todas as listas de compras não deletadas."""
    query = Lista.query.filter_by(deletado=False)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)

    # Eager load dos relacionamentos usados (evita N+1 queries)
    listas = query.options(
        selectinload(Lista.colaboradores),
        selectinload(Lista.area),
    ).all()

    # Buscar todos os nomes de itens em uma query (evita N+1 queries)
    lista_ids = [lista.id for lista in listas]
    itens_por_lista = {}
    if lista_ids:
        refs = (
            db.session.query(ListaItemRef.lista_id, ListaMaeItem.nome)
            .join(ListaMaeItem, ListaItemRef.item_id == ListaMaeItem.id)
            .filter(ListaItemRef.lista_id.in_(lista_ids))
            .all()
        )
        for lista_id, nome in refs:
            itens_por_lista.setdefault(lista_id, []).append(nome)

    # Serialização manual — evita to_dict() que carrega todos os backrefs lazily
    resultado = []
    for lista in listas:
        resultado.append({
            'id': lista.id,
            'nome': lista.nome,
            'descricao': lista.descricao,
            'data_criacao': lista.data_criacao.isoformat() if lista.data_criacao else None,
            'data_delecao': lista.data_delecao.isoformat() if lista.data_delecao else None,
            'area_id': lista.area_id,
            'area_nome': lista.area.nome if lista.area else None,
            'colaboradores': [
                {
                    'id': c.id,
                    'nome': c.nome,
                    'email': c.email,
                    'role': c.role.value if hasattr(c.role, 'value') else c.role,
                }
                for c in lista.colaboradores
            ],
            'itens_nomes': itens_por_lista.get(lista.id, []),
        })

    return resultado, 200

def get_lista_by_id(lista_id, restaurante_id=None):
    """Retorna uma lista específica."""
    if restaurante_id is None:
        return repositories.get_by_id(Lista, lista_id), 200
    lista = _get_lista_por_restaurante(lista_id, restaurante_id, include_deleted=True)
    return lista, 200

def assign_colaboradores_to_lista(lista_id, data, restaurante_id=None):
    """Atribui um ou mais colaboradores a uma lista."""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    colaborador_ids = data.get('colaborador_ids', [])
    if not colaborador_ids:
        # Se a lista de IDs está vazia, remove todos os colaboradores
        lista.colaboradores = []
    else:
        # Buscar todos os colaboradores válidos primeiro
        colaboradores_novos = []
        for user_id in colaborador_ids:
            user = repositories.get_by_id(Usuario, user_id)
            if (
                user
                and user.role == UserRoles.COLLABORATOR
                and (restaurante_id is None or user.restaurante_id == restaurante_id)
            ):
                colaboradores_novos.append(user)
        
        # Substituir de uma vez (mais eficiente)
        lista.colaboradores = colaboradores_novos
    
    db.session.commit()

    lista_dict = lista.to_dict()
    lista_dict["colaboradores"] = [colaborador.to_dict() for colaborador in lista.colaboradores]
    return lista_dict, 200

def unassign_colaborador_from_lista(lista_id, data, restaurante_id=None):
    """Desatribui um colaborador de uma lista."""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    colaborador_id = data.get('colaborador_id')
    if not colaborador_id:
        return {"error": "ID do colaborador não fornecido."}, 400

    user = repositories.get_by_id(Usuario, colaborador_id)
    if not user:
        return {"error": "Usuário não encontrado."}, 404

    if user in lista.colaboradores:
        lista.colaboradores.remove(user)
        db.session.commit()
        return {"message": "Colaborador desatribuído com sucesso."}, 200
    else:
        return {"error": "Colaborador não está atribuído a esta lista."}, 400

def update_lista(lista_id, data, restaurante_id=None):
    """Atualiza nome e/ou descrição de uma lista."""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    # Validar se nome já existe (se estiver sendo alterado)
    if 'nome' in data and data['nome']:
        novo_nome = data['nome'].strip()
        # Verificar se o nome mudou (case-insensitive)
        if novo_nome.lower() != lista.nome.lower():
            existing_query = Lista.query.filter(
                func.lower(Lista.nome) == func.lower(novo_nome),
                Lista.id != lista_id
            )
            if restaurante_id is not None:
                existing_query = existing_query.filter_by(restaurante_id=restaurante_id)
            existing = existing_query.first()
            if existing:
                return {"error": f"Já existe uma lista com o nome '{novo_nome}'."}, 409

    # Atualizar campos
    if 'nome' in data and data['nome']:
        lista.nome = data['nome'].strip()
    if 'descricao' in data:
        lista.descricao = data['descricao'].strip() if data['descricao'] else None
    if 'area_id' in data:
        raw = data['area_id']
        if raw is None or raw == '':
            lista.area_id = None
        else:
            try:
                lista.area_id = int(raw)
            except (ValueError, TypeError):
                lista.area_id = None

    log_event(
        acao="update",
        entidade="lista",
        entidade_id=lista.id,
        mensagem="Lista atualizada",
        restaurante_id=restaurante_id or lista.restaurante_id
    )
    db.session.commit()
    lista_dict = lista.to_dict()
    lista_dict['area_nome'] = lista.area.nome if lista.area else None
    return lista_dict, 200

def delete_lista(lista_id, restaurante_id=None):
    """Faz soft delete de uma lista (move para lixeira)."""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    # Soft delete
    lista.deletado = True
    lista.data_delecao = brasilia_now()
    log_event(
        acao="delete",
        entidade="lista",
        entidade_id=lista.id,
        mensagem="Lista movida para lixeira",
        restaurante_id=restaurante_id or lista.restaurante_id
    )
    db.session.commit()

    return {"message": "Lista movida para lixeira."}, 200


def get_deleted_listas(restaurante_id=None):
    """Retorna todas as listas deletadas (na lixeira)."""
    query = Lista.query.filter_by(deletado=True)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    listas = query.order_by(Lista.data_delecao.desc()).all()
    return [lista.to_dict() for lista in listas], 200


def restore_lista(lista_id, restaurante_id=None):
    """Restaura uma lista da lixeira."""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id, include_deleted=True)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    if not lista.deletado:
        return {"error": "Esta lista não está na lixeira."}, 400

    # Restaurar
    lista.deletado = False
    lista.data_delecao = None
    log_event(
        acao="restore",
        entidade="lista",
        entidade_id=lista.id,
        mensagem="Lista restaurada",
        restaurante_id=restaurante_id or lista.restaurante_id
    )
    db.session.commit()

    return {"message": "Lista restaurada com sucesso.", "lista": lista.to_dict()}, 200


def permanent_delete_lista(lista_id, restaurante_id=None):
    """Deleta permanentemente uma lista da lixeira."""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id, include_deleted=True)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    if not lista.deletado:
        return {"error": "Apenas listas na lixeira podem ser deletadas permanentemente."}, 400

    # Deletar permanentemente
    log_event(
        acao="permanent_delete",
        entidade="lista",
        entidade_id=lista.id,
        mensagem="Lista deletada permanentemente",
        restaurante_id=restaurante_id or lista.restaurante_id
    )
    db.session.delete(lista)
    db.session.commit()

    return {"message": "Lista deletada permanentemente."}, 200


def permanent_delete_listas_batch(lista_ids, restaurante_id=None):
    """Deleta permanentemente múltiplas listas em lote."""
    if not lista_ids or not isinstance(lista_ids, list):
        return {"error": "Lista de IDs inválida."}, 400

    deleted_count = 0
    errors = []

    for lista_id in lista_ids:
        lista = _get_lista_por_restaurante(lista_id, restaurante_id, include_deleted=True)
        if not lista:
            errors.append(f"Lista {lista_id} não encontrada")
            continue

        if not lista.deletado:
            errors.append(f"Lista {lista_id} não está na lixeira")
            continue

        try:
            db.session.delete(lista)
            deleted_count += 1
        except Exception as e:
            errors.append(f"Erro ao deletar lista {lista_id}: {str(e)}")

    db.session.commit()

    return {
        "message": f"{deleted_count} lista(s) deletada(s) permanentemente.",
        "deleted_count": deleted_count,
        "errors": errors if errors else None
    }, 200


# --- Serviços de Dashboard ---

def get_user_stats(user_id):
    """Retorna estatísticas para o dashboard de um usuário específico."""
    pending_submissions = Pedido.query.filter_by(usuario_id=user_id, status='PENDENTE').count()
    completed_lists = 0 # Placeholder, a lógica de "lista completa" precisa ser definida

    stats_data = {
        'pending_submissions': pending_submissions,
        'completed_lists': completed_lists,
    }
    
    return stats_data, 200

def get_estatisticas(restaurante_id):
    """Retorna estatísticas de itens completos/faltantes por lista e submissões."""
    listas = Lista.query.filter_by(restaurante_id=restaurante_id, deletado=False).all()

    # Pré-carregar catálogo de itens para evitar N+1
    items_dict = {
        i.id: i for i in ListaMaeItem.query.filter_by(restaurante_id=restaurante_id).all()
    }

    por_lista = []
    total_ok = 0
    total_faltantes = 0
    itens_detalhe = []

    for lista in listas:
        refs_all = ListaItemRef.query.filter_by(lista_id=lista.id).all()
        # Ignorar itens com quantidade_minima = 0, igual ao comportamento da lista mãe
        refs = [r for r in refs_all if r.quantidade_minima > 0]
        if not refs:
            continue
        itens_ok = sum(1 for r in refs if r.quantidade_atual >= r.quantidade_minima)
        itens_faltantes = sum(1 for r in refs if r.quantidade_atual < r.quantidade_minima)
        total_ok += itens_ok
        total_faltantes += itens_faltantes
        por_lista.append({
            "lista_id": lista.id,
            "lista_nome": lista.nome,
            "area_nome": lista.area.nome if lista.area else None,
            "area_id": lista.area_id,
            "total_itens": len(refs),
            "itens_ok": itens_ok,
            "itens_faltantes": itens_faltantes,
        })

        for ref in refs:
            item = items_dict.get(ref.item_id)
            itens_detalhe.append({
                "item_id": ref.item_id,
                "item_nome": item.nome if item else f"Item #{ref.item_id}",
                "unidade": item.unidade if item else "",
                "lista_id": lista.id,
                "lista_nome": lista.nome,
                "area_id": lista.area_id,
                "area_nome": lista.area.nome if lista.area else None,
                "quantidade_atual": float(ref.quantidade_atual),
                "quantidade_minima": float(ref.quantidade_minima),
            })

    # Submissões por status — últimos 30 dias
    cutoff = brasilia_now() - timedelta(days=30)
    lista_ids = [l.id for l in listas]
    submissoes_raw = (
        db.session.query(Submissao.status, func.count(Submissao.id))
        .filter(
            Submissao.lista_id.in_(lista_ids),
            Submissao.data_submissao >= cutoff,
            Submissao.arquivada == False,
        )
        .group_by(Submissao.status)
        .all()
    )
    status_dict = {"PENDENTE": 0, "APROVADO": 0, "REJEITADO": 0, "PARCIALMENTE_APROVADO": 0}
    for status, count in submissoes_raw:
        status_dict[status.value] = count

    return {
        "resumo": {
            "total_listas": len(por_lista),
            "total_itens": total_ok + total_faltantes,
            "itens_ok": total_ok,
            "itens_faltantes": total_faltantes,
        },
        "por_lista": sorted(por_lista, key=lambda x: x["itens_faltantes"], reverse=True),
        "submissoes_por_status": status_dict,
        "itens": itens_detalhe,
    }, 200


def get_dashboard_summary(restaurante_id=None):
    """Retorna dados agregados para o dashboard do admin."""
    users_query = Usuario.query
    if restaurante_id is not None:
        users_query = users_query.filter(Usuario.role != UserRoles.SUPER_ADMIN).filter_by(restaurante_id=restaurante_id)
    total_users = users_query.count()
    pending_users = users_query.filter_by(aprovado=False).count()
    listas_query = Lista.query
    if restaurante_id is not None:
        listas_query = listas_query.filter_by(restaurante_id=restaurante_id)
    total_lists = listas_query.count()
    if restaurante_id is not None:
        total_items = Item.query.join(Fornecedor).filter(
            Fornecedor.restaurante_id == restaurante_id
        ).count()
        if total_items == 0:
            total_items = ListaMaeItem.query.filter_by(restaurante_id=restaurante_id).count()
    else:
        total_items = Item.query.count()
        if total_items == 0:
            total_items = ListaMaeItem.query.count()
    pending_submissions_query = Submissao.query.filter_by(
        status=SubmissaoStatus.PENDENTE,
        arquivada=False
    )
    if restaurante_id is not None:
        pending_submissions_query = pending_submissions_query.join(Usuario, Submissao.usuario_id == Usuario.id).filter(
            Usuario.restaurante_id == restaurante_id
        )
    pending_submissions = pending_submissions_query.count()
    orders_today = Pedido.query.filter(
        func.date(Pedido.data_pedido) == brasilia_now().date()
    ).count()
    pending_cotacoes = Cotacao.query.filter_by(status=CotacaoStatus.PENDENTE).count()
    completed_cotacoes = Cotacao.query.filter_by(status=CotacaoStatus.CONCLUIDA).count()

    summary_data = {
        'total_users': total_users,
        'pending_users': pending_users,
        'total_lists': total_lists,
        'total_items': total_items,
        'pending_submissions': pending_submissions,
        'pending_cotacoes': pending_cotacoes,
        'orders_today': orders_today,
        'completed_cotacoes': completed_cotacoes,
    }
    
    return summary_data, 200

def get_activity_summary():
    """Retorna o número de pedidos por dia nos últimos 7 dias."""
    today = brasilia_now().date()
    dates = [today - timedelta(days=i) for i in range(6, -1, -1)] # Últimos 7 dias

    activity_data = []
    for d in dates:
        count = Pedido.query.filter(func.date(Pedido.data_pedido) == d).count()
        activity_data.append(count)

    labels = [d.strftime('%d/%m') for d in dates]

    return {"labels": labels, "data": activity_data}, 200


def get_admin_recent_activities(restaurante_id=None, limit=6):
    """Retorna atividades recentes do admin (submissoes, cotacoes e usuarios)."""
    activities = []

    submissoes_query = Submissao.query.join(Usuario, Submissao.usuario_id == Usuario.id)
    if restaurante_id is not None:
        submissoes_query = submissoes_query.filter(Usuario.restaurante_id == restaurante_id)
    submissoes = (
        submissoes_query
        .order_by(Submissao.data_submissao.desc())
        .limit(limit)
        .all()
    )
    for submissao in submissoes:
        if not submissao.data_submissao:
            continue
        lista_nome = submissao.lista.nome if submissao.lista else 'Lista'
        usuario_nome = submissao.usuario.nome if submissao.usuario else 'Usuário'
        activities.append((submissao.data_submissao, f'Lista "{lista_nome}" submetida por {usuario_nome}'))

    cotacoes_query = Cotacao.query.join(Fornecedor)
    if restaurante_id is not None:
        cotacoes_query = cotacoes_query.filter(Fornecedor.restaurante_id == restaurante_id)
    cotacoes = (
        cotacoes_query
        .order_by(Cotacao.data_cotacao.desc())
        .limit(limit)
        .all()
    )
    for cotacao in cotacoes:
        if not cotacao.data_cotacao:
            continue
        fornecedor_nome = cotacao.fornecedor.nome if cotacao.fornecedor else 'Fornecedor'
        activities.append((cotacao.data_cotacao, f'Cotação #{cotacao.id} criada para {fornecedor_nome}'))

    usuarios_query = Usuario.query.filter(Usuario.role != UserRoles.SUPER_ADMIN)
    if restaurante_id is not None:
        usuarios_query = usuarios_query.filter(Usuario.restaurante_id == restaurante_id)
    usuarios = (
        usuarios_query
        .order_by(Usuario.criado_em.desc())
        .limit(limit)
        .all()
    )
    for usuario in usuarios:
        if not usuario.criado_em:
            continue
        activities.append((usuario.criado_em, f'Usuário {usuario.nome} cadastrado'))

    activities.sort(key=lambda item: item[0], reverse=True)
    response = [
        {"time": item[0].strftime('%H:%M'), "description": item[1]}
        for item in activities[:limit]
    ]

    return {"activities": response}, 200


def get_collaborator_recent_activities(user_id, limit=6):
    """Retorna atividades recentes do colaborador."""
    activities = []

    submissoes = (
        Submissao.query
        .filter_by(usuario_id=user_id)
        .order_by(Submissao.data_submissao.desc())
        .limit(limit)
        .all()
    )
    for submissao in submissoes:
        if not submissao.data_submissao:
            continue
        lista_nome = submissao.lista.nome if submissao.lista else 'Lista'
        activities.append((submissao.data_submissao, f'Você submeteu a lista "{lista_nome}"'))

    pedidos = (
        Pedido.query
        .filter_by(usuario_id=user_id)
        .order_by(Pedido.data_pedido.desc())
        .limit(limit)
        .all()
    )
    status_map = {
        PedidoStatus.PENDENTE: 'pendente',
        PedidoStatus.APROVADO: 'aprovado',
        PedidoStatus.REJEITADO: 'rejeitado',
    }
    for pedido in pedidos:
        if not pedido.data_pedido:
            continue
        status_value = pedido.status.value if hasattr(pedido.status, 'value') else str(pedido.status)
        status_label = status_map.get(pedido.status, status_value.lower())
        activities.append((pedido.data_pedido, f'Pedido #{pedido.id} {status_label}'))

    activities.sort(key=lambda item: item[0], reverse=True)
    response = [
        {"time": item[0].strftime('%H:%M'), "description": item[1]}
        for item in activities[:limit]
    ]

    return {"activities": response}, 200


def get_catalogo_global(restaurante_id=None):
    """
    Retorna todos os itens do catálogo global.
    Usado pelo admin no card "Itens e Insumos".
    """
    # Query otimizada: uma única consulta com agregação
    from sqlalchemy import func as sql_func
    
    itens_query = db.session.query(
        ListaMaeItem.id,
        ListaMaeItem.restaurante_id,
        ListaMaeItem.nome,
        ListaMaeItem.unidade,
        ListaMaeItem.criado_em,
        ListaMaeItem.atualizado_em,
        sql_func.count(ListaItemRef.lista_id).label('total_listas')
    ).outerjoin(
        ListaItemRef, ListaMaeItem.id == ListaItemRef.item_id
    )

    if restaurante_id is not None:
        itens_query = itens_query.filter(ListaMaeItem.restaurante_id == restaurante_id)

    itens_query = itens_query.group_by(
        ListaMaeItem.id
    ).order_by(
        ListaMaeItem.nome
    ).all()

    itens_data = [{
        "id": item.id,
        "restaurante_id": item.restaurante_id,
        "nome": item.nome,
        "unidade": item.unidade,
        "total_listas": item.total_listas,
        "criado_em": item.criado_em.isoformat() if item.criado_em else None,
        "atualizado_em": item.atualizado_em.isoformat() if item.atualizado_em else None
    } for item in itens_query]

    return {"itens": itens_data, "total": len(itens_data)}, 200


def editar_item_catalogo_global(item_id, data, restaurante_id=None):
    """
    Edita um item do catálogo global (ListaMaeItem).
    Permite alterar nome e unidade.
    """
    query = ListaMaeItem.query.filter_by(id=item_id)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    item = query.first()
    if not item:
        return {"error": "Item não encontrado."}, 404
    if restaurante_id is None:
        restaurante_id = item.restaurante_id

    nome = data.get('nome', '').strip()
    unidade = data.get('unidade', '').strip()

    if not nome:
        return {"error": "Nome do item é obrigatório."}, 400
    if not unidade:
        return {"error": "Unidade é obrigatória."}, 400

    # Verifica se já existe outro item com o mesmo nome
    item_existente = ListaMaeItem.query.filter(
        ListaMaeItem.restaurante_id == restaurante_id,
        ListaMaeItem.nome == nome,
        ListaMaeItem.id != item_id
    ).first()
    
    if item_existente:
        return {"error": f"Já existe um item com o nome '{nome}'."}, 409

    nome_anterior = item.nome
    unidade_anterior = item.unidade
    item.nome = nome
    item.unidade = unidade
    item.atualizado_em = brasilia_now()

    log_event(
        acao="update",
        entidade="item_global",
        entidade_id=item.id,
        mensagem="Item atualizado no catalogo global",
        restaurante_id=restaurante_id,
        metadata={
            "nome_anterior": nome_anterior,
            "unidade_anterior": unidade_anterior,
            "nome": item.nome,
            "unidade": item.unidade
        }
    )
    db.session.commit()

    return {
        "message": "Item atualizado com sucesso.",
        "item": {
            "id": item.id,
            "nome": item.nome,
            "unidade": item.unidade,
            "atualizado_em": item.atualizado_em.isoformat()
        }
    }, 200


def criar_item_catalogo_global(data, restaurante_id=None):
    """Cria um novo item diretamente no catálogo global."""
    from sqlalchemy import func

    nome = data.get('nome', '').strip()
    unidade = data.get('unidade', '').strip()

    if not nome:
        return {"error": "Nome do item é obrigatório."}, 400
    if not unidade:
        return {"error": "Unidade é obrigatória."}, 400

    if restaurante_id is None:
        return {"error": "Restaurante é obrigatório."}, 400

    # Verifica duplicação (case-insensitive)
    item_existente = ListaMaeItem.query.filter(
        ListaMaeItem.restaurante_id == restaurante_id,
        func.lower(ListaMaeItem.nome) == func.lower(nome)
    ).first()

    if item_existente:
        return {"error": f"Já existe um item com o nome '{nome}'."}, 409

    item = ListaMaeItem(nome=nome, unidade=unidade, restaurante_id=restaurante_id)
    db.session.add(item)
    db.session.flush()
    log_event(
        acao="create",
        entidade="item_global",
        entidade_id=item.id,
        mensagem="Item criado no catalogo global",
        restaurante_id=restaurante_id
    )
    db.session.commit()

    return {"message": "Item criado com sucesso.", "item": item.to_dict()}, 201


def get_listas_status_submissoes(restaurante_id=None):
    """
    Retorna o status das submissões de listas com pedidos pendentes.

    Para cada lista, inclui:
    - ID e nome
    - Data da última submissão
    - Quantidade de pedidos pendentes
    
    Também inclui listas rápidas pendentes.
    """
    resultado = []
    
    # 1. Listas comuns com pedidos pendentes
    listas_query = Lista.query.filter_by(deletado=False)
    if restaurante_id is not None:
        listas_query = listas_query.filter_by(restaurante_id=restaurante_id)
    listas = listas_query.all()

    for lista in listas:
        estoques = Estoque.query.filter_by(lista_id=lista.id).all()
        if not estoques:
            continue

        datas_submissao = [e.data_ultima_submissao for e in estoques if e.data_ultima_submissao]
        ultima_submissao = max(datas_submissao) if datas_submissao else None

        item_ids = {e.item_id for e in estoques if e.item_id}
        if not item_ids:
            continue

        pedidos_pendentes = Pedido.query.filter(
            Pedido.item_id.in_(item_ids),
            Pedido.status == PedidoStatus.PENDENTE
        ).count()

        if pedidos_pendentes > 0:
            resultado.append({
                "id": lista.id,
                "nome": lista.nome,
                "area": lista.nome,
                "tipo": "lista_comum",
                "last_submission": ultima_submissao.isoformat() if ultima_submissao else None,
                "pending_submissions": pedidos_pendentes
            })
    
    # 2. Listas rápidas pendentes
    listas_rapidas_query = ListaRapida.query.filter_by(
        deletado=False,
        status=StatusListaRapida.PENDENTE
    )
    if restaurante_id is not None:
        listas_rapidas_query = listas_rapidas_query.join(Usuario).filter(
            Usuario.restaurante_id == restaurante_id
        )
    listas_rapidas_pendentes = listas_rapidas_query.all()
    
    for lista_rapida in listas_rapidas_pendentes:
        total_itens = lista_rapida.itens.count()
        resultado.append({
            "id": lista_rapida.id,
            "nome": lista_rapida.nome,
            "area": f"Lista Rápida - {lista_rapida.usuario.nome if lista_rapida.usuario else 'Usuário'}",
            "tipo": "lista_rapida",
            "last_submission": lista_rapida.submetido_em.isoformat() if lista_rapida.submetido_em else None,
            "pending_submissions": total_itens
        })

    return resultado, 200

def aprovar_todos_pedidos_lista(lista_id, restaurante_id=None):
    """
    Aprova todos os pedidos pendentes associados aos itens de uma lista.
    """
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    estoques = Estoque.query.filter_by(lista_id=lista_id).all()
    item_ids = {e.item_id for e in estoques if e.item_id}

    if not item_ids:
        return {"error": "Lista não possui itens."}, 404

    pedidos_pendentes = Pedido.query.filter(
        Pedido.item_id.in_(item_ids),
        Pedido.status == PedidoStatus.PENDENTE
    ).all()

    if not pedidos_pendentes:
        return {"error": "Nenhum pedido pendente encontrado para esta lista."}, 404

    for pedido in pedidos_pendentes:
        pedido.status = PedidoStatus.APROVADO

    db.session.commit()

    return {
        "message": f"Todos os {len(pedidos_pendentes)} pedidos pendentes foram aprovados com sucesso!",
        "pedidos_aprovados": len(pedidos_pendentes),
        "lista_nome": lista.nome
    }, 200

def get_collaborator_dashboard_summary(user_id):
    """Retorna estatísticas do dashboard para colaboradores."""
    from .models import Usuario, Pedido, Lista, Submissao, SubmissaoStatus
    from .extensions import db

    # Buscar o usuário
    usuario = db.session.get(Usuario, user_id)
    if not usuario:
        return {"error": "Usuário não encontrado"}, 404

    # Número de áreas atribuídas ao colaborador (via listas)
    minhas_areas = db.session.query(Lista).filter(
        Lista.colaboradores.any(id=user_id)
    ).count()

    # Submissões pendentes do colaborador
    pending_submissions = Submissao.query.filter_by(
        usuario_id=user_id,
        status=SubmissaoStatus.PENDENTE,
        arquivada=False
    ).count()

    # Submissões concluídas do colaborador
    completed_submissions = Submissao.query.filter(
        Submissao.usuario_id == user_id,
        Submissao.status.in_([
            SubmissaoStatus.APROVADO,
            SubmissaoStatus.REJEITADO,
            SubmissaoStatus.PARCIALMENTE_APROVADO,
        ]),
        Submissao.arquivada.is_(False)
    ).count()

    # Pedidos pendentes
    pedidos_pendentes = Pedido.query.filter_by(
        usuario_id=user_id,
        status=PedidoStatus.PENDENTE
    ).count()

    summary_data = {
        "minhas_areas": minhas_areas,
        "pending_submissions": pending_submissions,
        "completed_submissions": completed_submissions,
        "pedidos_pendentes": pedidos_pendentes
    }

    return summary_data, 200

def get_minhas_listas_status(user_id):
    """
    Retorna status das listas atribuídas ao colaborador.
    Inclui última submissão e quantidade de itens pendentes por lista.
    """
    usuario = repositories.get_by_id(Usuario, user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    if usuario.role == UserRoles.SUPER_ADMIN:
        listas = Lista.query.filter_by(deletado=False).all()
    elif _is_admin_or_super_admin(usuario):
        listas = Lista.query.filter_by(
            restaurante_id=usuario.restaurante_id,
            deletado=False
        ).all()
    else:
        listas = [
            lista
            for lista in usuario.listas_atribuidas
            if not lista.deletado and lista.restaurante_id == usuario.restaurante_id
        ]
    resultado = []

    for lista in listas:
        estoques = Estoque.query.filter_by(lista_id=lista.id).all()
        datas_submissao = [e.data_ultima_submissao for e in estoques if e.data_ultima_submissao]
        ultima_submissao = max(datas_submissao) if datas_submissao else None

        pending_items = sum(
            1 for estoque in estoques
            if float(estoque.quantidade_atual) < float(estoque.quantidade_minima)
        )

        resultado.append({
            "id": lista.id,
            "nome": lista.nome,
            "last_submission": ultima_submissao.isoformat() if ultima_submissao else None,
            "pending_items": pending_items
        })

    return resultado, 200

def get_minhas_areas_status(user_id):
    """
    Retorna status das áreas atribuídas ao colaborador.
    """
    usuario = repositories.get_by_id(Usuario, user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    pedidos = Pedido.query.filter_by(usuario_id=user_id).all()
    areas_status = {}

    for pedido in pedidos:
        if pedido.submissao and pedido.submissao.lista:
            lista = pedido.submissao.lista
            area_id = lista.id
            area_nome = lista.nome

            if area_id not in areas_status:
                areas_status[area_id] = {
                    "id": area_id,
                    "area": area_nome,
                    "last_submission": None,
                    "pending_items": 0
                }

            if pedido.data_pedido:
                if not areas_status[area_id]["last_submission"] or pedido.data_pedido > areas_status[area_id]["last_submission"]:
                    areas_status[area_id]["last_submission"] = pedido.data_pedido

            if pedido.status == PedidoStatus.PENDENTE:
                areas_status[area_id]["pending_items"] += 1

    resultado = []
    for area_status in areas_status.values():
        if area_status["last_submission"]:
            area_status["last_submission"] = area_status["last_submission"].strftime("%Y-%m-%d %H:%M")
        else:
            area_status["last_submission"] = "Nunca"
        resultado.append(area_status)

    return resultado, 200

# --- Serviços de Listas com Estoque (Nova Feature) ---

def get_minhas_listas(user_id):
    """Retorna todas as listas atribuídas a um colaborador."""
    current_app.logger.info(f"[GET_MINHAS_LISTAS] user_id: {user_id}")

    usuario = repositories.get_by_id(Usuario, user_id)
    if not usuario:
        current_app.logger.error(f"[GET_MINHAS_LISTAS] Usuário {user_id} não encontrado")
        return {"error": "Usuário não encontrado."}, 404

    # Admin/SuperAdmin veem TODAS as listas do restaurante, colaboradores apenas as atribuídas
    if usuario.role == UserRoles.SUPER_ADMIN:
        listas = Lista.query.filter_by(deletado=False).all()
    elif _is_admin_or_super_admin(usuario):
        listas = Lista.query.filter_by(
            restaurante_id=usuario.restaurante_id,
            deletado=False
        ).all()
    else:
        listas = [
            l for l in usuario.listas_atribuidas
            if not l.deletado and l.restaurante_id == usuario.restaurante_id
        ]
    
    current_app.logger.info(f"[GET_MINHAS_LISTAS] Usuário: {usuario.nome}, Listas atribuídas: {len(listas)}")
    for lista in listas:
        current_app.logger.info(f"  - Lista: {lista.id} - {lista.nome}")

    return {"listas": [lista.to_dict() for lista in listas]}, 200

def get_estoque_by_lista(lista_id, restaurante_id=None):
    """Retorna todos os estoques (itens) de uma lista específica."""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    estoques = Estoque.query.filter_by(lista_id=lista_id).all()
    return [estoque.to_dict() for estoque in estoques], 200

def get_lista_mae_consolidada(lista_id, restaurante_id=None):
    """
    Retorna a Lista Mãe consolidada com última submissão de cada item.
    Usada pelo admin para visualizar o consolidado de todas as submissões.
    """
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    estoques = Estoque.query.filter_by(lista_id=lista_id).all()

    itens_consolidados = []
    for estoque in estoques:
        usuario_submissao = None
        if estoque.usuario_ultima_submissao:
            usuario_submissao = {
                "id": estoque.usuario_ultima_submissao.id,
                "nome": estoque.usuario_ultima_submissao.nome
            }

        item_consolidado = {
            "estoque_id": estoque.id,
            "item_id": estoque.item_id,
            "item_nome": estoque.item.nome,
            "item_unidade": estoque.item.unidade_medida,
            "fornecedor_id": estoque.item.fornecedor_id,
            "fornecedor_nome": estoque.item.fornecedor.nome if estoque.item.fornecedor else None,
            "quantidade_minima": float(estoque.quantidade_minima),
            "quantidade_atual": float(estoque.quantidade_atual),
            "pedido": float(estoque.pedido) if estoque.pedido else estoque.calcular_pedido(),
            "data_ultima_submissao": estoque.data_ultima_submissao.isoformat() if estoque.data_ultima_submissao else None,
            "usuario_ultima_submissao": usuario_submissao
        }
        itens_consolidados.append(item_consolidado)

    consolidado = {
        "lista_id": lista.id,
        "lista_nome": lista.nome,
        "lista_descricao": lista.descricao,
        "data_criacao": lista.data_criacao.isoformat() if lista.data_criacao else None,
        "itens": itens_consolidados,
        "total_itens": len(itens_consolidados),
        "total_em_falta": sum(1 for item in itens_consolidados if item["pedido"] > 0),
        "total_pedido": sum(item["pedido"] for item in itens_consolidados)
    }

    return consolidado, 200

def adicionar_itens_na_lista(lista_id, items_data, restaurante_id=None):
    """
    Adiciona/atualiza itens de estoque em uma lista.

    items_data: [
        {"item_id": 1, "quantidade_minima": 10},
        {"item_id": 2, "quantidade_minima": 5}
    ]
    """
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    itens_adicionados = []
    itens_atualizados = []

    for item_data in items_data:
        item_id = item_data.get('item_id')
        quantidade_minima = item_data.get('quantidade_minima', 0)

        if not item_id:
            continue

        item = repositories.get_by_id(Item, item_id)
        if not item:
            continue

        # Verifica se o estoque já existe
        estoque_existente = Estoque.query.filter_by(
            lista_id=lista_id,
            item_id=item_id
        ).first()

        if estoque_existente:
            # Atualiza quantidade mínima
            estoque_existente.quantidade_minima = quantidade_minima
            itens_atualizados.append(item.nome)
        else:
            # Cria novo estoque vinculado à lista
            novo_estoque = Estoque(
                lista_id=lista_id,
                item_id=item_id,
                area_id=1,  # Padrão para listas (não é específico de área)
                quantidade_atual=0,
                quantidade_minima=quantidade_minima,
                pedido=0
            )
            db.session.add(novo_estoque)
            itens_adicionados.append(item.nome)

    log_event(
        acao="update",
        entidade="lista",
        entidade_id=lista_id,
        mensagem="Itens adicionados/atualizados na lista",
        restaurante_id=restaurante_id or lista.restaurante_id,
        metadata={
            "itens_adicionados": itens_adicionados,
            "itens_atualizados": itens_atualizados
        }
    )
    db.session.commit()

    return {
        "message": f"{len(itens_adicionados)} itens adicionados à lista.",
        "itens_adicionados": itens_adicionados
    }, 200

def obter_itens_da_lista(lista_id, restaurante_id=None):
    """Retorna todos os itens (estoques) vinculados a uma lista"""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    estoques = Estoque.query.filter_by(lista_id=lista_id).all()

    itens = []
    for estoque in estoques:
        itens.append({
            "estoque_id": estoque.id,
            "item_id": estoque.item_id,
            "item_nome": estoque.item.nome,
            "quantidade_minima": float(estoque.quantidade_minima),
            "quantidade_atual": float(estoque.quantidade_atual),
            "unidade_medida": estoque.item.unidade_medida,
            "fornecedor_id": estoque.item.fornecedor_id
        })

    return itens, 200

def remover_item_da_lista(lista_id, item_id, restaurante_id=None):
    """Remove um item (estoque) de uma lista"""
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    estoque = Estoque.query.filter_by(
        lista_id=lista_id,
        item_id=item_id
    ).first()

    if not estoque:
        return {"error": "Item não encontrado nesta lista."}, 404

    log_event(
        acao="update",
        entidade="lista",
        entidade_id=lista_id,
        mensagem="Item removido da lista",
        restaurante_id=restaurante_id or lista.restaurante_id,
        metadata={
            "estoque_id": estoque.id,
            "item_id": estoque.item_id,
            "item_nome": estoque.item.nome if estoque.item else None
        }
    )
    db.session.delete(estoque)
    db.session.commit()

    return {"message": f"Item removido da lista."}, 200


# ===== LISTA MAE ITENS (Nova Funcionalidade) =====

def normalize_item_nome(value):
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = re.sub(r"\s+", " ", normalized).strip().lower()
    return normalized

def sync_lista_mae_itens_para_estoque(lista_id):
    """
    FUNÇÃO LEGADA - DEPRECADA
    
    Esta função foi substituída pela arquitetura ListaItemRef.
    Mantida apenas para compatibilidade temporária.
    
    TODO: Remover esta função após garantir que nenhum código a chama.
    
    Sincroniza itens da lista (via ListaItemRef) com o estoque.
    Apenas cria estoque para itens com quantidade_minima > 0.
    """
    # FUNÇÃO DESABILITADA - Retorna imediatamente
    current_app.logger.warning(
        f"[DEPRECADO] sync_lista_mae_itens_para_estoque() foi chamada para lista {lista_id}. "
        "Esta função não é mais necessária com a arquitetura ListaItemRef."
    )
    return {"criados": 0, "atualizados": 0, "ignorados": 0, "warning": "Função deprecada"}
    
    # Código original comentado abaixo (para referência)
    """
    # Buscar referências de itens para esta lista
    refs = ListaItemRef.query.filter_by(lista_id=lista_id).all()
    if not refs:
        return {"criados": 0, "atualizados": 0, "ignorados": 0}

    # Carregar itens cadastrados para matching por nome
    itens_cadastrados = Item.query.all()
    itens_por_nome = {}
    for item in itens_cadastrados:
        chave = normalize_item_nome(item.nome)
        if not chave or chave in itens_por_nome:
            continue
        itens_por_nome[chave] = item

    criados = 0
    atualizados = 0
    ignorados = 0

    for ref in refs:
        nome_normalizado = normalize_item_nome(ref.item.nome)
        if not nome_normalizado:
            ignorados += 1
            continue

        item = itens_por_nome.get(nome_normalizado)
        if not item:
            ignorados += 1
            continue

        estoque = Estoque.query.filter_by(lista_id=lista_id, item_id=item.id).first()
        quantidade_minima = float(ref.quantidade_minima or 0)
        quantidade_atual = float(ref.quantidade_atual or 0)

        if not estoque:
            if quantidade_minima <= 0:
                ignorados += 1
                continue

            novo_estoque = Estoque(
                lista_id=lista_id,
                item_id=item.id,
                area_id=1,
                quantidade_atual=quantidade_atual,
                quantidade_minima=quantidade_minima,
                pedido=0
            )
            db.session.add(novo_estoque)
            criados += 1
            continue

        estoque.quantidade_minima = quantidade_minima
        if estoque.data_ultima_submissao is None:
            estoque.quantidade_atual = quantidade_atual
        estoque.pedido = estoque.calcular_pedido()
        atualizados += 1

    db.session.commit()

    return {"criados": criados, "atualizados": atualizados, "ignorados": ignorados}
    """

def adicionar_item_lista_mae(lista_id, data, restaurante_id=None):
    """
    Adiciona um item ao catálogo global e vincula à lista.
    Se o item já existe no catálogo, apenas cria/atualiza a referência.
    """
    try:
        lista = _get_lista_por_restaurante(lista_id, restaurante_id)
        if not lista:
            return {"error": "Lista não encontrada"}, 404
        if restaurante_id is None:
            return {"error": "Restaurante é obrigatório"}, 400

        nome = data.get('nome', '').strip()
        if not nome:
            return {"error": "Nome do item é obrigatório"}, 400

        # Buscar ou criar item no catálogo global (por nome, case-insensitive)
        item_criado = False
        item = ListaMaeItem.query.filter(
            ListaMaeItem.restaurante_id == restaurante_id,
            func.lower(ListaMaeItem.nome) == func.lower(nome)
        ).first()

        if not item:
            # Criar novo item no catálogo global
            item = ListaMaeItem(
                nome=nome,
                unidade=data.get('unidade', 'un'),
                restaurante_id=restaurante_id
            )
            db.session.add(item)
            db.session.flush()  # Para obter o ID
            item_criado = True

        # Verificar se já existe referência para esta lista
        ref = ListaItemRef.query.filter_by(lista_id=lista_id, item_id=item.id).first()
        ref_existente = ref is not None
        quantidade_atual_anterior = ref.quantidade_atual if ref else None
        quantidade_minima_anterior = ref.quantidade_minima if ref else None

        if ref:
            # Atualizar quantidades existentes
            ref.quantidade_atual = data.get('quantidade_atual', ref.quantidade_atual)
            ref.quantidade_minima = data.get('quantidade_minima', ref.quantidade_minima)
            ref.atualizado_em = brasilia_now()
        else:
            # Criar nova referência
            ref = ListaItemRef(
                lista_id=lista_id,
                item_id=item.id,
                quantidade_atual=data.get('quantidade_atual', 0),
                quantidade_minima=data.get('quantidade_minima', 1.0)
            )
            db.session.add(ref)

        log_event(
            acao="create" if not ref_existente else "update",
            entidade="lista",
            entidade_id=lista_id,
            mensagem="Item adicionado na lista mãe" if not ref_existente else "Item atualizado na lista mãe",
            restaurante_id=restaurante_id,
            metadata={
                "item_id": item.id,
                "item_nome": item.nome,
                "item_criado_catalogo": item_criado,
                "quantidade_atual_anterior": quantidade_atual_anterior,
                "quantidade_minima_anterior": quantidade_minima_anterior,
                "quantidade_atual": ref.quantidade_atual,
                "quantidade_minima": ref.quantidade_minima
            }
        )
        db.session.commit()
        # sync_lista_mae_itens_para_estoque(lista_id)  # REMOVIDO - Não mais necessário

        # Retornar no mesmo formato flat usado pelo GET /lista-mae
        return {
            "id": item.id,
            "nome": item.nome,
            "unidade": item.unidade,
            "quantidade_atual": ref.quantidade_atual,
            "quantidade_minima": ref.quantidade_minima,
            "pedido": ref.get_pedido(),
            "usa_threshold": ref.usa_threshold,
            "quantidade_por_fardo": ref.quantidade_por_fardo,
            "criado_em": ref.criado_em.isoformat() if ref.criado_em else None,
            "atualizado_em": ref.atualizado_em.isoformat() if ref.atualizado_em else None
        }, 201
    except Exception as e:
        import traceback
        from flask import current_app

        current_app.logger.error(f"[adicionar_item_lista_mae] Erro ao criar item:")
        current_app.logger.error(f"  Tipo: {type(e).__name__}")
        current_app.logger.error(f"  Mensagem: {str(e)}")
        current_app.logger.error(f"  Traceback: {traceback.format_exc()}")

        db.session.rollback()
        return {"error": f"{type(e).__name__}: {str(e)}"}, 500


def editar_item_lista_mae(lista_id, item_id, data, restaurante_id=None):
    """
    Edita a referência de um item em uma lista específica.
    Se o nome for alterado, busca ou cria um item no catálogo global
    e atualiza a referência para o novo item.
    """
    try:
        # Buscar a referência atual (item_id é o ID do item global)
        ref = ListaItemRef.query.filter_by(
            lista_id=lista_id,
            item_id=item_id
        ).first()

        if not ref:
            return {"error": "Item não encontrado nesta lista"}, 404
        if restaurante_id is None:
            return {"error": "Restaurante é obrigatório"}, 400
        lista = _get_lista_por_restaurante(lista_id, restaurante_id)
        if not lista:
            return {"error": "Lista não encontrada"}, 404
        if ref.item.restaurante_id != restaurante_id:
            return {"error": "Acesso negado."}, 403

        item_id_anterior = ref.item_id
        nome_anterior = ref.item.nome if ref.item else None
        unidade_anterior = ref.item.unidade if ref.item else None
        quantidade_atual_anterior = ref.quantidade_atual
        quantidade_minima_anterior = ref.quantidade_minima

        # Se o nome foi alterado, buscar ou criar novo item no catálogo
        if 'nome' in data and data['nome']:
            nome_novo = data['nome'].strip()
            
            # Busca case-insensitive por item com o mesmo nome
            item_existente = ListaMaeItem.query.filter(
                ListaMaeItem.restaurante_id == restaurante_id,
                db.func.lower(ListaMaeItem.nome) == nome_novo.lower()
            ).first()
            
            if item_existente:
                # Item já existe no catálogo, apenas atualizar a referência
                if item_existente.id != ref.item_id:
                    # Verificar se já existe uma referência para este item nesta lista
                    ref_existente = ListaItemRef.query.filter_by(
                        lista_id=lista_id,
                        item_id=item_existente.id
                    ).first()
                    
                    if ref_existente:
                        # Já existe referência, deletar a antiga e usar a existente
                        db.session.delete(ref)
                        ref = ref_existente
                    else:
                        # Atualizar o item_id da referência
                        ref.item_id = item_existente.id
            else:
                # Item não existe, criar novo no catálogo
                novo_item = ListaMaeItem(
                    nome=nome_novo,
                    unidade=data.get('unidade', ref.item.unidade),
                    restaurante_id=restaurante_id
                )
                db.session.add(novo_item)
                db.session.flush()  # Gera o ID do novo item
                
                # Atualizar a referência para o novo item
                ref.item_id = novo_item.id

        # Atualizar quantidades na referência
        if 'quantidade_atual' in data:
            ref.quantidade_atual = data['quantidade_atual']
        if 'quantidade_minima' in data:
            ref.quantidade_minima = data['quantidade_minima']
        ref.atualizado_em = brasilia_now()

        # Se unidade foi fornecida (e nome não foi alterado), atualizar no item global
        if 'unidade' in data and 'nome' not in data:
            ref.item.unidade = data['unidade']
            ref.item.atualizado_em = brasilia_now()

        log_event(
            acao="update",
            entidade="lista",
            entidade_id=lista_id,
            mensagem="Item atualizado na lista mãe",
            restaurante_id=restaurante_id,
            metadata={
                "item_id_anterior": item_id_anterior,
                "item_nome_anterior": nome_anterior,
                "item_unidade_anterior": unidade_anterior,
                "item_id": ref.item_id,
                "item_nome": ref.item.nome if ref.item else None,
                "item_unidade": ref.item.unidade if ref.item else None,
                "quantidade_atual_anterior": quantidade_atual_anterior,
                "quantidade_minima_anterior": quantidade_minima_anterior,
                "quantidade_atual": ref.quantidade_atual,
                "quantidade_minima": ref.quantidade_minima
            }
        )
        db.session.commit()
        return {
            "id": ref.item_id,
            "nome": ref.item.nome if ref.item else None,
            "unidade": ref.item.unidade if ref.item else None,
            "quantidade_atual": ref.quantidade_atual,
            "quantidade_minima": ref.quantidade_minima,
            "pedido": ref.get_pedido(),
            "criado_em": ref.criado_em.isoformat() if ref.criado_em else None,
            "atualizado_em": ref.atualizado_em.isoformat() if ref.atualizado_em else None
        }, 200
        
    except Exception as e:
        import traceback
        from flask import current_app

        current_app.logger.error(f"[editar_item_lista_mae] Erro ao editar item {item_id}:")
        current_app.logger.error(f"  Tipo: {type(e).__name__}")
        current_app.logger.error(f"  Mensagem: {str(e)}")
        current_app.logger.error(f"  Traceback: {traceback.format_exc()}")

        db.session.rollback()
        return {"error": f"{type(e).__name__}: {str(e)}"}, 500


def deletar_item_lista_mae(lista_id, item_id, restaurante_id=None):
    """
    Remove a referência de um item de uma lista específica.
    O item permanece no catálogo global para uso em outras listas.
    """
    try:
        ref = ListaItemRef.query.filter_by(
            lista_id=lista_id,
            item_id=item_id
        ).first()

        if not ref:
            return {"error": "Item não encontrado nesta lista"}, 404
        if restaurante_id is None:
            return {"error": "Restaurante é obrigatório"}, 400
        lista = _get_lista_por_restaurante(lista_id, restaurante_id)
        if not lista:
            return {"error": "Lista não encontrada"}, 404
        if ref.item.restaurante_id != restaurante_id:
            return {"error": "Acesso negado."}, 403

        log_event(
            acao="delete",
            entidade="lista",
            entidade_id=lista_id,
            mensagem="Item removido da lista mãe",
            restaurante_id=restaurante_id,
            metadata={
                "item_id": ref.item_id,
                "item_nome": ref.item.nome if ref.item else None
            }
        )
        db.session.delete(ref)
        db.session.commit()
        return {"message": "Item removido da lista com sucesso"}, 200
    except Exception as e:
        db.session.rollback()
        return {"error": str(e)}, 500


def obter_lista_mae(lista_id, restaurante_id=None):
    """
    Retorna a lista com todos os seus itens via tabela de referência.
    Os itens vêm do catálogo global com quantidades específicas desta lista.
    """
    from flask import current_app

    try:
        lista = _get_lista_por_restaurante(lista_id, restaurante_id)
        if not lista:
            return {"error": "Lista não encontrada"}, 404

        # Query única com JOIN — evita N queries lazy por item
        refs_query = (
            db.session.query(
                ListaItemRef.quantidade_atual,
                ListaItemRef.quantidade_minima,
                ListaItemRef.usa_threshold,
                ListaItemRef.quantidade_por_fardo,
                ListaItemRef.criado_em,
                ListaItemRef.atualizado_em,
                ListaMaeItem.id.label('item_id'),
                ListaMaeItem.nome,
                ListaMaeItem.unidade,
            )
            .join(ListaMaeItem, ListaItemRef.item_id == ListaMaeItem.id)
            .filter(ListaItemRef.lista_id == lista_id)
        )
        if restaurante_id is not None:
            refs_query = refs_query.filter(ListaMaeItem.restaurante_id == restaurante_id)
        refs = refs_query.all()

        # Montar dados dos itens com quantidades (get_pedido inline — sem acesso lazy)
        itens_data = []
        for row in refs:
            pedido = row.quantidade_por_fardo if row.quantidade_atual < row.quantidade_minima else 0
            itens_data.append({
                "id": row.item_id,
                "nome": row.nome,
                "unidade": row.unidade,
                "quantidade_atual": row.quantidade_atual,
                "quantidade_minima": row.quantidade_minima,
                "pedido": pedido,
                "usa_threshold": row.usa_threshold,
                "quantidade_por_fardo": row.quantidade_por_fardo,
                "criado_em": row.criado_em.isoformat() if row.criado_em else None,
                "atualizado_em": row.atualizado_em.isoformat() if row.atualizado_em else None,
            })

        # Buscar fornecedores atribuídos à lista
        fornecedores_data = []
        for fornecedor in lista.fornecedores:
            if restaurante_id is not None and fornecedor.restaurante_id != restaurante_id:
                continue
            fornecedores_data.append({
                "id": fornecedor.id,
                "nome": fornecedor.nome,
                "contato": fornecedor.contato,
                "meio_envio": fornecedor.meio_envio,
                "responsavel": fornecedor.responsavel,
                "observacao": fornecedor.observacao,
            })

        result = {
            "lista_id": lista.id,
            "lista_nome": lista.nome,
            "lista_descricao": lista.descricao,
            "data_criacao": lista.data_criacao.isoformat(),
            "fornecedores": fornecedores_data,
            "itens": itens_data,
            "total_itens": len(itens_data)
        }

        return result, 200
    except Exception as e:
        current_app.logger.error(f"[obter_lista_mae] ERRO: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500


def atribuir_fornecedor_lista_mae(lista_id, data, restaurante_id=None):
    """
    Atribui um fornecedor a múltiplos itens de uma lista e gera Pedidos.

    Recebe:
        lista_id: ID da Lista
        data: {
            "item_ids": [1, 2, 3],  # IDs dos itens globais
            "fornecedor_id": 5      # ID do fornecedor
        }

    Retorna:
        Lista de pedidos criados
    """
    try:
        item_ids = data.get('item_ids', [])
        fornecedor_id = data.get('fornecedor_id')

        if not item_ids or not fornecedor_id:
            return {"error": "item_ids e fornecedor_id são obrigatórios"}, 400

        lista = db.session.get(Lista, lista_id)
        if not lista:
            return {"error": "Lista não encontrada"}, 404

        if restaurante_id is None:
            return {"error": "Restaurante é obrigatório"}, 400

        fornecedor = _get_fornecedor_por_id(fornecedor_id, restaurante_id)
        if not fornecedor:
            return {"error": "Fornecedor não encontrado"}, 404

        # Buscar referências de itens vinculados a esta lista
        refs = ListaItemRef.query.join(ListaMaeItem).filter(
            ListaItemRef.lista_id == lista_id,
            ListaItemRef.item_id.in_(item_ids),
            ListaMaeItem.restaurante_id == restaurante_id
        ).all()

        if len(refs) != len(item_ids):
            return {"error": "Um ou mais itens não encontrados na lista"}, 404

        pedidos_criados = []
        usuario_id = get_current_user_id()

        for ref in refs:
            # Calcular quantidade do pedido a partir da referência
            quantidade_pedido = ref.get_pedido()

            if quantidade_pedido > 0:
                pedido_existe = Pedido.query.filter_by(
                    item_id=ref.item_id,
                    fornecedor_id=fornecedor_id,
                    status=PedidoStatus.PENDENTE
                ).first()

                if not pedido_existe:
                    novo_pedido = Pedido(
                        item_id=ref.item_id,
                        fornecedor_id=fornecedor_id,
                        quantidade_solicitada=quantidade_pedido,
                        usuario_id=usuario_id,
                        status=PedidoStatus.PENDENTE,
                        data_pedido=brasilia_now()
                    )
                    db.session.add(novo_pedido)
                    pedidos_criados.append(novo_pedido)

        log_event(
            acao="update",
            entidade="lista",
            entidade_id=lista_id,
            mensagem="Fornecedor atribuído à lista mãe",
            restaurante_id=restaurante_id,
            metadata={
                "fornecedor_id": fornecedor_id,
                "item_ids": item_ids,
                "pedidos_criados": len(pedidos_criados)
            }
        )
        db.session.commit()

        return {
            "message": f"{len(pedidos_criados)} pedido(s) criado(s)",
            "pedidos": [p.to_dict() for p in pedidos_criados],
            "total_pedidos": len(pedidos_criados)
        }, 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[ATRIBUIR FORNECEDOR] ERRO: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500

def copiar_itens_entre_listas(lista_origem_id, data, restaurante_id=None):
    """
    Copia itens de uma lista para outra (ou cria nova lista).
    Mantém os itens na lista origem e cria referências na lista destino.
    """
    try:
        print(f"[copiar_itens_entre_listas] Iniciando cópia da lista #{lista_origem_id}")
        
        item_ids = data.get('item_ids', [])
        lista_destino_id = data.get('lista_destino_id')
        nome_nova_lista = data.get('nome_nova_lista')
        area_id = data.get('area_id')
        
        if not item_ids:
            return {"error": "Nenhum item selecionado"}, 400
        
        # Validar lista origem
        lista_origem = _get_lista_por_restaurante(lista_origem_id, restaurante_id)
        if not lista_origem:
            return {"error": "Lista de origem não encontrada"}, 404
        
        # Determinar lista destino (existente ou criar nova)
        if lista_destino_id:
            lista_destino = _get_lista_por_restaurante(lista_destino_id, restaurante_id)
            if not lista_destino:
                return {"error": "Lista de destino não encontrada"}, 404
        elif nome_nova_lista and area_id:
            if restaurante_id is None:
                return {"error": "Restaurante é obrigatório."}, 400
            # Criar nova lista
            nova_lista = Lista(nome=nome_nova_lista, area_id=area_id, restaurante_id=restaurante_id)
            db.session.add(nova_lista)
            db.session.flush()
            lista_destino = nova_lista
            print(f"[copiar_itens_entre_listas] Nova lista criada: #{lista_destino.id}")
        else:
            return {"error": "Forneça lista_destino_id ou (nome_nova_lista + area_id)"}, 400
        
        # Buscar itens da lista origem
        refs_origem = ListaItemRef.query.filter(
            ListaItemRef.lista_id == lista_origem_id,
            ListaItemRef.item_id.in_(item_ids)
        ).all()
        
        if not refs_origem:
            return {"error": "Nenhum item encontrado na lista de origem"}, 404
        
        print(f"[copiar_itens_entre_listas] Encontrados {len(refs_origem)} itens na origem")
        
        # Buscar itens já existentes na lista destino
        refs_destino_existentes = ListaItemRef.query.filter_by(
            lista_id=lista_destino.id
        ).all()
        itens_existentes_ids = {ref.item_id for ref in refs_destino_existentes}
        
        itens_copiados = 0
        itens_ignorados = 0
        itens_ignorados_lista = []
        
        for ref_origem in refs_origem:
            if ref_origem.item_id in itens_existentes_ids:
                # Item já existe na lista destino - ignorar
                item = repositories.get_by_id(ListaMaeItem, ref_origem.item_id)
                itens_ignorados += 1
                if item:
                    itens_ignorados_lista.append(item.nome)
                print(f"[copiar_itens_entre_listas] Item #{ref_origem.item_id} ignorado (já existe)")
                continue
            
            # Copiar item (criar nova referência)
            nova_ref = ListaItemRef(
                lista_id=lista_destino.id,
                item_id=ref_origem.item_id,
                quantidade_minima=ref_origem.quantidade_minima,
                quantidade_atual=ref_origem.quantidade_atual
            )
            db.session.add(nova_ref)
            itens_copiados += 1
            print(f"[copiar_itens_entre_listas] Item #{ref_origem.item_id} copiado")

        log_event(
            acao="update",
            entidade="lista",
            entidade_id=lista_destino.id,
            mensagem="Itens copiados entre listas",
            restaurante_id=restaurante_id,
            metadata={
                "lista_origem_id": lista_origem_id,
                "lista_destino_id": lista_destino.id,
                "itens_copiados": itens_copiados,
                "itens_ignorados": itens_ignorados,
                "item_ids": item_ids
            }
        )
        db.session.commit()
        
        print(f"[copiar_itens_entre_listas] Cópia concluída: {itens_copiados} copiados, {itens_ignorados} ignorados")
        
        message = f"{itens_copiados} item(ns) copiado(s) com sucesso"
        if itens_ignorados > 0:
            message += f". {itens_ignorados} item(ns) ignorado(s) por já existirem"
        
        return {
            "message": message,
            "itens_copiados": itens_copiados,
            "itens_ignorados": itens_ignorados,
            "itens_ignorados_lista": itens_ignorados_lista,
            "lista_destino_id": lista_destino.id,
            "lista_destino_nome": lista_destino.nome
        }, 200
        
    except Exception as e:
        db.session.rollback()
        print(f"[copiar_itens_entre_listas] EXCEÇÃO: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": f"Erro ao copiar itens: {str(e)}"}, 500


def mover_itens_entre_listas(lista_origem_id, data, restaurante_id=None):
    """
    Move itens de uma lista para outra (ou cria nova lista).
    Remove os itens da lista origem e cria referências na lista destino.
    """
    try:
        print(f"[mover_itens_entre_listas] Iniciando movimentação da lista #{lista_origem_id}")
        
        item_ids = data.get('item_ids', [])
        lista_destino_id = data.get('lista_destino_id')
        nome_nova_lista = data.get('nome_nova_lista')
        area_id = data.get('area_id')
        
        if not item_ids:
            return {"error": "Nenhum item selecionado"}, 400
        
        # Validar lista origem
        lista_origem = _get_lista_por_restaurante(lista_origem_id, restaurante_id)
        if not lista_origem:
            return {"error": "Lista de origem não encontrada"}, 404
        
        # Determinar lista destino (existente ou criar nova)
        if lista_destino_id:
            lista_destino = _get_lista_por_restaurante(lista_destino_id, restaurante_id)
            if not lista_destino:
                return {"error": "Lista de destino não encontrada"}, 404
        elif nome_nova_lista and area_id:
            if restaurante_id is None:
                return {"error": "Restaurante é obrigatório."}, 400
            # Criar nova lista
            nova_lista = Lista(nome=nome_nova_lista, area_id=area_id, restaurante_id=restaurante_id)
            db.session.add(nova_lista)
            db.session.flush()
            lista_destino = nova_lista
            print(f"[mover_itens_entre_listas] Nova lista criada: #{lista_destino.id}")
        else:
            return {"error": "Forneça lista_destino_id ou (nome_nova_lista + area_id)"}, 400
        
        # Buscar itens da lista origem
        refs_origem = ListaItemRef.query.filter(
            ListaItemRef.lista_id == lista_origem_id,
            ListaItemRef.item_id.in_(item_ids)
        ).all()
        
        if not refs_origem:
            return {"error": "Nenhum item encontrado na lista de origem"}, 404
        
        print(f"[mover_itens_entre_listas] Encontrados {len(refs_origem)} itens na origem")
        
        # Buscar itens já existentes na lista destino
        refs_destino_existentes = ListaItemRef.query.filter_by(
            lista_id=lista_destino.id
        ).all()
        itens_existentes_ids = {ref.item_id for ref in refs_destino_existentes}
        
        itens_movidos = 0
        itens_ignorados = 0
        itens_ignorados_lista = []
        
        for ref_origem in refs_origem:
            if ref_origem.item_id in itens_existentes_ids:
                # Item já existe na lista destino - ignorar
                item = repositories.get_by_id(ListaMaeItem, ref_origem.item_id)
                itens_ignorados += 1
                if item:
                    itens_ignorados_lista.append(item.nome)
                print(f"[mover_itens_entre_listas] Item #{ref_origem.item_id} ignorado (já existe)")
                continue
            
            # Mover item (criar nova referência e deletar antiga)
            nova_ref = ListaItemRef(
                lista_id=lista_destino.id,
                item_id=ref_origem.item_id,
                quantidade_minima=ref_origem.quantidade_minima,
                quantidade_atual=ref_origem.quantidade_atual
            )
            db.session.add(nova_ref)
            db.session.delete(ref_origem)
            itens_movidos += 1
            print(f"[mover_itens_entre_listas] Item #{ref_origem.item_id} movido")

        log_event(
            acao="update",
            entidade="lista",
            entidade_id=lista_destino.id,
            mensagem="Itens movidos entre listas",
            restaurante_id=restaurante_id,
            metadata={
                "lista_origem_id": lista_origem_id,
                "lista_destino_id": lista_destino.id,
                "itens_movidos": itens_movidos,
                "itens_ignorados": itens_ignorados,
                "item_ids": item_ids
            }
        )
        db.session.commit()
        
        print(f"[mover_itens_entre_listas] Movimentação concluída: {itens_movidos} movidos, {itens_ignorados} ignorados")
        
        message = f"{itens_movidos} item(ns) movido(s) com sucesso"
        if itens_ignorados > 0:
            message += f". {itens_ignorados} item(ns) ignorado(s) por já existirem"
        
        return {
            "message": message,
            "itens_movidos": itens_movidos,
            "itens_ignorados": itens_ignorados,
            "itens_ignorados_lista": itens_ignorados_lista,
            "lista_destino_id": lista_destino.id,
            "lista_destino_nome": lista_destino.nome
        }, 200
        
    except Exception as e:
        db.session.rollback()
        print(f"[mover_itens_entre_listas] EXCEÇÃO: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": f"Erro ao mover itens: {str(e)}"}, 500

def importar_items_em_lote(lista_id, data, restaurante_id=None):
    """
    Importa múltiplos itens em lote para uma lista.
    Itens são adicionados ao catálogo global e vinculados à lista.

    Recebe:
        lista_id: ID da Lista
        data: {
            "nomes": ["Alga Nori", "Arroz Grao Curto", "BAO com vegetais", ...]
        }

    Retorna:
        Quantidade de itens importados
    """
    try:
        nomes = data.get('nomes', [])

        if not nomes or not isinstance(nomes, list):
            return {"error": "nomes deve ser um array de strings"}, 400

        lista = _get_lista_por_restaurante(lista_id, restaurante_id)
        if not lista:
            return {"error": "Lista não encontrada"}, 404
        if restaurante_id is None:
            return {"error": "Restaurante é obrigatório"}, 400

        items_criados = 0
        items_vinculados = 0
        items_duplicados = 0

        for nome in nomes:
            nome_limpo = nome.strip()

            if len(nome_limpo) < 2:
                continue

            # Buscar item no catálogo global (case-insensitive)
            item = ListaMaeItem.query.filter(
                ListaMaeItem.restaurante_id == restaurante_id,
                func.lower(ListaMaeItem.nome) == func.lower(nome_limpo)
            ).first()

            if not item:
                # Criar novo item no catálogo global
                item = ListaMaeItem(
                    nome=nome_limpo,
                    unidade='un',
                    restaurante_id=restaurante_id
                )
                db.session.add(item)
                db.session.flush()
                items_criados += 1

            # Verificar se já está vinculado a esta lista
            ref_existe = ListaItemRef.query.filter_by(
                lista_id=lista_id,
                item_id=item.id
            ).first()

            if ref_existe:
                items_duplicados += 1
                continue

            # Criar referência
            ref = ListaItemRef(
                lista_id=lista_id,
                item_id=item.id,
                quantidade_atual=0,
                quantidade_minima=1.0
            )
            db.session.add(ref)
            items_vinculados += 1

        log_event(
            acao="update",
            entidade="lista",
            entidade_id=lista_id,
            mensagem="Importação em lote para lista mãe",
            restaurante_id=restaurante_id,
            metadata={
                "items_vinculados": items_vinculados,
                "items_criados": items_criados,
                "items_duplicados": items_duplicados
            }
        )
        db.session.commit()

        return {
            "message": f"{items_vinculados} item(ns) vinculado(s) à lista, {items_criados} novo(s) no catálogo",
            "items_vinculados": items_vinculados,
            "items_criados": items_criados,
            "items_duplicados": items_duplicados
        }, 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[IMPORTAR ITEMS] ERRO: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500


def get_lista_colaborador(user_id, lista_id):
    """Retorna dados da lista para um colaborador."""
    # Verificar se o usuário tem acesso a esta lista
    usuario = repositories.get_by_id(Usuario, user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    lista = repositories.get_by_id(Lista, lista_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    if lista.restaurante_id != usuario.restaurante_id and usuario.role != UserRoles.SUPER_ADMIN:
        return {"error": "Acesso negado. Esta lista não pertence ao seu restaurante."}, 403

    # Verificar se o colaborador está atribuído a esta lista (Admin/SuperAdmin têm acesso automático)
    if not _is_admin_or_super_admin(usuario):
        if lista not in usuario.listas_atribuidas:
            return {"error": "Acesso negado. Esta lista não foi atribuída a você."}, 403

    return lista.to_dict(), 200


def get_estoque_lista_colaborador(user_id, lista_id):
    """Retorna itens da lista via ListaItemRef (sem usar tabela Estoque)."""
    # Verificar se o usuário tem acesso a esta lista
    usuario = repositories.get_by_id(Usuario, user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    lista = repositories.get_by_id(Lista, lista_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    if lista.restaurante_id != usuario.restaurante_id and usuario.role != UserRoles.SUPER_ADMIN:
        return {"error": "Acesso negado. Esta lista não pertence ao seu restaurante."}, 403

    # Verificar se o colaborador está atribuído a esta lista (Admin/SuperAdmin têm acesso automático)
    if not _is_admin_or_super_admin(usuario):
        if lista not in usuario.listas_atribuidas:
            return {"error": "Acesso negado. Esta lista não foi atribuída a você."}, 403

    # Buscar itens da lista via ListaItemRef com EAGER LOADING (otimização)
    refs = ListaItemRef.query.options(
        db.joinedload(ListaItemRef.item)  # Carrega item junto, evita N+1 queries
    ).filter_by(lista_id=lista_id).all()

    # Preparar resposta com dados do item
    itens_data = []
    for ref in refs:
        # Pular itens sem quantidade mínima definida
        if ref.quantidade_minima <= 0:
            continue
        
        item_dict = {
            "id": ref.item_id,  # Usar item_id como identificador (compatibilidade com frontend)
            "item_id": ref.item_id,
            "lista_id": ref.lista_id,
            "quantidade_atual": float(ref.quantidade_atual),
            "quantidade_minima": float(ref.quantidade_minima),
            "pedido": float(ref.get_pedido()),
            "usa_threshold": ref.usa_threshold,
            "quantidade_por_fardo": float(ref.quantidade_por_fardo),
            "item": {
                "id": ref.item.id,
                "nome": ref.item.nome,
                "unidade_medida": ref.item.unidade  # ListaMaeItem usa 'unidade', não 'unidade_medida'
            }
        }
        itens_data.append(item_dict)

    return itens_data, 200


def get_estoque_lista_admin(lista_id, restaurante_id=None):
    """
    Retorna itens da lista via ListaItemRef para admin (sem verificação de atribuição).
    Mesmo formato do colaborador, usado para editar submissões.
    """
    lista = _get_lista_por_restaurante(lista_id, restaurante_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    # Buscar itens da lista via ListaItemRef com EAGER LOADING
    refs = ListaItemRef.query.options(
        db.joinedload(ListaItemRef.item)
    ).filter_by(lista_id=lista_id).all()

    # Preparar resposta com dados do item
    itens_data = []
    for ref in refs:
        # Pular itens sem quantidade mínima definida
        if ref.quantidade_minima <= 0:
            continue
        
        item_dict = {
            "id": ref.item_id,
            "item_id": ref.item_id,
            "lista_id": ref.lista_id,
            "quantidade_atual": float(ref.quantidade_atual),
            "quantidade_minima": float(ref.quantidade_minima),
            "pedido": float(ref.get_pedido()),
            "usa_threshold": ref.usa_threshold,
            "quantidade_por_fardo": float(ref.quantidade_por_fardo),
            "item": {
                "id": ref.item.id,
                "nome": ref.item.nome,
                "unidade_medida": ref.item.unidade
            }
        }
        itens_data.append(item_dict)

    return itens_data, 200


def update_estoque_colaborador(user_id, estoque_id, data):
    """
    Atualiza quantidade_atual de um item via ListaItemRef.
    NOTA: estoque_id agora é interpretado como item_id (para compatibilidade).
    """
    # Verificar se o usuário existe
    usuario = repositories.get_by_id(Usuario, user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    # Validar que apenas quantidade_atual está sendo atualizado
    if "quantidade_atual" not in data:
        return {"error": "Dados inválidos. Forneça quantidade_atual."}, 400

    # Validar que o valor é um número válido
    try:
        quantidade_atual = float(data.get("quantidade_atual", 0))
        if quantidade_atual < 0:
            return {"error": "A quantidade não pode ser negativa."}, 400
    except (ValueError, TypeError):
        return {"error": "Quantidade deve ser um número válido."}, 400

    # Buscar ListaItemRef - precisa de lista_id também
    # Por ora, vamos buscar por item_id em todas as listas do colaborador
    item_id = estoque_id  # estoque_id é na verdade o item_id
    
    # Buscar qual lista contém este item e está atribuída ao colaborador
    ref = None
    for lista in usuario.listas_atribuidas:
        ref = ListaItemRef.query.filter_by(
            lista_id=lista.id,
            item_id=item_id
        ).first()
        if ref:
            break
    
    if not ref:
        return {"error": "Item não encontrado nas suas listas."}, 404

    last_known_update = data.get("last_known_update")
    if last_known_update:
        client_dt = _parse_iso_datetime(last_known_update)
        if client_dt is None:
            return {"error": "Timestamp invalido."}, 400
        server_dt = _normalize_datetime(ref.atualizado_em)
        client_dt = _normalize_datetime(client_dt)
        if server_dt and client_dt and server_dt > client_dt:
            return {
                "error": "Conflito de versao",
                "server_data": ref.to_dict()
            }, 409

    try:
        # Atualizar apenas quantidade_atual
        ref.quantidade_atual = quantidade_atual
        ref.atualizado_em = brasilia_now()
        
        db.session.commit()

        return {
            "message": "Estoque atualizado com sucesso.",
            "estoque": {
                "id": ref.item_id,
                "quantidade_atual": float(ref.quantidade_atual),
                "quantidade_minima": float(ref.quantidade_minima),
                "pedido": float(ref.get_pedido())
            }
        }, 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[UPDATE ESTOQUE COLABORADOR] ERRO: {type(e).__name__}: {str(e)}")
        return {"error": str(e)}, 500


# --- CSV Export/Import para Fornecedores ---

def exportar_fornecedores_csv(restaurante_id=None):
    """Exporta fornecedores em formato CSV."""
    try:
        query = Fornecedor.query
        if restaurante_id is not None:
            query = query.filter_by(restaurante_id=restaurante_id)
        fornecedores = query.all()

        # Construir CSV
        csv_lines = ["Nome,Contato,Meio de Envio,Responsável,Observação"]

        for fornecedor in fornecedores:
            contato = (fornecedor.contato or "").replace(",", ";").replace("\n", " ")
            meio_envio = (fornecedor.meio_envio or "").replace(",", ";").replace("\n", " ")
            responsavel = (fornecedor.responsavel or "").replace(",", ";").replace("\n", " ")
            observacao = (fornecedor.observacao or "").replace(",", ";").replace("\n", " ")

            csv_lines.append(f'"{fornecedor.nome}","{contato}","{meio_envio}","{responsavel}","{observacao}"')

        csv_content = "\n".join(csv_lines)
        return {"csv": csv_content, "filename": "fornecedores.csv"}, 200
    except Exception as e:
        return {"error": str(e)}, 500


def importar_fornecedores_csv(csv_content, restaurante_id=None):
    """Importa fornecedores a partir de conteúdo CSV."""
    try:
        import csv
        import io

        if restaurante_id is None:
            return {"error": "Restaurante é obrigatório para importação."}, 400

        # Parse CSV
        csv_reader = csv.DictReader(io.StringIO(csv_content))

        fornecedores_criados = 0
        fornecedores_duplicados = 0
        erros = []

        for row in csv_reader:
            nome = row.get('Nome', '').strip()
            if not nome:
                continue

            # Verificar se já existe
            fornecedor_query = Fornecedor.query.filter_by(nome=nome)
            if restaurante_id is not None:
                fornecedor_query = fornecedor_query.filter_by(restaurante_id=restaurante_id)
            fornecedor_existe = fornecedor_query.first()
            if fornecedor_existe:
                fornecedores_duplicados += 1
                continue

            # Criar novo fornecedor
            novo_fornecedor = Fornecedor(
                nome=nome,
                contato=row.get('Contato', '').strip(),
                meio_envio=row.get('Meio de Envio', '').strip(),
                responsavel=row.get('Responsável', '').strip(),
                observacao=row.get('Observação', '').strip(),
                restaurante_id=restaurante_id
            )
            db.session.add(novo_fornecedor)
            fornecedores_criados += 1

        db.session.commit()

        return {
            "message": f"{fornecedores_criados} fornecedor(es) criado(s)",
            "fornecedores_criados": fornecedores_criados,
            "fornecedores_duplicados": fornecedores_duplicados
        }, 201

    except Exception as e:
        db.session.rollback()
        return {"error": str(e)}, 500

def clear_database_except_users(user_id, data):
    """
    Limpa todas as tabelas do banco de dados exceto a tabela de usuários.
    Requer que o usuário forneça sua senha para confirmação.
    """
    try:
        # Validação de entrada
        senha = data.get('senha')
        if not senha:
            return {"error": "Senha é obrigatória para confirmar a operação"}, 400

        # Busca o usuário e verifica a senha
        usuario = db.session.get(Usuario, user_id)
        if not usuario:
            return {"error": "Usuário não encontrado"}, 404

        if not check_password_hash(usuario.senha_hash, senha):
            return {"error": "Senha incorreta"}, 401

        # Verifica se é SUPER_ADMIN
        if usuario.role != UserRoles.SUPER_ADMIN:
            return {"error": "Apenas SUPER_ADMIN pode limpar o banco de dados"}, 403

        # Começa a limpeza das tabelas na ordem correta (respeitando foreign keys)
        print("🗑️ Iniciando limpeza do banco de dados...")

        # 1. Limpar tabelas de associação (many-to-many)
        db.session.execute(db.text('DELETE FROM fornecedor_lista'))
        db.session.execute(db.text('DELETE FROM lista_colaborador'))
        print("✅ Tabelas de associação limpas")

        # 2. Limpar tabelas dependentes
        CotacaoItem.query.delete()
        print("✅ Itens de cotação removidos")

        Cotacao.query.delete()
        print("✅ Cotações removidas")

        Pedido.query.delete()
        print("✅ Pedidos removidos")

        Estoque.query.delete()
        print("✅ Estoques removidos")

        ListaMaeItem.query.delete()
        print("✅ Itens da lista mãe removidos")

        # 3. Limpar listas
        Lista.query.delete()
        print("✅ Listas removidas")

        # 4. Limpar itens
        Item.query.delete()
        print("✅ Itens removidos")

        # 5. Limpar fornecedores
        Fornecedor.query.delete()
        print("✅ Fornecedores removidos")

        # 6. Limpar áreas
        Area.query.delete()
        print("✅ Áreas removidas")

        # Commit das alterações
        db.session.commit()
        print("✨ Banco de dados limpo com sucesso!")

        return {
            "message": "Banco de dados limpo com sucesso! Apenas usuários foram mantidos.",
            "cleared_tables": [
                "fornecedor_lista",
                "lista_colaborador",
                "cotacao_itens",
                "cotacoes",
                "pedidos",
                "estoques",
                "lista_mae_itens",
                "listas",
                "itens",
                "fornecedores",
                "areas"
            ]
        }, 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao limpar banco de dados: {str(e)}")
        return {"error": f"Erro ao limpar banco de dados: {str(e)}"}, 500

def populate_database_with_mock_data(restaurante_id=None):
    """
    Popula o banco de dados com dados fictícios para teste.
    NÃO vincula colaboradores às listas - admin faz isso manualmente.
    """
    try:
        import random
        from datetime import timedelta

        print("🌱 Iniciando população do banco de dados...")
        restaurante_id = restaurante_id or _get_restaurante_padrao_id()
        if not restaurante_id:
            return {"error": "Restaurante não encontrado."}, 404

        # 1. Criar Áreas
        print("📍 Criando áreas...")
        areas_nomes = ["Cozinha", "Almoxarifado", "Limpeza", "Manutenção", "Escritório"]
        areas = []
        for nome in areas_nomes:
            if not Area.query.filter_by(nome=nome).first():
                area = Area(nome=nome)
                db.session.add(area)
                areas.append(area)
        db.session.commit()
        areas = Area.query.all()
        print(f"✅ {len(areas)} áreas no banco")

        # 2. Criar Fornecedores
        print("🏪 Criando fornecedores...")
        fornecedores_data = [
            {
                "nome": "Distribuidora ABC",
                "contato": "(11) 1234-5678",
                "meio_envio": "WhatsApp",
                "responsavel": "João Silva",
                "observacao": "Fornecedor principal de alimentos"
            },
            {
                "nome": "Produtos Limpeza XYZ",
                "contato": "(11) 8765-4321",
                "meio_envio": "Email",
                "responsavel": "Maria Santos",
                "observacao": "Produtos de limpeza e higiene"
            },
            {
                "nome": "Papelaria Moderna",
                "contato": "(11) 5555-9999",
                "meio_envio": "WhatsApp",
                "responsavel": "Pedro Costa",
                "observacao": "Material de escritório"
            },
            {
                "nome": "Ferramentas e Cia",
                "contato": "(11) 3333-7777",
                "meio_envio": "Telefone",
                "responsavel": "Carlos Oliveira",
                "observacao": "Ferramentas e material de manutenção"
            }
        ]

        fornecedores = []
        for forn_data in fornecedores_data:
            if not Fornecedor.query.filter_by(nome=forn_data["nome"], restaurante_id=restaurante_id).first():
                fornecedor = Fornecedor(**forn_data, restaurante_id=restaurante_id)
                db.session.add(fornecedor)
                fornecedores.append(fornecedor)
        db.session.commit()
        fornecedores = Fornecedor.query.filter_by(restaurante_id=restaurante_id).all()
        print(f"✅ {len(fornecedores)} fornecedores no banco")

        # 3. Criar Itens
        print("📦 Criando itens...")
        itens_data = [
            # Itens da Distribuidora ABC
            {"nome": "Arroz", "unidade_medida": "Kg", "fornecedor_nome": "Distribuidora ABC"},
            {"nome": "Feijão", "unidade_medida": "Kg", "fornecedor_nome": "Distribuidora ABC"},
            {"nome": "Óleo", "unidade_medida": "Litro", "fornecedor_nome": "Distribuidora ABC"},
            {"nome": "Açúcar", "unidade_medida": "Kg", "fornecedor_nome": "Distribuidora ABC"},
            {"nome": "Sal", "unidade_medida": "Kg", "fornecedor_nome": "Distribuidora ABC"},
            # Itens da Produtos Limpeza XYZ
            {"nome": "Detergente", "unidade_medida": "Unidade", "fornecedor_nome": "Produtos Limpeza XYZ"},
            {"nome": "Sabão em Pó", "unidade_medida": "Kg", "fornecedor_nome": "Produtos Limpeza XYZ"},
            {"nome": "Desinfetante", "unidade_medida": "Litro", "fornecedor_nome": "Produtos Limpeza XYZ"},
            {"nome": "Álcool Gel", "unidade_medida": "Litro", "fornecedor_nome": "Produtos Limpeza XYZ"},
            # Itens da Papelaria Moderna
            {"nome": "Papel A4", "unidade_medida": "Resma", "fornecedor_nome": "Papelaria Moderna"},
            {"nome": "Caneta Azul", "unidade_medida": "Unidade", "fornecedor_nome": "Papelaria Moderna"},
            {"nome": "Grampeador", "unidade_medida": "Unidade", "fornecedor_nome": "Papelaria Moderna"},
            # Itens da Ferramentas e Cia
            {"nome": "Martelo", "unidade_medida": "Unidade", "fornecedor_nome": "Ferramentas e Cia"},
            {"nome": "Chave de Fenda", "unidade_medida": "Unidade", "fornecedor_nome": "Ferramentas e Cia"},
            {"nome": "Fita Isolante", "unidade_medida": "Rolo", "fornecedor_nome": "Ferramentas e Cia"},
        ]

        itens = []
        for item_data in itens_data:
            if not Item.query.filter_by(nome=item_data["nome"]).first():
                fornecedor = Fornecedor.query.filter_by(nome=item_data["fornecedor_nome"], restaurante_id=restaurante_id).first()
                if fornecedor:
                    item = Item(
                        nome=item_data["nome"],
                        unidade_medida=item_data["unidade_medida"],
                        fornecedor_id=fornecedor.id
                    )
                    db.session.add(item)
                    itens.append(item)
        db.session.commit()
        itens = Item.query.all()
        print(f"✅ {len(itens)} itens no banco")

        # 4. Criar Listas de Compras (SEM vincular colaboradores)
        print("📋 Criando listas de compras...")
        listas_data = [
            {
                "nome": "Lista Mensal - Alimentos",
                "descricao": "Lista mensal de compras de alimentos"
            },
            {
                "nome": "Lista Semanal - Limpeza",
                "descricao": "Lista semanal de produtos de limpeza"
            },
            {
                "nome": "Lista Escritório",
                "descricao": "Material de escritório trimestral"
            }
        ]

        listas = []
        for lista_data in listas_data:
            if not Lista.query.filter_by(nome=lista_data["nome"], restaurante_id=restaurante_id).first():
                lista = Lista(restaurante_id=restaurante_id, **lista_data)
                db.session.add(lista)
                listas.append(lista)
        db.session.commit()
        listas = Lista.query.all()

        # Associar fornecedores às listas
        lista_alimentos = Lista.query.filter_by(nome="Lista Mensal - Alimentos", restaurante_id=restaurante_id).first()
        lista_limpeza = Lista.query.filter_by(nome="Lista Semanal - Limpeza", restaurante_id=restaurante_id).first()
        lista_escritorio = Lista.query.filter_by(nome="Lista Escritório", restaurante_id=restaurante_id).first()

        forn_abc = Fornecedor.query.filter_by(nome="Distribuidora ABC", restaurante_id=restaurante_id).first()
        forn_limpeza = Fornecedor.query.filter_by(nome="Produtos Limpeza XYZ", restaurante_id=restaurante_id).first()
        forn_papelaria = Fornecedor.query.filter_by(nome="Papelaria Moderna", restaurante_id=restaurante_id).first()

        if lista_alimentos and forn_abc and forn_abc not in lista_alimentos.fornecedores:
            lista_alimentos.fornecedores.append(forn_abc)
        if lista_limpeza and forn_limpeza and forn_limpeza not in lista_limpeza.fornecedores:
            lista_limpeza.fornecedores.append(forn_limpeza)
        if lista_escritorio and forn_papelaria and forn_papelaria not in lista_escritorio.fornecedores:
            lista_escritorio.fornecedores.append(forn_papelaria)

        db.session.commit()
        print(f"✅ {len(listas)} listas no banco (colaboradores NÃO vinculados)")

        # 5. Criar itens da Lista Mãe
        print("📝 Criando itens no catálogo global e vinculando às listas...")
        lista_mae_itens_data = [
            # Lista Mensal - Alimentos
            {"lista_nome": "Lista Mensal - Alimentos", "nome": "Arroz Tipo 1", "unidade": "Kg", "qtd_atual": 50, "qtd_min": 100},
            {"lista_nome": "Lista Mensal - Alimentos", "nome": "Feijão Preto", "unidade": "Kg", "qtd_atual": 30, "qtd_min": 80},
            {"lista_nome": "Lista Mensal - Alimentos", "nome": "Óleo de Soja", "unidade": "Litro", "qtd_atual": 15, "qtd_min": 40},
            # Lista Semanal - Limpeza
            {"lista_nome": "Lista Semanal - Limpeza", "nome": "Detergente Neutro", "unidade": "Unidade", "qtd_atual": 5, "qtd_min": 20},
            {"lista_nome": "Lista Semanal - Limpeza", "nome": "Desinfetante Pinho", "unidade": "Litro", "qtd_atual": 8, "qtd_min": 15},
            # Lista Escritório
            {"lista_nome": "Lista Escritório", "nome": "Papel Sulfite A4", "unidade": "Resma", "qtd_atual": 10, "qtd_min": 30},
        ]

        itens_criados = 0
        refs_criadas = 0
        for item_data in lista_mae_itens_data:
            lista = Lista.query.filter_by(nome=item_data["lista_nome"]).first()
            if lista:
                # Buscar ou criar item no catálogo global
                item = ListaMaeItem.query.filter(
                    ListaMaeItem.restaurante_id == restaurante_id,
                    func.lower(ListaMaeItem.nome) == func.lower(item_data["nome"])
                ).first()

                if not item:
                    item = ListaMaeItem(
                        nome=item_data["nome"],
                        unidade=item_data["unidade"],
                        restaurante_id=restaurante_id
                    )
                    db.session.add(item)
                    db.session.flush()
                    itens_criados += 1

                # Verificar se referência já existe
                ref_existe = ListaItemRef.query.filter_by(
                    lista_id=lista.id,
                    item_id=item.id
                ).first()

                if not ref_existe:
                    ref = ListaItemRef(
                        lista_id=lista.id,
                        item_id=item.id,
                        quantidade_atual=item_data["qtd_atual"],
                        quantidade_minima=item_data["qtd_min"]
                    )
                    db.session.add(ref)
                    refs_criadas += 1

        db.session.commit()
        total_lista_mae = ListaMaeItem.query.filter_by(restaurante_id=restaurante_id).count()
        total_refs = ListaItemRef.query.count()
        print(f"✅ {total_lista_mae} itens no catálogo global, {total_refs} vínculos com listas")

        # 6. Criar Estoques
        print("📊 Criando registros de estoque...")
        estoque_count = 0
        for area in areas[:3]:  # Primeiras 3 áreas
            for item in itens[:10]:  # Primeiros 10 itens
                # Verifica se já existe
                existe = Estoque.query.filter_by(item_id=item.id, area_id=area.id).first()
                if not existe:
                    qtd_atual = random.uniform(5, 50)
                    qtd_minima = random.uniform(20, 100)
                    estoque = Estoque(
                        item_id=item.id,
                        area_id=area.id,
                        quantidade_atual=qtd_atual,
                        quantidade_minima=qtd_minima,
                        pedido=max(qtd_minima - qtd_atual, 0)
                    )
                    db.session.add(estoque)
                    estoque_count += 1

        db.session.commit()
        total_estoques = Estoque.query.count()
        print(f"✅ {total_estoques} registros de estoque no banco")

        # 7. Criar Pedidos
        print("🛒 Criando pedidos...")
        admin = Usuario.query.filter_by(role=UserRoles.ADMIN, restaurante_id=restaurante_id).first()
        pedido_count = 0

        if admin and len(itens) > 0:
            for i in range(10):
                item = random.choice(itens)
                # Verifica se item tem fornecedor
                if item.fornecedor_id:
                    pedido = Pedido(
                        item_id=item.id,
                        fornecedor_id=item.fornecedor_id,
                        quantidade_solicitada=random.uniform(10, 100),
                        data_pedido=brasilia_now() - timedelta(days=random.randint(1, 30)),
                        usuario_id=admin.id,
                        status=random.choice([PedidoStatus.PENDENTE, PedidoStatus.APROVADO, PedidoStatus.REJEITADO])
                    )
                    db.session.add(pedido)
                    pedido_count += 1

            db.session.commit()

        total_pedidos = Pedido.query.count()
        print(f"✅ {total_pedidos} pedidos no banco")

        # 8. Criar Cotações
        print("💰 Criando cotações...")
        cotacao_count = 0
        for fornecedor in fornecedores[:2]:
            cotacao = Cotacao(
                fornecedor_id=fornecedor.id,
                data_cotacao=brasilia_now() - timedelta(days=random.randint(1, 15)),
                status=random.choice([CotacaoStatus.PENDENTE, CotacaoStatus.CONCLUIDA])
            )
            db.session.add(cotacao)
            db.session.commit()

            # Adicionar itens à cotação
            itens_fornecedor = Item.query.filter_by(fornecedor_id=fornecedor.id).all()
            for item in itens_fornecedor[:3]:
                cotacao_item = CotacaoItem(
                    cotacao_id=cotacao.id,
                    item_id=item.id,
                    quantidade=random.uniform(10, 50),
                    preco_unitario=random.uniform(5, 100)
                )
                db.session.add(cotacao_item)
            cotacao_count += 1

        db.session.commit()
        total_cotacoes = Cotacao.query.count()
        print(f"✅ {total_cotacoes} cotações no banco")

        print("\n✨ Banco de dados populado com sucesso!")

        return {
            "message": "Banco de dados populado com dados fictícios com sucesso!",
            "data": {
                "areas": len(areas),
                "fornecedores": len(fornecedores),
                "itens": len(itens),
                "listas": len(listas),
                "itens_lista_mae": total_lista_mae,
                "estoques": total_estoques,
                "pedidos": total_pedidos,
                "cotacoes": total_cotacoes,
                "nota": "Colaboradores NÃO foram vinculados às listas - faça isso manualmente"
            }
        }, 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao popular banco de dados: {str(e)}")
        return {"error": f"Erro ao popular banco de dados: {str(e)}"}, 500


def export_lista_to_csv(lista_id, restaurante_id=None):
    """
    Exporta os itens da lista para formato CSV.
    Usa a tabela de referência ListaItemRef para obter quantidades.
    """
    try:
        lista = _get_lista_por_restaurante(lista_id, restaurante_id)
        if not lista:
            return {"error": "Lista não encontrada."}, 404

        # Buscar referências de itens para esta lista
        refs = ListaItemRef.query.filter_by(lista_id=lista_id).all()

        if not refs:
            return {"error": "Lista não possui itens para exportar."}, 400

        # Gerar nome do arquivo com nome da lista e data
        data_atual = brasilia_now().strftime("%Y-%m-%d")
        nome_limpo = "".join(c for c in lista.nome if c.isalnum() or c in (' ', '-', '_')).strip()
        nome_limpo = nome_limpo.replace(' ', '_')
        filename = f"{nome_limpo}_{data_atual}.csv"

        # Gerar conteúdo CSV
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)

        # Cabeçalho
        writer.writerow(['nome', 'unidade', 'quantidade_atual', 'quantidade_minima'])

        # Dados (item global + quantidades da referência)
        for ref in refs:
            writer.writerow([
                ref.item.nome,
                ref.item.unidade,
                float(ref.quantidade_atual) if ref.quantidade_atual else 0,
                float(ref.quantidade_minima) if ref.quantidade_minima else 0
            ])

        csv_content = output.getvalue()
        output.close()

        return {
            "csv_content": csv_content,
            "filename": filename
        }, 200

    except Exception as e:
        print(f"❌ Erro ao exportar lista para CSV: {str(e)}")
        return {"error": f"Erro ao exportar lista: {str(e)}"}, 500


def import_lista_from_csv(lista_id, csv_file, restaurante_id=None):
    """
    Importa itens para uma lista a partir de um arquivo CSV.
    Itens são adicionados ao catálogo global e vinculados à lista.
    Substitui os vínculos existentes pelos do CSV.
    """
    try:
        if restaurante_id is None:
            return {"error": "Restaurante é obrigatório."}, 400

        lista = _get_lista_por_restaurante(lista_id, restaurante_id)
        if not lista:
            return {"error": "Lista não encontrada."}, 404

        import csv
        from io import StringIO

        if hasattr(csv_file, 'read'):
            content = csv_file.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
        else:
            content = csv_file

        csv_reader = csv.DictReader(StringIO(content))

        required_headers = {'nome', 'unidade', 'quantidade_atual', 'quantidade_minima'}
        headers = set(csv_reader.fieldnames or [])

        if not required_headers.issubset(headers):
            return {
                "error": f"CSV deve conter os cabeçalhos: {', '.join(required_headers)}"
            }, 400

        # Remover vínculos existentes da lista (não os itens globais)
        ListaItemRef.query.filter_by(lista_id=lista_id).delete()

        itens_importados = 0
        itens_criados = 0

        for row in csv_reader:
            try:
                nome = row['nome'].strip()
                if not nome:
                    continue

                # Buscar ou criar item no catálogo global
                item = ListaMaeItem.query.filter(
                    ListaMaeItem.restaurante_id == restaurante_id,
                    func.lower(ListaMaeItem.nome) == func.lower(nome)
                ).first()

                if not item:
                    item = ListaMaeItem(
                        nome=nome,
                        unidade=row['unidade'].strip() if row['unidade'] else 'un',
                        restaurante_id=restaurante_id
                    )
                    db.session.add(item)
                    db.session.flush()
                    itens_criados += 1

                # Criar referência
                ref = ListaItemRef(
                    lista_id=lista_id,
                    item_id=item.id,
                    quantidade_atual=float(row['quantidade_atual']) if row['quantidade_atual'] else 0,
                    quantidade_minima=float(row['quantidade_minima']) if row['quantidade_minima'] else 1.0
                )
                db.session.add(ref)
                itens_importados += 1

            except (ValueError, KeyError) as e:
                db.session.rollback()
                return {
                    "error": f"Erro ao processar linha do CSV: {str(e)}"
                }, 400

        db.session.commit()

        return {
            "message": f"Lista importada com sucesso! {itens_importados} itens vinculados, {itens_criados} novos no catálogo.",
            "itens_importados": itens_importados,
            "itens_criados": itens_criados
        }, 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao importar lista de CSV: {str(e)}")
        return {"error": f"Erro ao importar lista: {str(e)}"}, 500


def create_lista_from_csv(nome, descricao, csv_file, restaurante_id=None):
    """
    Cria uma nova lista e importa itens a partir de um arquivo CSV.
    Itens são adicionados ao catálogo global e vinculados à nova lista.
    """
    try:
        if not nome or not nome.strip():
            return {"error": "Nome da lista é obrigatório."}, 400
        if restaurante_id is None:
            return {"error": "Restaurante é obrigatório."}, 400

        nome_limpo = nome.strip()
        lista_query = Lista.query.filter(func.lower(Lista.nome) == func.lower(nome_limpo))
        if restaurante_id is not None:
            lista_query = lista_query.filter_by(restaurante_id=restaurante_id)
        lista_existente = lista_query.first()
        if lista_existente:
            return {"error": f"Já existe uma lista com o nome '{nome_limpo}'."}, 409

        nova_lista = Lista(
            nome=nome_limpo,
            descricao=descricao.strip() if descricao else None,
            restaurante_id=restaurante_id
        )
        db.session.add(nova_lista)
        db.session.flush()

        import csv
        from io import StringIO

        if hasattr(csv_file, 'read'):
            content = csv_file.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
        else:
            content = csv_file

        lines = content.strip().split('\n')

        is_csv_format = False
        if lines and ',' in lines[0]:
            first_line_lower = lines[0].lower().strip()
            if 'nome' in first_line_lower.split(','):
                is_csv_format = True

        itens_vinculados = 0
        itens_criados = 0

        def find_or_create_item_and_ref(nome_item, unidade='un', quantidade_atual=0, quantidade_minima=1.0):
            """Helper para buscar/criar item global e criar referência"""
            nonlocal itens_criados, itens_vinculados

            # Buscar item no catálogo global
            item = ListaMaeItem.query.filter(
                ListaMaeItem.restaurante_id == restaurante_id,
                func.lower(ListaMaeItem.nome) == func.lower(nome_item)
            ).first()

            if not item:
                item = ListaMaeItem(nome=nome_item, unidade=unidade, restaurante_id=restaurante_id)
                db.session.add(item)
                db.session.flush()
                itens_criados += 1

            # Criar referência
            ref = ListaItemRef(
                lista_id=nova_lista.id,
                item_id=item.id,
                quantidade_atual=quantidade_atual,
                quantidade_minima=quantidade_minima
            )
            db.session.add(ref)
            itens_vinculados += 1

        if is_csv_format:
            csv_reader = csv.DictReader(StringIO(content))

            required_headers = {'nome'}
            headers = set(csv_reader.fieldnames or [])

            if not required_headers.issubset(headers):
                db.session.rollback()
                return {
                    "error": "CSV deve conter no mínimo a coluna 'nome'. Opcionais: unidade, quantidade_atual, quantidade_minima"
                }, 400

            for row in csv_reader:
                try:
                    nome_item = row.get('nome', '').strip()
                    if not nome_item:
                        continue

                    unidade = row.get('unidade', 'un').strip() if row.get('unidade') else 'un'
                    quantidade_atual = 0
                    quantidade_minima = 1.0

                    if row.get('quantidade_atual'):
                        try:
                            quantidade_atual = float(row['quantidade_atual'])
                        except ValueError:
                            pass

                    if row.get('quantidade_minima'):
                        try:
                            valor = float(row['quantidade_minima'])
                            if valor > 0:
                                quantidade_minima = valor
                        except ValueError:
                            pass

                    find_or_create_item_and_ref(nome_item, unidade, quantidade_atual, quantidade_minima)

                except Exception as e:
                    db.session.rollback()
                    return {"error": f"Erro ao processar linha do CSV: {str(e)}"}, 400
        else:
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                try:
                    find_or_create_item_and_ref(line, 'un', 0, 1.0)
                except Exception as e:
                    db.session.rollback()
                    return {"error": f"Erro ao processar linha: {str(e)}"}, 400

        db.session.commit()

        return {
            "message": f"Lista '{nome}' criada com sucesso! {itens_vinculados} itens vinculados, {itens_criados} novos no catálogo.",
            "lista_id": nova_lista.id,
            "itens_vinculados": itens_vinculados,
            "itens_criados": itens_criados
        }, 201

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao criar lista de CSV: {str(e)}")
        return {"error": f"Erro ao criar lista: {str(e)}"}, 500


def export_data_bulk(tipos_dados, restaurante_id=None):
    """
    Exporta múltiplos tipos de dados em um arquivo ZIP contendo CSVs.

    Args:
        tipos_dados: Lista de strings indicando quais dados exportar
                    ['usuarios', 'listas', 'itens', 'fornecedores', 'areas', 'pedidos', 'cotacoes', 'estoque']

    Returns:
        BytesIO contendo o arquivo ZIP, ou erro
    """
    import io
    import zipfile
    import csv

    try:
        # Criar arquivo ZIP em memória
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:

            # Exportar Usuários
            if 'usuarios' in tipos_dados:
                usuarios_query = Usuario.query
                if restaurante_id is not None:
                    usuarios_query = usuarios_query.filter(
                        Usuario.role != UserRoles.SUPER_ADMIN,
                        Usuario.restaurante_id == restaurante_id
                    )
                usuarios = usuarios_query.all()
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['ID', 'Nome', 'Email', 'Username', 'Role', 'Aprovado', 'Ativo', 'Data Criação'])

                for user in usuarios:
                    csv_writer.writerow([
                        user.id,
                        user.nome,
                        user.email,
                        user.username or '',
                        user.role.value,
                        'Sim' if user.aprovado else 'Não',
                        'Sim' if user.ativo else 'Não',
                        user.data_criacao.strftime('%Y-%m-%d %H:%M:%S') if user.data_criacao else ''
                    ])

                zip_file.writestr('usuarios.csv', csv_buffer.getvalue())

            # Exportar Listas
            if 'listas' in tipos_dados:
                listas_query = Lista.query.filter_by(deletado=False)
                if restaurante_id is not None:
                    listas_query = listas_query.filter_by(restaurante_id=restaurante_id)
                listas = listas_query.all()
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['ID', 'Nome', 'Descrição', 'Data Criação', 'Colaboradores'])

                for lista in listas:
                    colaboradores_nomes = ', '.join([c.nome for c in lista.colaboradores])
                    csv_writer.writerow([
                        lista.id,
                        lista.nome,
                        lista.descricao or '',
                        lista.data_criacao.strftime('%Y-%m-%d %H:%M:%S'),
                        colaboradores_nomes
                    ])

                zip_file.writestr('listas.csv', csv_buffer.getvalue())

            # Exportar Itens
            if 'itens' in tipos_dados:
                itens_query = Item.query
                if restaurante_id is not None:
                    itens_query = itens_query.join(Fornecedor).filter(Fornecedor.restaurante_id == restaurante_id)
                itens = itens_query.all()
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['ID', 'Nome', 'Unidade Medida', 'Fornecedor', 'Data Criação'])

                for item in itens:
                    csv_writer.writerow([
                        item.id,
                        item.nome,
                        item.unidade_medida,
                        item.fornecedor.nome if item.fornecedor else '',
                        item.data_criacao.strftime('%Y-%m-%d %H:%M:%S') if item.data_criacao else ''
                    ])

                zip_file.writestr('itens.csv', csv_buffer.getvalue())

            # Exportar Fornecedores
            if 'fornecedores' in tipos_dados:
                fornecedores_query = Fornecedor.query
                if restaurante_id is not None:
                    fornecedores_query = fornecedores_query.filter_by(restaurante_id=restaurante_id)
                fornecedores = fornecedores_query.all()
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['ID', 'Nome', 'Contato', 'Telefone', 'Email', 'Responsável', 'Observação'])

                for fornecedor in fornecedores:
                    csv_writer.writerow([
                        fornecedor.id,
                        fornecedor.nome,
                        fornecedor.contato or '',
                        fornecedor.telefone or '',
                        fornecedor.email or '',
                        fornecedor.responsavel or '',
                        fornecedor.observacao or ''
                    ])

                zip_file.writestr('fornecedores.csv', csv_buffer.getvalue())

            # Exportar Áreas
            if 'areas' in tipos_dados:
                areas = Area.query.all()
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['ID', 'Nome', 'Descrição'])

                for area in areas:
                    csv_writer.writerow([
                        area.id,
                        area.nome,
                        area.descricao or ''
                    ])

                zip_file.writestr('areas.csv', csv_buffer.getvalue())

            # Exportar Pedidos
            if 'pedidos' in tipos_dados:
                pedidos_query = Pedido.query
                if restaurante_id is not None:
                    pedidos_query = pedidos_query.join(
                        Fornecedor, Pedido.fornecedor_id == Fornecedor.id
                    ).filter(Fornecedor.restaurante_id == restaurante_id)
                pedidos = pedidos_query.all()
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['ID', 'Item', 'Fornecedor', 'Quantidade', 'Status', 'Usuário', 'Data Pedido'])

                for pedido in pedidos:
                    csv_writer.writerow([
                        pedido.id,
                        pedido.item.nome if pedido.item else '',
                        pedido.fornecedor.nome if pedido.fornecedor else '',
                        pedido.quantidade_solicitada,
                        pedido.status.value if pedido.status else '',
                        pedido.usuario.nome if pedido.usuario else '',
                        pedido.data_pedido.strftime('%Y-%m-%d %H:%M:%S') if pedido.data_pedido else ''
                    ])

                zip_file.writestr('pedidos.csv', csv_buffer.getvalue())

            # Exportar Cotações
            if 'cotacoes' in tipos_dados:
                cotacoes_query = Cotacao.query
                if restaurante_id is not None:
                    cotacoes_query = cotacoes_query.join(
                        Fornecedor, Cotacao.fornecedor_id == Fornecedor.id
                    ).filter(Fornecedor.restaurante_id == restaurante_id)
                cotacoes = cotacoes_query.all()
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['ID', 'Fornecedor', 'Status', 'Data Criação', 'Total Itens'])

                for cotacao in cotacoes:
                    csv_writer.writerow([
                        cotacao.id,
                        cotacao.fornecedor.nome if cotacao.fornecedor else '',
                        cotacao.status.value if cotacao.status else '',
                        cotacao.data_criacao.strftime('%Y-%m-%d %H:%M:%S') if cotacao.data_criacao else '',
                        len(cotacao.itens) if cotacao.itens else 0
                    ])

                zip_file.writestr('cotacoes.csv', csv_buffer.getvalue())

            # Exportar Estoque
            if 'estoque' in tipos_dados:
                estoques_query = Estoque.query
                if restaurante_id is not None:
                    estoques_query = estoques_query.join(Item).join(Fornecedor).filter(
                        Fornecedor.restaurante_id == restaurante_id
                    )
                estoques = estoques_query.all()
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['ID', 'Item', 'Área', 'Quantidade Atual', 'Quantidade Mínima'])

                for estoque in estoques:
                    csv_writer.writerow([
                        estoque.id,
                        estoque.item.nome if estoque.item else '',
                        estoque.area.nome if estoque.area else '',
                        estoque.quantidade_atual,
                        estoque.quantidade_minima
                    ])

                zip_file.writestr('estoque.csv', csv_buffer.getvalue())

        # Resetar posição do buffer para o início
        zip_buffer.seek(0)

        return zip_buffer, 200

    except Exception as e:
        print(f"❌ Erro ao exportar dados em lote: {str(e)}")
        return {"error": f"Erro ao exportar dados: {str(e)}"}, 500


# ===== IMPORTAÇÃO COM ESTOQUE (NOVA FUNCIONALIDADE) =====

def preview_importacao_estoque(data):
    """
    Faz preview da importação antes de executar
    
    Args:
        data: {
            'texto': str - Texto com dados para importar,
            'area_id': int - ID da área (opcional para preview),
            'fornecedor_id': int - ID do fornecedor (opcional para preview),
            'formato_forcado': str - 'simples' ou 'completo' (opcional)
        }
    
    Returns:
        Tuple (response_dict, status_code)
    """
    from .import_parser import parse_texto_importacao
    
    try:
        texto = data.get('texto', '').strip()
        formato_forcado = data.get('formato_forcado')  # Novo parâmetro
        
        if not texto:
            return {"error": "Texto para importação é obrigatório"}, 400
        
        # Parse do texto COM formato forçado se especificado
        if formato_forcado:
            resultado = parse_texto_importacao(texto, formato_forcado=formato_forcado)
        else:
            resultado = parse_texto_importacao(texto)
        
        if not resultado['sucesso']:
            return {
                "error": "Nenhum item válido encontrado",
                "erros": resultado['erros']
            }, 400
        
        # Retorna preview
        return {
            "formato": resultado['formato'],
            "total_itens": resultado['total_itens'],
            "total_erros": resultado['total_erros'],
            "itens": resultado['itens'],
            "erros": resultado['erros']
        }, 200
        
    except Exception as e:
        print(f"❌ Erro ao fazer preview: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": f"Erro ao processar preview: {str(e)}"}, 500


def executar_importacao_estoque(data, restaurante_id=None):
    """
    Executa a importação de itens com estoque
    
    Args:
        data: {
            'texto': str - Texto com dados para importar,
            'area_id': int - ID da área,
            'fornecedor_id': int - ID do fornecedor,
            'atualizar_existentes': bool - Se deve atualizar itens existentes (default: True)
        }
    
    Returns:
        Tuple (response_dict, status_code)
    """
    from .import_parser import parse_texto_importacao
    
    try:
        texto = data.get('texto', '').strip()
        area_id = data.get('area_id')
        fornecedor_id = data.get('fornecedor_id')
        atualizar_existentes = data.get('atualizar_existentes', True)
        
        # Validações
        if not texto:
            return {"error": "Texto para importação é obrigatório"}, 400
        
        if not area_id:
            return {"error": "Área é obrigatória"}, 400
        
        if not fornecedor_id:
            return {"error": "Fornecedor é obrigatório"}, 400
        
        # Valida se área existe
        area = Area.query.get(area_id)
        if not area:
            return {"error": f"Área com ID {area_id} não encontrada"}, 404
        
        # Valida se fornecedor existe
        fornecedor = _get_fornecedor_por_id(fornecedor_id, restaurante_id)
        if not fornecedor:
            return {"error": f"Fornecedor com ID {fornecedor_id} não encontrado"}, 404
        
        # Parse do texto
        resultado = parse_texto_importacao(texto)
        
        if not resultado['sucesso']:
            return {
                "error": "Nenhum item válido encontrado",
                "erros": resultado['erros']
            }, 400
        
        itens_criados = []
        itens_atualizados = []
        erros_processamento = []
        
        # Processa cada item
        for item_data in resultado['itens']:
            try:
                nome_item = item_data['nome']
                qtd_atual = item_data.get('quantidade_atual')
                qtd_minima = item_data.get('quantidade_minima')
                
                # Busca item existente no estoque
                estoque_existente = Estoque.query.filter_by(
                    nome_item=nome_item,
                    area_id=area_id
                ).first()
                
                if estoque_existente:
                    # Atualiza item existente
                    if atualizar_existentes:
                        if qtd_atual is not None:
                            estoque_existente.quantidade_atual = qtd_atual
                        if qtd_minima is not None:
                            estoque_existente.quantidade_minima = qtd_minima
                        
                        db.session.commit()
                        itens_atualizados.append({
                            'id': estoque_existente.id,
                            'nome': nome_item,
                            'quantidade_atual': estoque_existente.quantidade_atual,
                            'quantidade_minima': estoque_existente.quantidade_minima
                        })
                    else:
                        # Pula se não deve atualizar
                        erros_processamento.append(f"Item '{nome_item}' já existe e não foi atualizado")
                else:
                    # Cria novo item no estoque
                    novo_estoque = Estoque(
                        nome_item=nome_item,
                        area_id=area_id,
                        fornecedor_id=fornecedor_id,
                        quantidade_atual=qtd_atual if qtd_atual is not None else 0,
                        quantidade_minima=qtd_minima if qtd_minima is not None else 0,
                        unidade_medida='UN'  # Padrão, usuário configura depois
                    )
                    db.session.add(novo_estoque)
                    db.session.flush()  # Para obter o ID
                    
                    itens_criados.append({
                        'id': novo_estoque.id,
                        'nome': nome_item,
                        'quantidade_atual': novo_estoque.quantidade_atual,
                        'quantidade_minima': novo_estoque.quantidade_minima
                    })
                
            except Exception as e:
                db.session.rollback()
                erros_processamento.append(f"Erro ao processar '{item_data['nome']}': {str(e)}")
                continue
        
        # Commit final
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {"error": f"Erro ao salvar no banco: {str(e)}"}, 500
        
        return {
            "sucesso": True,
            "formato": resultado['formato'],
            "total_criados": len(itens_criados),
            "total_atualizados": len(itens_atualizados),
            "total_erros": len(erros_processamento),
            "itens_criados": itens_criados,
            "itens_atualizados": itens_atualizados,
            "erros": erros_processamento + resultado['erros']
        }, 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao executar importação: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": f"Erro ao executar importação: {str(e)}"}, 500


def get_navbar_preferences(user_id):
    """Busca preferências de navbar do usuário."""
    from .models import NavbarPreference
    
    pref = NavbarPreference.query.filter_by(usuario_id=user_id).first()
    
    if not pref:
        # Retorna estado padrão (todas categorias recolhidas)
        return {
            "categorias_estado": {},
            "usuario_id": user_id
        }, 200
    
    return pref.to_dict(), 200


def save_navbar_preferences(user_id, data):
    """Salva ou atualiza preferências de navbar do usuário."""
    from .models import NavbarPreference
    from .extensions import db
    
    categorias_estado = data.get('categorias_estado', {})
    
    pref = NavbarPreference.query.filter_by(usuario_id=user_id).first()
    
    if pref:
        # Atualizar existente
        pref.categorias_estado = categorias_estado
        pref.atualizado_em = brasilia_now()
    else:
        # Criar novo
        pref = NavbarPreference(
            usuario_id=user_id,
            categorias_estado=categorias_estado
        )
        db.session.add(pref)
    
    db.session.commit()
    
    return {"message": "Preferências salvas com sucesso", "data": pref.to_dict()}, 200


_NAVBAR_ROLE_KEYS = {'ADMIN', 'SUPER_ADMIN', 'COLLABORATOR', 'SUPPLIER'}


def _normalize_navbar_role(role):
    if not role:
        return None
    role_value = str(role).upper()
    return role_value if role_value in _NAVBAR_ROLE_KEYS else None


def _extract_navbar_layout_for_role(layout_payload, role=None):
    if not isinstance(layout_payload, dict):
        return None
    if isinstance(layout_payload.get('groups'), list):
        return layout_payload
    by_role = layout_payload.get('by_role')
    if isinstance(by_role, dict) and role:
        role_layout = by_role.get(role)
        if isinstance(role_layout, dict):
            return role_layout
    default_layout = layout_payload.get('default')
    if isinstance(default_layout, dict):
        return default_layout
    return None


def _upsert_navbar_layout_for_role(layout_payload, role, layout_data):
    if not role:
        return layout_data

    if isinstance(layout_payload, dict) and isinstance(layout_payload.get('groups'), list):
        store = {"default": layout_payload, "by_role": {}}
    elif isinstance(layout_payload, dict):
        store = layout_payload
    else:
        store = {"by_role": {}}

    by_role = store.get("by_role")
    if not isinstance(by_role, dict):
        by_role = {}
        store["by_role"] = by_role
    by_role[role] = layout_data
    return store


def get_navbar_layout(role=None):
    """Retorna o layout da navbar (com suporte a role)."""
    from .models import NavbarLayout

    role_key = _normalize_navbar_role(role)
    layout = NavbarLayout.query.first()
    if not layout:
        return {"layout": None}, 200
    layout_data = _extract_navbar_layout_for_role(layout.layout or {}, role_key)
    return {"layout": layout_data}, 200


def save_navbar_layout(data, role=None):
    """Salva o layout da navbar (com suporte a role)."""
    from .models import NavbarLayout
    from .extensions import db
    from sqlalchemy.orm.attributes import flag_modified

    layout_data = data.get('layout')
    if not isinstance(layout_data, dict) or not isinstance(layout_data.get('groups', []), list):
        return {"error": "Layout inválido."}, 400

    role_key = _normalize_navbar_role(role or data.get('role'))
    layout = NavbarLayout.query.first()
    if layout:
        layout.layout = _upsert_navbar_layout_for_role(layout.layout or {}, role_key, layout_data)
        flag_modified(layout, 'layout')
        layout.atualizado_em = brasilia_now()
    else:
        layout_payload = _upsert_navbar_layout_for_role({}, role_key, layout_data)
        layout = NavbarLayout(layout=layout_payload)
        db.session.add(layout)

    db.session.commit()
    return {"layout": layout_data}, 200


def get_navbar_activity(user_id):
    """Retorna os últimos itens acessados pelo usuário."""
    from .models import NavbarActivity

    items = (
        NavbarActivity.query
        .filter_by(usuario_id=user_id)
        .order_by(NavbarActivity.criado_em.desc())
        .limit(6)
        .all()
    )

    return {"items": [item.item_key for item in items]}, 200


def log_navbar_activity(user_id, data):
    """Registra um clique na navbar para o usuário."""
    from .models import NavbarActivity
    from .extensions import db

    item_key = (data or {}).get("item_key", "")
    if not item_key or not isinstance(item_key, str):
        return {"error": "item_key inválido."}, 400

    item_key = item_key.strip()
    if not item_key:
        return {"error": "item_key inválido."}, 400

    activity = NavbarActivity.query.filter_by(usuario_id=user_id, item_key=item_key).first()
    timestamp = brasilia_now()
    if activity:
        activity.criado_em = timestamp
    else:
        activity = NavbarActivity(usuario_id=user_id, item_key=item_key, criado_em=timestamp)
        db.session.add(activity)

    db.session.flush()

    # Manter apenas os registros mais recentes para este usuário (máximo 20)
    excess_activities = (
        NavbarActivity.query
        .filter_by(usuario_id=user_id)
        .order_by(NavbarActivity.criado_em.desc())
        .offset(20)
        .all()
    )
    for old in excess_activities:
        db.session.delete(old)

    db.session.commit()
    return {"message": "Atividade registrada."}, 200


# ===== SUGESTÕES DE ITENS =====

def adicionar_item_temporario_lista_rapida(lista_rapida_id, data):
    """
    Adiciona item temporário (não do catálogo global) a uma lista rápida.
    Este item só existirá nesta lista específica.
    """
    nome_item = data.get('nome_item', '').strip()
    unidade = data.get('unidade', '').strip() or 'un'
    prioridade = data.get('prioridade', 'PREVENCAO').upper()
    observacao = data.get('observacao', '').strip()

    if not nome_item:
        return {"error": "Nome do item é obrigatório."}, 400

    # Verifica se a lista rápida existe
    lista_rapida = ListaRapida.query.get(lista_rapida_id)
    if not lista_rapida:
        return {"error": "Lista rápida não encontrada."}, 404

    if lista_rapida.status != StatusListaRapida.RASCUNHO:
        return {"error": "Só é possível adicionar itens em listas com status RASCUNHO."}, 400

    # Validar prioridade
    try:
        prioridade_enum = PrioridadeItem[prioridade]
    except KeyError:
        return {"error": "Prioridade inválida."}, 400

    # Criar item temporário
    item = ListaRapidaItem(
        lista_rapida_id=lista_rapida_id,
        item_global_id=None,  # Item temporário
        item_nome_temp=nome_item,
        item_unidade_temp=unidade,
        prioridade=prioridade_enum,
        observacao=observacao if observacao else None
    )

    db.session.add(item)
    db.session.commit()

    return {
        "message": "Item temporário adicionado com sucesso!",
        "item": item.to_dict()
    }, 201


def criar_sugestao_item(user_id, data, restaurante_id=None):
    """
    Cria uma sugestão de novo item para o catálogo global.
    Pode ser vinculada a uma lista tradicional, a uma lista rápida, OU genérica (sem lista).
    """
    lista_id = data.get('lista_id')
    lista_rapida_id = data.get('lista_rapida_id')
    nome_item = data.get('nome_item', '').strip()
    unidade = data.get('unidade', '').strip() or 'un'  # Padrão 'un' se não informar
    quantidade = data.get('quantidade') or 1  # Padrão 1 se não informar
    mensagem_usuario = data.get('mensagem_usuario', '').strip()

    # MUDANÇA: Permite ambos None (sugestão genérica)
    # Validação: não podem ter ambos lista_id E lista_rapida_id
    if lista_id and lista_rapida_id:
        return {"error": "Informe apenas lista_id OU lista_rapida_id, não ambos."}, 400

    if not nome_item:
        return {"error": "Nome do item é obrigatório."}, 400

    # Verifica se as listas existem (se informadas)
    if lista_id:
        lista = Lista.query.get(lista_id)
        if not lista:
            return {"error": "Lista não encontrada."}, 404

    if lista_rapida_id:
        lista_rapida = ListaRapida.query.get(lista_rapida_id)
        if not lista_rapida:
            return {"error": "Lista rápida não encontrada."}, 404

    restaurante_id = restaurante_id or _get_usuario_restaurante_id(user_id)
    if not restaurante_id:
        return {"error": "Restaurante do usuário não encontrado."}, 400

    # Verifica se item já existe no catálogo global
    item_existente = ListaMaeItem.query.filter(
        ListaMaeItem.restaurante_id == restaurante_id,
        func.lower(ListaMaeItem.nome) == func.lower(nome_item)
    ).first()
    if item_existente:
        return {"error": f"Item '{nome_item}' já existe no catálogo global."}, 409

    # Cria a sugestão
    sugestao = SugestaoItem(
        usuario_id=user_id,
        lista_id=lista_id,
        lista_rapida_id=lista_rapida_id,
        nome_item=nome_item,
        unidade=unidade,
        quantidade=float(quantidade),
        mensagem_usuario=mensagem_usuario if mensagem_usuario else None,
        status=SugestaoStatus.PENDENTE
    )

    db.session.add(sugestao)
    db.session.commit()

    return {
        "message": "Sugestão enviada com sucesso! Aguarde aprovação do administrador.",
        "sugestao": sugestao.to_dict()
    }, 201


def listar_minhas_sugestoes(user_id):
    """Lista todas as sugestões do usuário logado."""
    sugestoes = SugestaoItem.query.filter_by(usuario_id=user_id).order_by(
        SugestaoItem.criado_em.desc()
    ).all()
    
    return {
        "sugestoes": [s.to_dict() for s in sugestoes],
        "total": len(sugestoes)
    }, 200


def listar_sugestoes_pendentes(restaurante_id=None):
    """Lista todas as sugestões pendentes para o admin."""
    try:
        query = SugestaoItem.query.filter_by(status=SugestaoStatus.PENDENTE)
        if restaurante_id is not None:
            query = query.join(Usuario, SugestaoItem.usuario_id == Usuario.id).filter(Usuario.restaurante_id == restaurante_id)
        sugestoes = query.order_by(SugestaoItem.criado_em.asc()).all()

        return {
            "sugestoes": [s.to_dict() for s in sugestoes],
            "total": len(sugestoes)
        }, 200
    except Exception as e:
        print(f"[ERRO] listar_sugestoes_pendentes: {str(e)}")
        raise


def contar_sugestoes_pendentes(restaurante_id=None):
    """Conta quantas sugestões estão pendentes."""
    query = SugestaoItem.query.filter_by(status=SugestaoStatus.PENDENTE)
    if restaurante_id is not None:
        query = query.join(Usuario, SugestaoItem.usuario_id == Usuario.id).filter(Usuario.restaurante_id == restaurante_id)
    count = query.count()
    return {"count": count}, 200


def aprovar_sugestao(sugestao_id, admin_id, data, restaurante_id=None):
    """
    Admin aprova a sugestão:
    1. Cria item no catálogo global
    2. APENAS se lista_id existir: Adiciona item à lista via ListaItemRef
    3. Sugestões genéricas (sem lista) ou de listas rápidas: apenas cria no catálogo
    4. Atualiza status da sugestão
    """
    sugestao = SugestaoItem.query.get(sugestao_id)
    if not sugestao:
        return {"error": "Sugestão não encontrada."}, 404
    
    if sugestao.status != SugestaoStatus.PENDENTE:
        return {"error": "Sugestão já foi respondida."}, 400

    mensagem_admin = data.get('mensagem_admin', '').strip()
    unidade = data.get('unidade', '').strip() or sugestao.unidade
    quantidade = data.get('quantidade') or sugestao.quantidade

    if not unidade:
        return {"error": "Unidade é obrigatória."}, 400
    if not quantidade or quantidade <= 0:
        return {"error": "Quantidade deve ser maior que zero."}, 400

    try:
        restaurante_sugestao_id = sugestao.usuario.restaurante_id if sugestao.usuario else None
        if not restaurante_sugestao_id:
            return {"error": "Restaurante da sugestão não encontrado."}, 400
        if restaurante_id is not None and restaurante_sugestao_id != restaurante_id:
            return {"error": "Acesso negado."}, 403

        item_global = ListaMaeItem.query.filter(
            ListaMaeItem.restaurante_id == restaurante_sugestao_id,
            func.lower(ListaMaeItem.nome) == func.lower(sugestao.nome_item)
        ).first()
        item_criado = False
        if not item_global:
            # 1. Cria item no catálogo global
            item_global = ListaMaeItem(
                nome=sugestao.nome_item,
                unidade=unidade,
                restaurante_id=restaurante_sugestao_id
            )
            db.session.add(item_global)
            db.session.flush()  # Para obter o ID
            item_criado = True

        # 2. Adiciona à lista do usuário (apenas se for lista tradicional)
        # Sugestões de listas rápidas não criam ListaItemRef
        if sugestao.lista_id:
            ref_existe = ListaItemRef.query.filter_by(
                lista_id=sugestao.lista_id,
                item_id=item_global.id
            ).first()
            if not ref_existe:
                item_ref = ListaItemRef(
                    lista_id=sugestao.lista_id,
                    item_id=item_global.id,
                    quantidade_atual=0,
                    quantidade_minima=float(quantidade)
                )
                db.session.add(item_ref)

        # 3. Atualiza sugestão
        sugestao.status = SugestaoStatus.APROVADA
        sugestao.admin_id = admin_id
        sugestao.mensagem_admin = mensagem_admin if mensagem_admin else None
        sugestao.item_global_id = item_global.id
        sugestao.unidade = unidade
        sugestao.quantidade = float(quantidade)
        sugestao.respondido_em = brasilia_now()

        db.session.commit()

        mensagem = "Sugestão aprovada! Item adicionado ao catálogo global."
        if not item_criado:
            mensagem = "Sugestão aprovada! Item já existia no catálogo global."
        if sugestao.lista_id:
            mensagem += " Item também adicionado à lista."

        return {
            "message": mensagem,
            "sugestao": sugestao.to_dict(),
            "item_global": {
                "id": item_global.id,
                "nome": item_global.nome,
                "unidade": item_global.unidade
            }
        }, 200

    except Exception as e:
        db.session.rollback()
        return {"error": f"Erro ao aprovar sugestão: {str(e)}"}, 500


def rejeitar_sugestao(sugestao_id, admin_id, data, restaurante_id=None):
    """
    Admin rejeita a sugestão.
    Apenas atualiza o status e adiciona mensagem.
    """
    sugestao = SugestaoItem.query.get(sugestao_id)
    if not sugestao:
        return {"error": "Sugestão não encontrada."}, 404
    
    if sugestao.status != SugestaoStatus.PENDENTE:
        return {"error": "Sugestão já foi respondida."}, 400

    restaurante_sugestao_id = sugestao.usuario.restaurante_id if sugestao.usuario else None
    if restaurante_id is not None and restaurante_sugestao_id != restaurante_id:
        return {"error": "Acesso negado."}, 403

    mensagem_admin = data.get('mensagem_admin', '').strip()
    if not mensagem_admin:
        return {"error": "Mensagem explicativa é obrigatória ao rejeitar."}, 400

    sugestao.status = SugestaoStatus.REJEITADA
    sugestao.admin_id = admin_id
    sugestao.mensagem_admin = mensagem_admin
    sugestao.respondido_em = brasilia_now()

    db.session.commit()

    return {
        "message": "Sugestão rejeitada.",
        "sugestao": sugestao.to_dict()
    }, 200


# ===== LISTA RÁPIDA =====

def criar_lista_rapida(user_id, data):
    """Cria uma nova lista rápida."""
    from .models import ListaRapida, StatusListaRapida
    from .extensions import db
    
    nome = (data.get('nome') or '').strip()
    descricao = (data.get('descricao') or '').strip()
    
    if not nome:
        return {"error": "Nome da lista é obrigatório."}, 400
    
    lista = ListaRapida(
        usuario_id=user_id,
        nome=nome,
        descricao=descricao if descricao else None,
        status=StatusListaRapida.RASCUNHO
    )
    
    db.session.add(lista)
    db.session.commit()
    
    return {"message": "Lista rápida criada!", "lista": lista.to_dict()}, 201


def listar_minhas_listas_rapidas(user_id):
    """Lista todas as listas rápidas do usuário."""
    from .models import ListaRapida
    
    listas = ListaRapida.query.filter_by(
        usuario_id=user_id,
        deletado=False
    ).order_by(ListaRapida.criado_em.desc()).all()
    
    return {
        "listas": [l.to_dict() for l in listas],
        "total": len(listas)
    }, 200


def obter_lista_rapida(lista_id, user_id):
    """Obtém detalhes de uma lista rápida com seus itens."""
    from .models import ListaRapida
    
    lista = ListaRapida.query.filter_by(
        id=lista_id,
        usuario_id=user_id,
        deletado=False
    ).first()
    
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    
    lista_dict = lista.to_dict()
    lista_dict['itens'] = [item.to_dict() for item in lista.itens.all()]
    
    return lista_dict, 200


def adicionar_item_lista_rapida(lista_id, user_id, data):
    """Adiciona item à lista rápida."""
    from .models import ListaRapida, ListaRapidaItem, ListaMaeItem, PrioridadeItem, StatusListaRapida
    from .extensions import db
    
    restaurante_id = _get_usuario_restaurante_id(user_id)
    
    lista = ListaRapida.query.filter_by(
        id=lista_id,
        usuario_id=user_id,
        deletado=False
    ).first()
    
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    
    if lista.status != StatusListaRapida.RASCUNHO:
        return {"error": "Não é possível adicionar itens a uma lista já submetida."}, 400
    
    item_global_id = data.get('item_global_id')
    prioridade_str = data.get('prioridade', 'precisa_comprar')
    observacao = (data.get('observacao') or '').strip()
    
    if not item_global_id:
        return {"error": "ID do item é obrigatório."}, 400
    
    # Verifica se item existe
    item_global = ListaMaeItem.query.get(item_global_id)
    if not item_global:
        return {"error": "Item não encontrado no catálogo."}, 404
    if restaurante_id is not None and item_global.restaurante_id != restaurante_id:
        return {"error": "Item não pertence ao seu restaurante."}, 403
    
    # Verifica se já foi adicionado
    item_existente = ListaRapidaItem.query.filter_by(
        lista_rapida_id=lista_id,
        item_global_id=item_global_id
    ).first()
    
    if item_existente:
        return {"error": "Item já está na lista."}, 409
    
    # Converte string para enum
    try:
        prioridade = PrioridadeItem(prioridade_str)
    except ValueError:
        prioridade = PrioridadeItem.PRECISA_COMPRAR
    
    item = ListaRapidaItem(
        lista_rapida_id=lista_id,
        item_global_id=item_global_id,
        prioridade=prioridade,
        observacao=observacao if observacao else None
    )
    
    db.session.add(item)
    db.session.commit()
    
    return {"message": "Item adicionado!", "item": item.to_dict()}, 201


def adicionar_multiplos_itens_lista_rapida(lista_id, user_id, data):
    """Adiciona múltiplos itens à lista rápida de uma vez."""
    from .models import ListaRapida, ListaRapidaItem, ListaMaeItem, PrioridadeItem, StatusListaRapida
    from .extensions import db
    
    restaurante_id = _get_usuario_restaurante_id(user_id)
    
    lista = ListaRapida.query.filter_by(
        id=lista_id,
        usuario_id=user_id,
        deletado=False
    ).first()
    
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    
    if lista.status != StatusListaRapida.RASCUNHO:
        return {"error": "Não é possível adicionar itens a uma lista já submetida."}, 400
    
    itens_data = data.get('itens', [])
    if not itens_data:
        return {"error": "Nenhum item informado."}, 400
    
    itens_adicionados = []
    erros = []
    
    for item_data in itens_data:
        try:
            item_global_id = item_data.get('item_global_id')
            prioridade_str = item_data.get('prioridade', 'precisa_comprar')
            observacao = (item_data.get('observacao') or '').strip()
            
            if not item_global_id:
                continue
            
            # Verifica se item existe
            item_global = ListaMaeItem.query.get(item_global_id)
            if not item_global:
                erros.append(f"Item ID {item_global_id} não encontrado")
                continue
            if restaurante_id is not None and item_global.restaurante_id != restaurante_id:
                erros.append(f"Item ID {item_global_id} não pertence ao seu restaurante")
                continue
            
            # Verifica se já foi adicionado
            item_existente = ListaRapidaItem.query.filter_by(
                lista_rapida_id=lista_id,
                item_global_id=item_global_id
            ).first()
            
            if item_existente:
                continue
            
            # Converte string para enum
            try:
                prioridade = PrioridadeItem(prioridade_str)
            except ValueError:
                prioridade = PrioridadeItem.PRECISA_COMPRAR
            
            item = ListaRapidaItem(
                lista_rapida_id=lista_id,
                item_global_id=item_global_id,
                prioridade=prioridade,
                observacao=observacao if observacao else None
            )
            
            db.session.add(item)
            itens_adicionados.append(item)
        except Exception as e:
            erros.append(f"Erro ao adicionar item: {str(e)}")
    
    if itens_adicionados:
        db.session.commit()
    
    return {
        "message": f"{len(itens_adicionados)} item(ns) adicionado(s)!",
        "itens_adicionados": len(itens_adicionados),
        "erros": erros if erros else None
    }, 201


def remover_item_lista_rapida(lista_id, item_id, user_id):
    """Remove item da lista rápida."""
    from .models import ListaRapida, ListaRapidaItem, StatusListaRapida
    from .extensions import db
    
    lista = ListaRapida.query.filter_by(
        id=lista_id,
        usuario_id=user_id,
        deletado=False
    ).first()
    
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    
    if lista.status != StatusListaRapida.RASCUNHO:
        return {"error": "Não é possível remover itens de uma lista já submetida."}, 400
    
    item = ListaRapidaItem.query.filter_by(
        id=item_id,
        lista_rapida_id=lista_id
    ).first()
    
    if not item:
        return {"error": "Item não encontrado na lista."}, 404
    
    db.session.delete(item)
    db.session.commit()
    
    return {"message": "Item removido com sucesso."}, 200


def atualizar_prioridade_item(lista_id, item_id, user_id, data):
    """Atualiza a prioridade de um item."""
    from .models import ListaRapida, ListaRapidaItem, PrioridadeItem, StatusListaRapida
    from .extensions import db
    
    lista = ListaRapida.query.filter_by(
        id=lista_id,
        usuario_id=user_id,
        deletado=False
    ).first()
    
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    
    if lista.status != StatusListaRapida.RASCUNHO:
        return {"error": "Não é possível editar uma lista já submetida."}, 400
    
    item = ListaRapidaItem.query.filter_by(
        id=item_id,
        lista_rapida_id=lista_id
    ).first()
    
    if not item:
        return {"error": "Item não encontrado na lista."}, 404
    
    prioridade_str = data.get('prioridade')
    
    try:
        item.prioridade = PrioridadeItem(prioridade_str)
        db.session.commit()
        return {"message": "Prioridade atualizada!", "item": item.to_dict()}, 200
    except ValueError:
        return {"error": "Prioridade inválida."}, 400


def submeter_lista_rapida(lista_id, user_id):
    """Submete lista rápida para aprovação do admin."""
    from .models import ListaRapida, StatusListaRapida, UserRoles, TipoNotificacao
    from .extensions import db

    lista = ListaRapida.query.filter_by(
        id=lista_id,
        usuario_id=user_id,
        deletado=False
    ).first()

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    if lista.status != StatusListaRapida.RASCUNHO:
        return {"error": "Lista já foi submetida."}, 400

    if lista.itens.count() == 0:
        return {"error": "Adicione pelo menos um item antes de submeter."}, 400

    # Obter informações do usuário
    usuario = Usuario.query.get(user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    lista.status = StatusListaRapida.PENDENTE
    lista.submetido_em = brasilia_now()

    db.session.commit()

    # Criar notificações para todos os admins do restaurante
    try:
        admins = Usuario.query.filter(
            Usuario.restaurante_id == usuario.restaurante_id,
            Usuario.role == UserRoles.ADMIN,  # Apenas ADMINs do restaurante
            Usuario.ativo == True,
            Usuario.aprovado == True
        ).all()

        for admin in admins:
            criar_notificacao(
                usuario_id=admin.id,
                titulo=f"Nova lista rápida de {usuario.nome}",
                tipo=TipoNotificacao.SUBMISSAO_LISTA_RAPIDA,
                mensagem=f"O usuário {usuario.nome} submeteu uma nova lista rápida para aprovação.",
                lista_rapida_id=lista.id,
                restaurante_id=usuario.restaurante_id
            )
    except Exception as e:
        print(f"[AVISO] Erro ao criar notificações para lista rápida {lista.id}: {str(e)}")
        # Não falhar o fluxo principal se houver erro nas notificações

    return {"message": "Lista submetida com sucesso!", "lista": lista.to_dict()}, 200


def deletar_lista_rapida(lista_id, user_id):
    """Soft delete de lista rápida."""
    from .models import ListaRapida, StatusListaRapida
    from .extensions import db
    
    lista = ListaRapida.query.filter_by(
        id=lista_id,
        usuario_id=user_id,
        deletado=False
    ).first()
    
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    
    if lista.status == StatusListaRapida.PENDENTE:
        return {"error": "Não é possível deletar uma lista pendente de aprovação."}, 400
    
    lista.deletado = True
    db.session.commit()
    
    return {"message": "Lista deletada com sucesso."}, 200


def listar_minhas_listas_rapidas(user_id):
    """Lista todas as listas rápidas do usuário (não só rascunhos)."""
    from .models import ListaRapida
    
    listas = ListaRapida.query.filter_by(
        usuario_id=user_id,
        deletado=False
    ).order_by(ListaRapida.criado_em.desc()).all()
    
    resultado = []
    for lista in listas:
        lista_dict = lista.to_dict()
        lista_dict['total_itens'] = lista.itens.count()
        resultado.append(lista_dict)
    
    return {"listas": resultado, "total": len(resultado)}, 200


# Admin

def _validar_lista_rapida_restaurante(lista, restaurante_id):
    if restaurante_id is None:
        return None
    if not lista or not lista.usuario or lista.usuario.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    return None

def listar_listas_rapidas_pendentes(restaurante_id=None):
    """Admin lista todas as listas rápidas pendentes."""
    from .models import ListaRapida, StatusListaRapida

    query = ListaRapida.query.filter_by(
        status=StatusListaRapida.PENDENTE,
        deletado=False
    )
    if restaurante_id is not None:
        query = query.join(Usuario, ListaRapida.usuario_id == Usuario.id).filter(Usuario.restaurante_id == restaurante_id)
    listas = query.order_by(ListaRapida.submetido_em.asc()).all()

    return {
        "listas": [l.to_dict() for l in listas],
        "total": len(listas)
    }, 200


def obter_lista_rapida_admin(lista_id, restaurante_id=None):
    """Admin obtém detalhes de qualquer lista rápida."""
    from .models import ListaRapida
    
    lista = ListaRapida.query.filter_by(
        id=lista_id,
        deletado=False
    ).first()
    
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro
    
    lista_dict = lista.to_dict()
    lista_dict['itens'] = [item.to_dict() for item in lista.itens.all()]
    
    return lista_dict, 200


def aprovar_lista_rapida(lista_id, admin_id, data, restaurante_id=None):
    """Admin aprova lista rápida."""
    from .models import ListaRapida, StatusListaRapida, TipoNotificacao
    from .extensions import db

    lista = ListaRapida.query.get(lista_id)

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if lista.status != StatusListaRapida.PENDENTE:
        return {"error": "Lista não está pendente."}, 400

    mensagem_admin = data.get('mensagem_admin', '').strip()

    lista.status = StatusListaRapida.APROVADA
    lista.admin_id = admin_id
    lista.mensagem_admin = mensagem_admin if mensagem_admin else None
    lista.respondido_em = brasilia_now()

    db.session.commit()

    # Notificar o usuário que submeteu a lista
    try:
        criar_notificacao(
            usuario_id=lista.usuario_id,
            titulo="Lista rápida aprovada",
            tipo=TipoNotificacao.LISTA_APROVADA,
            mensagem="Sua lista rápida foi aprovada.",
            lista_rapida_id=lista.id,
            restaurante_id=lista.usuario.restaurante_id
        )
    except Exception as e:
        print(f"[AVISO] Erro ao criar notificação de aprovação para lista rápida {lista.id}: {str(e)}")

    return {"message": "Lista aprovada com sucesso!", "lista": lista.to_dict()}, 200


def rejeitar_lista_rapida(lista_id, admin_id, data, restaurante_id=None):
    """Admin rejeita lista rápida."""
    from .models import ListaRapida, StatusListaRapida, TipoNotificacao
    from .extensions import db

    lista = ListaRapida.query.get(lista_id)

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if lista.status != StatusListaRapida.PENDENTE:
        return {"error": "Lista não está pendente."}, 400

    mensagem_admin = data.get('mensagem_admin', '').strip()

    if not mensagem_admin:
        return {"error": "Mensagem de rejeição é obrigatória."}, 400

    lista.status = StatusListaRapida.REJEITADA
    lista.admin_id = admin_id
    lista.mensagem_admin = mensagem_admin
    lista.respondido_em = brasilia_now()

    db.session.commit()

    # Notificar o usuário que submeteu a lista
    try:
        criar_notificacao(
            usuario_id=lista.usuario_id,
            titulo="Lista rápida rejeitada",
            tipo=TipoNotificacao.LISTA_REJEITADA,
            mensagem=f"Sua lista rápida foi rejeitada. Motivo: {mensagem_admin}",
            lista_rapida_id=lista.id,
            restaurante_id=lista.usuario.restaurante_id
        )
    except Exception as e:
        print(f"[AVISO] Erro ao criar notificação de rejeição para lista rápida {lista.id}: {str(e)}")

    return {"message": "Lista rejeitada.", "lista": lista.to_dict()}, 200


def reverter_lista_rapida_para_pendente(lista_id, restaurante_id=None):
    """Reverte uma lista rápida APROVADA ou REJEITADA para PENDENTE."""
    from .models import ListaRapida, StatusListaRapida
    from .extensions import db

    lista = ListaRapida.query.get(lista_id)

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if lista.status == StatusListaRapida.PENDENTE:
        return {"error": "Lista já está PENDENTE."}, 400

    if lista.status == StatusListaRapida.RASCUNHO:
        return {"error": "Não é possível reverter uma lista em rascunho."}, 400

    # Reverter para PENDENTE
    lista.status = StatusListaRapida.PENDENTE
    lista.admin_id = None
    lista.mensagem_admin = None
    lista.respondido_em = None

    db.session.commit()

    return {"message": "Lista revertida para PENDENTE.", "lista": lista.to_dict()}, 200


def arquivar_lista_rapida(lista_id, restaurante_id=None):
    """Arquiva uma lista rápida APROVADA ou REJEITADA."""
    from .models import ListaRapida, StatusListaRapida
    from .extensions import db

    lista = ListaRapida.query.filter_by(id=lista_id, deletado=False).first()

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if lista.arquivada:
        return {"error": "Lista já está arquivada."}, 400

    lista.arquivada = True
    lista.arquivada_em = brasilia_now()
    db.session.commit()

    return {"message": "Lista arquivada com sucesso."}, 200


def desarquivar_lista_rapida(lista_id, restaurante_id=None):
    """Remove o status de arquivada de uma lista rápida e restaura para rascunho."""
    from .models import ListaRapida, StatusListaRapida
    from .extensions import db

    lista = ListaRapida.query.filter_by(id=lista_id, deletado=False).first()

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if not lista.arquivada:
        return {"error": "Lista não está arquivada."}, 400

    lista.arquivada = False
    lista.arquivada_em = None
    # Restaura para rascunho para permitir nova edição e submissão
    lista.status = StatusListaRapida.RASCUNHO
    lista.submetido_em = None
    lista.respondido_em = None
    lista.mensagem_admin = None
    lista.admin_id = None
    db.session.commit()

    return {"message": "Lista desarquivada com sucesso."}, 200


def deletar_lista_rapida_permanente(lista_id, restaurante_id=None):
    """Exclui permanentemente uma lista rápida arquivada."""
    from .models import ListaRapida
    from .extensions import db

    lista = ListaRapida.query.filter_by(id=lista_id, deletado=False).first()

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if not lista.arquivada:
        return {"error": "Apenas listas arquivadas podem ser excluídas permanentemente."}, 400

    db.session.delete(lista)
    db.session.commit()

    return {"message": f"Lista rápida #{lista_id} excluída permanentemente."}, 200


def contar_listas_rapidas_pendentes(restaurante_id=None):
    """Conta quantas listas rápidas estão pendentes."""
    from .models import ListaRapida, StatusListaRapida

    query = ListaRapida.query.filter_by(
        status=StatusListaRapida.PENDENTE,
        deletado=False
    )
    if restaurante_id is not None:
        query = query.join(Usuario, ListaRapida.usuario_id == Usuario.id).filter(Usuario.restaurante_id == restaurante_id)
    count = query.count()

    return {"count": count}, 200


def adicionar_item_lista_rapida_admin(lista_id, data, restaurante_id=None):
    """Admin adiciona item à lista rápida PENDENTE."""
    from .models import ListaRapida, ListaRapidaItem, ListaMaeItem, PrioridadeItem, StatusListaRapida
    from .extensions import db

    lista = ListaRapida.query.filter_by(id=lista_id, deletado=False).first()

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if lista.status != StatusListaRapida.PENDENTE:
        return {"error": "Apenas listas PENDENTES podem ser editadas."}, 400

    item_global_id = data.get('item_global_id')
    if not item_global_id:
        return {"error": "item_global_id é obrigatório."}, 400

    # Verificar se item global existe
    item_global = ListaMaeItem.query.get(item_global_id)
    if not item_global:
        return {"error": "Item não encontrado no catálogo global."}, 404
    if restaurante_id is not None and item_global.restaurante_id != restaurante_id:
        return {"error": "Item não pertence ao restaurante da lista."}, 403

    # Verificar se já existe
    item_existente = ListaRapidaItem.query.filter_by(
        lista_rapida_id=lista_id,
        item_global_id=item_global_id
    ).first()

    if item_existente:
        return {"error": "Item já está na lista."}, 409

    # Criar item
    prioridade_str = data.get('prioridade', 'precisa_comprar')
    try:
        prioridade = PrioridadeItem(prioridade_str)
    except ValueError:
        prioridade = PrioridadeItem.PRECISA_COMPRAR

    novo_item = ListaRapidaItem(
        lista_rapida_id=lista_id,
        item_global_id=item_global_id,
        prioridade=prioridade,
        observacao=data.get('observacao', '').strip() or None
    )

    db.session.add(novo_item)
    db.session.commit()

    return {"message": "Item adicionado.", "item": novo_item.to_dict()}, 201


def remover_item_lista_rapida_admin(lista_id, item_id, restaurante_id=None):
    """Admin remove item da lista rápida PENDENTE."""
    from .models import ListaRapida, ListaRapidaItem, StatusListaRapida
    from .extensions import db

    lista = ListaRapida.query.filter_by(id=lista_id, deletado=False).first()

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if lista.status != StatusListaRapida.PENDENTE:
        return {"error": "Apenas listas PENDENTES podem ser editadas."}, 400

    item = ListaRapidaItem.query.filter_by(
        id=item_id,
        lista_rapida_id=lista_id
    ).first()

    if not item:
        return {"error": "Item não encontrado na lista."}, 404

    db.session.delete(item)
    db.session.commit()

    return {"message": "Item removido."}, 200


def editar_item_lista_rapida_admin(lista_id, item_id, data, restaurante_id=None):
    """Admin edita observação e prioridade de item da lista rápida PENDENTE."""
    from .models import ListaRapida, ListaRapidaItem, PrioridadeItem, StatusListaRapida
    from .extensions import db

    lista = ListaRapida.query.filter_by(id=lista_id, deletado=False).first()

    if not lista:
        return {"error": "Lista não encontrada."}, 404

    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro

    if lista.status != StatusListaRapida.PENDENTE:
        return {"error": "Apenas listas PENDENTES podem ser editadas."}, 400

    item = ListaRapidaItem.query.filter_by(
        id=item_id,
        lista_rapida_id=lista_id
    ).first()

    if not item:
        return {"error": "Item não encontrado na lista."}, 404

    # Atualizar observação
    if 'observacao' in data:
        item.observacao = data['observacao'].strip() or None

    # Atualizar prioridade
    if 'prioridade' in data:
        try:
            item.prioridade = PrioridadeItem(data['prioridade'])
        except ValueError:
            return {"error": "Prioridade inválida."}, 400

    db.session.commit()

    return {"message": "Item atualizado.", "item": item.to_dict()}, 200


# ==================== CHECKLIST SERVICES ====================

def converter_submissao_para_checklist(submissao_id: int, data: dict):
    """
    Converte uma submissão aprovada em checklist de compras.

    Args:
        submissao_id: ID da submissão
        data: {
            "incluir_fornecedor": bool,
            "incluir_observacoes": bool
        }
    """
    from kaizen_app.models import Submissao, SubmissaoStatus, Checklist, ChecklistItem, ChecklistStatus

    submissao = Submissao.query.get(submissao_id)
    if not submissao:
        raise ValueError("Submissão não encontrada")

    if submissao.status != SubmissaoStatus.APROVADO:
        raise ValueError("Apenas submissões aprovadas podem ser convertidas em checklist")

    # Buscar pedidos aprovados da submissão
    pedidos_aprovados = [p for p in submissao.pedidos if p.status.value == 'APROVADO']

    if not pedidos_aprovados:
        raise ValueError("Nenhum pedido aprovado encontrado nesta submissão")

    # Criar checklist
    checklist = Checklist(
        submissao_id=submissao_id,
        nome=f"Checklist - {submissao.lista.nome}",
        status=ChecklistStatus.ATIVO
    )
    db.session.add(checklist)
    db.session.flush()  # Obter ID do checklist

    # Criar itens do checklist
    incluir_fornecedor = data.get('incluir_fornecedor', False)
    incluir_observacoes = data.get('incluir_observacoes', False)

    for pedido in pedidos_aprovados:
        item = ChecklistItem(
            checklist_id=checklist.id,
            pedido_id=pedido.id,
            item_nome=pedido.item.nome,
            quantidade=pedido.quantidade_solicitada,
            unidade=pedido.item.unidade,
            fornecedor_nome=pedido.item.fornecedor.nome if incluir_fornecedor and pedido.item.fornecedor else None,
            observacao=None,
            marcado=False
        )
        db.session.add(item)

    db.session.commit()
    return checklist.to_dict()


def converter_lista_rapida_para_checklist(lista_rapida_id: int, data: dict):
    """
    Converte uma lista rápida aprovada em checklist de compras.

    Args:
        lista_rapida_id: ID da lista rápida
        data: {
            'nome': str (opcional),
            'itens': [{'item_id': int, 'quantidade': float ou None}],
            'incluir_fornecedor': bool,
            'incluir_observacoes': bool
        }
    """
    from kaizen_app.models import ListaRapida, StatusListaRapida, Checklist, ChecklistItem, ChecklistStatus

    lista_rapida = ListaRapida.query.get(lista_rapida_id)
    if not lista_rapida:
        raise ValueError("Lista rápida não encontrada")

    if lista_rapida.status != StatusListaRapida.APROVADA:
        raise ValueError("Apenas listas rápidas aprovadas podem ser convertidas em checklist")

    # Buscar itens da lista rápida
    itens_lista = lista_rapida.itens.all()
    if not itens_lista:
        raise ValueError("Lista rápida não possui itens")

    # Criar mapeamento de quantidades informadas pelo admin
    quantidades_map = {}
    itens_data = data.get('itens', [])
    for item_data in itens_data:
        item_id = item_data.get('item_id')
        quantidade = item_data.get('quantidade')
        if item_id:
            quantidades_map[item_id] = quantidade

    # Criar checklist
    nome_checklist = data.get('nome') or f"Checklist - {lista_rapida.nome}"
    checklist = Checklist(
        submissao_id=None,  # Não vem de submissão
        nome=nome_checklist,
        status=ChecklistStatus.ATIVO
    )
    db.session.add(checklist)
    db.session.flush()  # Obter ID do checklist

    # Criar itens do checklist
    incluir_fornecedor = data.get('incluir_fornecedor', False)
    incluir_observacoes = data.get('incluir_observacoes', False)

    for item_lista in itens_lista:
        # Pegar quantidade informada pelo admin (pode ser None)
        quantidade = quantidades_map.get(item_lista.id)

        checklist_item = ChecklistItem(
            checklist_id=checklist.id,
            pedido_id=None,  # Não vem de pedido
            item_nome=item_lista.item_global.nome,
            quantidade=quantidade if quantidade is not None else None,
            unidade=item_lista.item_global.unidade if quantidade is not None else None,
            fornecedor_nome=item_lista.item_global.fornecedor.nome if incluir_fornecedor and item_lista.item_global.fornecedor else None,
            observacao=item_lista.observacao if incluir_observacoes and item_lista.observacao else None,
            marcado=False
        )
        db.session.add(checklist_item)

    db.session.commit()

    # Retornar com dados enriquecidos
    dados = checklist.to_dict()
    itens = list(checklist.itens.order_by('item_nome'))

    # Incluir dados da lista rápida original
    dados['lista_rapida'] = {
        'id': lista_rapida.id,
        'nome': lista_rapida.nome
    }

    dados['itens'] = [item.to_dict() for item in itens]
    dados['total_itens'] = len(itens)
    dados['itens_marcados'] = sum(1 for item in itens if item.marcado)
    dados['progresso_percentual'] = 0

    return dados


def listar_checklists(status_filter=None):
    """Lista checklists com filtro opcional de status."""
    from kaizen_app.models import Checklist, ChecklistStatus

    query = Checklist.query

    if status_filter:
        try:
            status_enum = ChecklistStatus[status_filter]
            query = query.filter_by(status=status_enum)
        except KeyError:
            raise ValueError(f"Status inválido: {status_filter}")

    checklists = query.order_by(Checklist.criado_em.desc()).all()

    # Enriquecer com dados calculados
    resultado = []
    for checklist in checklists:
        dados = checklist.to_dict()
        itens = list(checklist.itens)
        total_itens = len(itens)
        itens_marcados = sum(1 for item in itens if item.marcado)

        # Incluir dados da submissao e lista
        if checklist.submissao:
            dados['submissao'] = {
                'lista': {
                    'nome': checklist.submissao.lista.nome
                }
            }

        dados['total_itens'] = total_itens
        dados['itens_marcados'] = itens_marcados
        dados['progresso_percentual'] = (itens_marcados / total_itens * 100) if total_itens > 0 else 0

        resultado.append(dados)

    return resultado


def obter_checklist_por_id(checklist_id: int):
    """Retorna checklist com todos os itens."""
    from kaizen_app.models import Checklist

    checklist = Checklist.query.get(checklist_id)
    if not checklist:
        raise ValueError("Checklist não encontrado")

    dados = checklist.to_dict()
    itens = list(checklist.itens.order_by('item_nome'))

    # Incluir dados da submissao e lista para exibição no frontend
    if checklist.submissao:
        dados['submissao'] = {
            'lista': {
                'nome': checklist.submissao.lista.nome
            }
        }

    dados['itens'] = [item.to_dict() for item in itens]
    dados['total_itens'] = len(itens)
    dados['itens_marcados'] = sum(1 for item in itens if item.marcado)
    dados['progresso_percentual'] = (dados['itens_marcados'] / dados['total_itens'] * 100) if dados['total_itens'] > 0 else 0

    return dados


def marcar_item_checklist(item_id: int, marcado: bool):
    """Marca ou desmarca um item do checklist."""
    from kaizen_app.models import ChecklistItem

    item = ChecklistItem.query.get(item_id)
    if not item:
        raise ValueError("Item não encontrado")

    item.marcado = marcado
    item.marcado_em = brasilia_now() if marcado else None

    db.session.commit()
    return item.to_dict()


def finalizar_checklist(checklist_id: int):
    """Finaliza um checklist ativo."""
    from kaizen_app.models import Checklist, ChecklistStatus

    checklist = Checklist.query.get(checklist_id)
    if not checklist:
        raise ValueError("Checklist não encontrado")

    if checklist.status == ChecklistStatus.FINALIZADO:
        raise ValueError("Checklist já está finalizado")

    checklist.status = ChecklistStatus.FINALIZADO
    checklist.finalizado_em = brasilia_now()

    db.session.commit()
    return checklist.to_dict()


def reabrir_checklist(checklist_id: int):
    """Reabre um checklist finalizado para permitir edições."""
    from kaizen_app.models import Checklist, ChecklistStatus

    checklist = Checklist.query.get(checklist_id)
    if not checklist:
        raise ValueError("Checklist não encontrado")

    if checklist.status == ChecklistStatus.ATIVO:
        raise ValueError("Checklist já está ativo")

    checklist.status = ChecklistStatus.ATIVO
    checklist.finalizado_em = None

    db.session.commit()
    return checklist.to_dict()


def compartilhar_checklist_whatsapp(checklist_id: int):
    """Gera texto formatado para compartilhar checklist via WhatsApp."""
    from kaizen_app.models import Checklist

    checklist = Checklist.query.get(checklist_id)
    if not checklist:
        raise ValueError("Checklist não encontrado")

    itens = list(checklist.itens.order_by('item_nome'))
    total_itens = len(itens)
    itens_marcados = sum(1 for item in itens if item.marcado)

    # Separar itens pendentes e concluídos
    itens_pendentes = [item for item in itens if not item.marcado]
    itens_concluidos = [item for item in itens if item.marcado]

    # Montar texto com formatação WhatsApp
    linhas = [
        f"*CHECKLIST DE COMPRAS*",
        f"*{checklist.nome}*",
        f"Data: {checklist.criado_em.strftime('%d/%m/%Y %H:%M')}",
        "",
    ]

    # EDGE CASE: Checklist vazio
    if total_itens == 0:
        linhas.append("_Nenhum item no checklist_")
        linhas.append("")
        linhas.append(f"*Progresso: 0/0 itens*")
        return {"texto": "\n".join(linhas)}

    # Seção de itens PENDENTES
    if itens_pendentes:
        linhas.append("*📋 ITENS PENDENTES*")
        for item in itens_pendentes:
            linha = f"[ ] {item.item_nome}"

            if item.quantidade is not None:
                qtd_texto = f"{float(item.quantidade):.1f}{item.unidade}"
                linha += f" - {qtd_texto}"

            if item.fornecedor_nome:
                linha += f" ({item.fornecedor_nome})"

            linhas.append(linha)
        linhas.append("")

    # Seção de itens CONCLUÍDOS
    if itens_concluidos:
        linhas.append("*✅ ITENS CONCLUÍDOS*")
        for item in itens_concluidos:
            # Construir linha completa ANTES de riscar
            linha_completa = f"[x] {item.item_nome}"

            if item.quantidade is not None:
                qtd_texto = f"{float(item.quantidade):.1f}{item.unidade}"
                linha_completa += f" - {qtd_texto}"

            if item.fornecedor_nome:
                linha_completa += f" ({item.fornecedor_nome})"

            # Riscar TODA a linha (nome + quantidade + fornecedor)
            linhas.append(f"~{linha_completa}~")
        linhas.append("")

    # Progresso com percentual
    progresso_percentual = (itens_marcados / total_itens * 100) if total_itens > 0 else 0
    linhas.append(f"*Progresso: {itens_marcados}/{total_itens} itens ({progresso_percentual:.0f}%)*")

    return {"texto": "\n".join(linhas)}


def compartilhar_lista_rapida_whatsapp(lista_rapida_id: int):
    """Gera texto formatado para compartilhar lista rápida via WhatsApp."""
    from kaizen_app.models import ListaRapida, ListaRapidaItem, PrioridadeItem

    lista = ListaRapida.query.get(lista_rapida_id)
    if not lista:
        raise ValueError("Lista rápida não encontrada")

    # Filtrar apenas itens não descartados
    itens = list(lista.itens.filter_by(descartado=False).order_by(ListaRapidaItem.prioridade.desc(), ListaRapidaItem.item_global_id))
    total_itens = len(itens)

    # Montar texto com formatação WhatsApp
    linhas = [
        f"*LISTA DE COMPRAS*",
        f"*{lista.nome}*",
        f"Data: {lista.criado_em.strftime('%d/%m/%Y %H:%M')}",
        "",
    ]

    for item in itens:
        # Prioridade como emoji
        if item.prioridade == PrioridadeItem.URGENTE:
            prioridade = "🔴"
        elif item.prioridade == PrioridadeItem.PRECISA_COMPRAR:
            prioridade = "🟡"
        else:  # PREVENCAO
            prioridade = "🟢"

        # Nome do item
        item_nome = item.item_global.nome if item.item_global else "Item desconhecido"
        linha = f"[ ] {prioridade} {item_nome}"

        # Adicionar unidade
        if item.item_global and item.item_global.unidade:
            linha += f" ({item.item_global.unidade})"

        # Adicionar observação se houver
        if item.observacao:
            linha += f"\n    _{item.observacao}_"

        linhas.append(linha)

    linhas.append("")
    linhas.append(f"*Total: {total_itens} itens*")

    return {"texto": "\n".join(linhas)}


# ========================================
# FUNDIR SUBMISSÕES (MERGE)
# ========================================

def merge_submissoes_preview(submissao_ids: list, restaurante_id=None):
    """
    Funde múltiplas submissões APROVADAS em um preview de pedido único.
    Agrupa por lista_mae_item_id, soma quantidade_solicitada
    (somente pedidos com PedidoStatus.APROVADO).
    """
    if not submissao_ids or len(submissao_ids) < 2:
        return {"error": "Informe ao menos 2 submissões para fundir."}, 400

    submissoes = Submissao.query.filter(Submissao.id.in_(submissao_ids)).all()

    if len(submissoes) != len(submissao_ids):
        ids_encontrados = {s.id for s in submissoes}
        ids_faltando = [i for i in submissao_ids if i not in ids_encontrados]
        return {"error": f"Submissões não encontradas: {ids_faltando}"}, 404

    # Validar que todas estão APROVADAS
    nao_aprovadas = [s for s in submissoes if s.status != SubmissaoStatus.APROVADO]
    if nao_aprovadas:
        ids_invalidas = [s.id for s in nao_aprovadas]
        return {
            "error": f"Submissões {ids_invalidas} não estão APROVADAS. "
                     "Apenas submissões aprovadas podem ser fundidas."
        }, 422

    # Validar restaurante (segurança multi-tenant)
    if restaurante_id:
        fora_do_restaurante = [
            s for s in submissoes
            if s.lista and s.lista.restaurante_id != restaurante_id
        ]
        if fora_do_restaurante:
            ids_fora = [s.id for s in fora_do_restaurante]
            return {"error": f"Acesso negado. Submissões {ids_fora} pertencem a outro restaurante."}, 403

    # ---- MERGE LOGIC ----
    # { item_id: { "item_id", "item_nome", "item_unidade", "quantidade_total", "breakdown" } }
    merged = {}

    for submissao in submissoes:
        pedidos_aprovados = [p for p in submissao.pedidos if p.status == PedidoStatus.APROVADO]

        for pedido in pedidos_aprovados:
            item_id = pedido.lista_mae_item_id
            item = pedido.item  # ListaMaeItem via relationship

            if item_id not in merged:
                merged[item_id] = {
                    "item_id": item_id,
                    "item_nome": item.nome if item else f"Item #{item_id}",
                    "item_unidade": item.unidade if item else "un",
                    "quantidade_total": 0.0,
                    "breakdown": []
                }

            qtd = float(pedido.quantidade_solicitada)
            merged[item_id]["quantidade_total"] += qtd
            merged[item_id]["breakdown"].append({
                "submissao_id": submissao.id,
                "lista_id": submissao.lista_id,
                "lista_nome": submissao.lista.nome if submissao.lista else f"Lista #{submissao.lista_id}",
                "usuario_nome": submissao.usuario.nome if submissao.usuario else "Desconhecido",
                "quantidade": qtd
            })

    itens_fundidos = sorted(merged.values(), key=lambda x: x["item_nome"])

    return {
        "submissao_ids": submissao_ids,
        "listas": [
            {
                "submissao_id": s.id,
                "lista_id": s.lista_id,
                "lista_nome": s.lista.nome if s.lista else f"Lista #{s.lista_id}"
            }
            for s in submissoes
        ],
        "itens": itens_fundidos,
        "total_itens": len(itens_fundidos),
    }, 200


def merge_submissoes_whatsapp(submissao_ids: list, restaurante_id=None):
    """Gera texto WhatsApp formatado para o pedido fundido."""
    preview, status = merge_submissoes_preview(submissao_ids, restaurante_id)
    if status != 200:
        return preview, status

    listas_nomes = " + ".join(l["lista_nome"] for l in preview["listas"])
    linhas = [
        "*📋 PEDIDO FUNDIDO*",
        f"*Listas:* {listas_nomes}",
        f"*Data:* {brasilia_now().strftime('%d/%m/%Y %H:%M')}",
        "",
        "*Itens:*",
    ]

    for item in preview["itens"]:
        qtd = item["quantidade_total"]
        qtd_str = f"{qtd:.0f}" if qtd == int(qtd) else f"{qtd:.2f}"
        linhas.append(f"• {item['item_nome']} — *{qtd_str} {item['item_unidade']}*")

    linhas += [
        "",
        f"*Total: {preview['total_itens']} itens*",
        "---",
        "Sistema Kaizen"
    ]

    return {"texto": "\n".join(linhas)}, 200


# ========================================
# RESTAURANTES (Multi-Tenant)
# ========================================

def listar_restaurantes():
    """Lista todos os restaurantes (apenas SUPER_ADMIN) com credenciais do admin."""
    from .models import UserRoles, SolicitacaoRestaurante

    restaurantes = Restaurante.query.filter_by(deletado=False).order_by(Restaurante.nome).all()

    restaurantes_com_credenciais = []
    for r in restaurantes:
        restaurante_dict = r.to_dict()

        # Buscar admin do restaurante (usuário com role ADMIN)
        admin = Usuario.query.filter_by(
            restaurante_id=r.id,
            role=UserRoles.ADMIN,
            ativo=True
        ).first()

        if admin:
            restaurante_dict['usuario_admin_email'] = admin.email

            # Buscar senha armazenada na solicitação de origem
            solicitacao = SolicitacaoRestaurante.query.filter_by(
                restaurante_criado_id=r.id
            ).first()

            if solicitacao and solicitacao.senha_gerada:
                restaurante_dict['usuario_admin_senha'] = solicitacao.senha_gerada
            else:
                # Se não encontrar na solicitação, senha não está disponível
                restaurante_dict['usuario_admin_senha'] = '(Definida no cadastro - não armazenada)'
        else:
            restaurante_dict['usuario_admin_email'] = 'Nenhum admin atribuído'
            restaurante_dict['usuario_admin_senha'] = 'N/A'

        restaurantes_com_credenciais.append(restaurante_dict)

    return {
        "restaurantes": restaurantes_com_credenciais,
        "total": len(restaurantes_com_credenciais)
    }, 200


def obter_restaurante(restaurante_id):
    """Obtém um restaurante por ID."""
    restaurante = Restaurante.query.filter_by(id=restaurante_id, deletado=False).first()
    if not restaurante:
        return {"error": "Restaurante não encontrado."}, 404
    return {"restaurante": restaurante.to_dict()}, 200


def criar_restaurante(data):
    """Cria um novo restaurante."""
    from slugify import slugify
    from sqlalchemy import func

    nome = data.get('nome', '').strip()

    if not nome:
        return {"error": "Nome do restaurante é obrigatório."}, 400

    if len(nome) < 3:
        return {"error": "Nome deve ter pelo menos 3 caracteres."}, 400

    existente = Restaurante.query.filter(
        func.lower(Restaurante.nome) == func.lower(nome)
    ).first()

    if existente:
        return {"error": f"Já existe um restaurante com o nome '{nome}'."}, 409

    slug_base = slugify(nome)
    slug = slug_base
    contador = 1

    while Restaurante.query.filter_by(slug=slug).first():
        slug = f"{slug_base}-{contador}"
        contador += 1

    nome = nome.strip()[:200]

    restaurante = Restaurante(nome=nome, slug=slug, ativo=True)
    db.session.add(restaurante)
    db.session.flush()
    log_event(
        acao="create",
        entidade="restaurante",
        entidade_id=restaurante.id,
        mensagem="Restaurante criado",
        restaurante_id=restaurante.id
    )
    db.session.commit()

    return {"message": "Restaurante criado com sucesso.", "restaurante": restaurante.to_dict()}, 201


def atualizar_restaurante(restaurante_id, data):
    """Atualiza dados de um restaurante."""
    from slugify import slugify

    restaurante = Restaurante.query.filter_by(id=restaurante_id, deletado=False).first()
    if not restaurante:
        return {"error": "Restaurante não encontrado."}, 404

    nome = data.get('nome', '').strip()

    if not nome:
        return {"error": "Nome do restaurante é obrigatório."}, 400

    restaurante.nome = nome

    novo_slug = slugify(nome)
    if novo_slug != restaurante.slug:
        slug = novo_slug
        contador = 1
        while Restaurante.query.filter_by(slug=slug).filter(Restaurante.id != restaurante_id).first():
            slug = f"{novo_slug}-{contador}"
            contador += 1
        restaurante.slug = slug

    if 'ativo' in data:
        restaurante.ativo = bool(data['ativo'])

    log_event(
        acao="update",
        entidade="restaurante",
        entidade_id=restaurante.id,
        mensagem="Restaurante atualizado",
        restaurante_id=restaurante.id
    )
    db.session.commit()

    return {
        "message": "Restaurante atualizado com sucesso.",
        "restaurante": restaurante.to_dict()
    }, 200


def deletar_restaurante(restaurante_id):
    """Hard delete de um restaurante (remoção física do banco)."""
    from .models import Restaurante, Usuario, Fornecedor, ListaMaeItem, Item, Estoque, Pedido

    restaurante = Restaurante.query.filter_by(id=restaurante_id).first()
    if not restaurante:
        return {"error": "Restaurante não encontrado."}, 404

    try:
        # 1. Deletar Pedidos relacionados aos fornecedores e itens do restaurante
        fornecedores_ids = [f.id for f in Fornecedor.query.filter_by(restaurante_id=restaurante_id).all()]
        lista_mae_itens_ids = [item.id for item in ListaMaeItem.query.filter_by(restaurante_id=restaurante_id).all()]

        if fornecedores_ids:
            Pedido.query.filter(Pedido.fornecedor_id.in_(fornecedores_ids)).delete(synchronize_session=False)
        if lista_mae_itens_ids:
            Pedido.query.filter(Pedido.lista_mae_item_id.in_(lista_mae_itens_ids)).delete(synchronize_session=False)

        # 2. Deletar Estoques dos Itens que pertencem aos Fornecedores do restaurante
        if fornecedores_ids:
            itens_ids = [item.id for item in Item.query.filter(Item.fornecedor_id.in_(fornecedores_ids)).all()]
            if itens_ids:
                Estoque.query.filter(Estoque.item_id.in_(itens_ids)).delete(synchronize_session=False)
                # 3. Deletar Itens que pertencem aos Fornecedores
                Item.query.filter(Item.fornecedor_id.in_(fornecedores_ids)).delete(synchronize_session=False)

        # 4. Deletar ListaMaeItem do restaurante
        ListaMaeItem.query.filter_by(restaurante_id=restaurante_id).delete()

        # 5. Deletar Fornecedores do restaurante
        Fornecedor.query.filter_by(restaurante_id=restaurante_id).delete()

        # 6. Deletar Usuários do restaurante
        Usuario.query.filter_by(restaurante_id=restaurante_id).delete()

        # 7. Deletar Restaurante
        log_event(
            acao="delete",
            entidade="restaurante",
            entidade_id=restaurante.id,
            mensagem="Restaurante deletado",
            restaurante_id=restaurante.id
        )
        db.session.delete(restaurante)

        db.session.commit()
        return {"message": "Restaurante e todos os dados relacionados foram deletados permanentemente."}, 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERRO] deletar_restaurante: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": f"Erro ao deletar restaurante: {str(e)}"}, 500


def listar_usuarios_restaurante(restaurante_id, admin_id):
    """Lista todos os usuários de um restaurante com suas senhas."""
    from .models import Usuario

    # Validar permissão
    admin = Usuario.query.get(admin_id)
    if admin.role != UserRoles.SUPER_ADMIN:
        if restaurante_id != admin.restaurante_id:
            return {"error": "Você não tem permissão para acessar usuários deste restaurante."}, 403

    usuarios = Usuario.query.filter_by(restaurante_id=restaurante_id).all()

    return {
        "usuarios": [u.to_dict_with_password() for u in usuarios],
        "total": len(usuarios)
    }, 200


def alterar_senha_usuario(usuario_id, nova_senha, admin_id):
    """Altera manualmente a senha de um usuário (admin digita nova senha)."""
    from .models import Usuario

    usuario = Usuario.query.get(usuario_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    # Validar que admin tem permissão (mesmo restaurante ou SUPER_ADMIN)
    admin = Usuario.query.get(admin_id)
    if admin.role != UserRoles.SUPER_ADMIN:
        if usuario.restaurante_id != admin.restaurante_id:
            return {"error": "Você não tem permissão para alterar este usuário."}, 403

    # Atualizar senha
    usuario.senha_hash = generate_password_hash(nova_senha)
    usuario.senha_texto_puro = nova_senha
    db.session.commit()

    return {
        "message": "Senha alterada com sucesso.",
        "usuario": usuario.to_dict_with_password()
    }, 200


def resetar_senha_usuario(usuario_id, admin_id):
    """Reseta senha para senha aleatória de 12 caracteres."""
    from .models import Usuario
    import secrets
    import string

    usuario = Usuario.query.get(usuario_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    # Validar permissão
    admin = Usuario.query.get(admin_id)
    if admin.role != UserRoles.SUPER_ADMIN:
        if usuario.restaurante_id != admin.restaurante_id:
            return {"error": "Você não tem permissão para resetar este usuário."}, 403

    # Gerar nova senha aleatória
    caracteres = string.ascii_letters + string.digits
    nova_senha = ''.join(secrets.choice(caracteres) for _ in range(12))

    # Atualizar senha
    usuario.senha_hash = generate_password_hash(nova_senha)
    usuario.senha_texto_puro = nova_senha
    db.session.commit()

    return {
        "message": "Senha resetada com sucesso.",
        "usuario": usuario.to_dict_with_password(),
        "nova_senha": nova_senha
    }, 200


# ========================================
# USUÁRIOS (Multi-Tenant)
# ========================================

def atribuir_usuario_restaurante(usuario_id, restaurante_id):
    """Atribui um usuário a um restaurante."""
    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    if usuario.role == UserRoles.SUPER_ADMIN:
        return {"error": "SUPER_ADMIN não pode ser atribuído a restaurante."}, 400

    restaurante = Restaurante.query.filter_by(id=restaurante_id, deletado=False).first()
    if not restaurante:
        return {"error": "Restaurante não encontrado."}, 404

    usuario.restaurante_id = restaurante.id
    db.session.commit()

    return {
        "message": "Restaurante atribuído com sucesso.",
        "usuario": usuario.to_dict()
    }, 200


def criar_admin_restaurante(data):
    """SUPER_ADMIN cria um novo ADMIN vinculado a restaurante."""
    try:
        nome = data.get('nome', '').strip()
        email = data.get('email', '').strip()
        senha = data.get('senha')
        restaurante_id = data.get('restaurante_id')

        if not nome or not email or not senha or not restaurante_id:
            return {"error": "Nome, email, senha e restaurante são obrigatórios."}, 400

        if Usuario.query.filter_by(email=email).first():
            return {"error": "E-mail já cadastrado."}, 409

        restaurante = Restaurante.query.filter_by(id=restaurante_id, deletado=False).first()
        if not restaurante:
            return {"error": "Restaurante não encontrado."}, 404

        admin = Usuario(
            nome=nome,
            email=email,
            senha_hash=generate_password_hash(senha),
            role=UserRoles.ADMIN,
            aprovado=True,
            ativo=True,
            restaurante_id=restaurante.id
        )

        db.session.add(admin)
        db.session.flush()
        log_event(
            acao="create",
            entidade="usuario",
            entidade_id=admin.id,
            mensagem="Admin criado para restaurante",
            restaurante_id=restaurante.id,
            usuario_id=admin.id
        )
        db.session.commit()

        return {"message": "Admin criado com sucesso.", "usuario": admin.to_dict()}, 201

    except Exception as e:
        db.session.rollback()
        return {"error": f"Erro ao criar admin: {str(e)}"}, 500


def descartar_item_lista_rapida(lista_id, item_id, restaurante_id=None):
    """Admin marca um item como descartado (suficiente/rejeitado)."""
    from .models import ListaRapida, ListaRapidaItem
    from .extensions import db
    
    lista = ListaRapida.query.get(lista_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    
    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro
    
    item = ListaRapidaItem.query.filter_by(
        id=item_id,
        lista_rapida_id=lista_id
    ).first()
    
    if not item:
        return {"error": "Item não encontrado nesta lista."}, 404
    
    item.descartado = True
    db.session.commit()
    
    return {
        "message": "Item marcado como descartado.",
        "item": item.to_dict()
    }, 200


def restaurar_item_lista_rapida(lista_id, item_id, restaurante_id=None):
    """Admin remove a marca de descartado de um item."""
    from .models import ListaRapida, ListaRapidaItem
    from .extensions import db
    
    lista = ListaRapida.query.get(lista_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    
    erro = _validar_lista_rapida_restaurante(lista, restaurante_id)
    if erro:
        return erro
    
    item = ListaRapidaItem.query.filter_by(
        id=item_id,
        lista_rapida_id=lista_id
    ).first()
    
    if not item:
        return {"error": "Item não encontrado nesta lista."}, 404
    
    item.descartado = False
    db.session.commit()
    
    return {
        "message": "Item restaurado.",
        "item": item.to_dict()
    }, 200


def reverter_pedido(pedido_id):
    """Reverte um pedido rejeitado para PENDENTE."""
    pedido = repositories.get_by_id(Pedido, pedido_id)

    if not pedido:
        return {"error": "Pedido não encontrado."}, 404

    if pedido.status != PedidoStatus.REJEITADO:
        return {"error": f"Apenas pedidos rejeitados podem ser revertidos. Status atual: {pedido.status.value}"}, 400

    pedido.status = PedidoStatus.PENDENTE
    db.session.commit()

    return {"message": "Pedido revertido para PENDENTE com sucesso.", "pedido": pedido.to_dict()}, 200


def criar_notificacao(usuario_id, titulo, tipo, mensagem=None, lista_id=None, lista_rapida_id=None, submissao_id=None, sugestao_id=None, pedido_id=None, restaurante_id=None):
    """Cria uma notificação para um usuário."""
    from .models import Notificacao, TipoNotificacao, Usuario
    from .extensions import db

    try:
        # Se restaurante_id não foi passado, inferir do usuário
        if restaurante_id is None:
            usuario = Usuario.query.get(usuario_id)
            if usuario:
                restaurante_id = usuario.restaurante_id

        # Converter string para enum se necessário
        if isinstance(tipo, str):
            tipo = TipoNotificacao(tipo)

        notificacao = Notificacao(
            usuario_id=usuario_id,
            titulo=titulo,
            mensagem=mensagem,
            tipo=tipo,
            restaurante_id=restaurante_id,
            lista_id=lista_id,
            lista_rapida_id=lista_rapida_id,
            submissao_id=submissao_id,
            sugestao_id=sugestao_id,
            pedido_id=pedido_id
        )

        db.session.add(notificacao)
        db.session.commit()
        return notificacao
    except Exception as e:
        db.session.rollback()
        print(f"[ERRO] criar_notificacao: {str(e)}")
        raise


def listar_notificacoes_usuario(usuario_id, restaurante_id=None, nao_lidas_primeiro=True, limit=50):
    """Lista notificações de um usuário filtradas por restaurante."""
    from .models import Notificacao

    query = Notificacao.query.filter_by(usuario_id=usuario_id)

    # Filtro por restaurante
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    # SUPER_ADMIN (restaurante_id=None) vê apenas notificações globais
    else:
        query = query.filter_by(restaurante_id=None)

    if nao_lidas_primeiro:
        # Ordenar não lidas primeiro, depois por data
        query = query.order_by(
            Notificacao.lida.asc(),  # False (não lidas) vem primeiro
            Notificacao.criado_em.desc()
        )
    else:
        query = query.order_by(Notificacao.criado_em.desc())

    notificacoes = query.limit(limit).all()

    return {
        "notificacoes": [n.to_dict() for n in notificacoes],
        "total": len(notificacoes),
        "nao_lidas": sum(1 for n in notificacoes if not n.lida)
    }, 200


def marcar_notificacao_lida(notificacao_id, usuario_id):
    """Marca uma notificação como lida."""
    from .models import Notificacao
    from .extensions import db

    notificacao = Notificacao.query.filter_by(id=notificacao_id, usuario_id=usuario_id).first()

    if not notificacao:
        return {"error": "Notificação não encontrada."}, 404

    if notificacao.lida:
        return {"message": "Notificação já estava lida."}, 200

    notificacao.lida = True
    notificacao.lido_em = brasilia_now()
    db.session.commit()

    return {"message": "Notificação marcada como lida.", "notificacao": notificacao.to_dict()}, 200


def marcar_todas_notificacoes_lidas(usuario_id):
    """Marca todas as notificações de um usuário como lidas."""
    from .models import Notificacao
    from .extensions import db

    notificacoes = Notificacao.query.filter_by(usuario_id=usuario_id, lida=False).all()

    for n in notificacoes:
        n.lida = True
        n.lido_em = brasilia_now()

    db.session.commit()

    return {
        "message": f"{len(notificacoes)} notificação(ões) marcada(s) como lida(s).",
        "total_marcadas": len(notificacoes)
    }, 200


# ===== CONVITE DE USUÁRIOS =====

def criar_convite_usuario(admin_id: int, role: str, admin_role: str, restaurante_id: int = None):
    """Cria um novo token de convite para registro de usuário."""
    import os
    from flask import current_app

    # Validar role
    role_upper = role.upper()
    if role_upper not in ['ADMIN', 'COLLABORATOR', 'SUPER_ADMIN']:
        return {"error": "Role inválido. Use ADMIN, COLLABORATOR ou SUPER_ADMIN"}, 400

    admin_role_upper = (admin_role or '').upper()

    if admin_role_upper == 'SUPER_ADMIN':
        if not restaurante_id:
            return {"error": "SUPER_ADMIN deve especificar o restaurante"}, 400

        try:
            restaurante_id = int(restaurante_id)
        except (TypeError, ValueError):
            return {"error": "restaurante_id inválido"}, 400

        restaurante = Restaurante.query.filter_by(id=restaurante_id, deletado=False).first()
        if not restaurante:
            return {"error": "Restaurante não encontrado"}, 404

        final_restaurante_id = restaurante_id
    else:
        final_restaurante_id = _get_usuario_restaurante_id(admin_id)
        if final_restaurante_id is None:
            final_restaurante_id = _get_restaurante_padrao_id()

    # Converter string para enum
    role_enum = UserRoles[role_upper]

    # Gerar token único
    token = ConviteToken.gerar_token()

    # Criar convite
    convite = ConviteToken(
        token=token,
        role=role_enum,
        criado_por_id=admin_id,
        restaurante_id=final_restaurante_id
    )

    db.session.add(convite)
    db.session.commit()

    # Gerar link completo
    base_url = os.environ.get('FRONTEND_URL', 'https://kaizen-compras.up.railway.app')
    link_convite = f"{base_url}/convite?token={token}"

    return {
        "id": convite.id,
        "token": token,
        "link": link_convite,
        "role": role_upper,
        "restaurante_id": final_restaurante_id,
        "restaurante_nome": convite.restaurante.nome if convite.restaurante else None,
        "criado_em": convite.criado_em.strftime('%d/%m/%Y %H:%M'),
        "message": "Convite criado com sucesso!"
    }, 201


def listar_convites(admin_id: int, admin_role: str):
    """Lista convites criados (todos para SUPER_ADMIN, só seus para ADMIN)."""
    # SUPER_ADMIN vê todos, ADMIN vê só os seus
    if admin_role == 'SUPER_ADMIN':
        convites = ConviteToken.query.order_by(ConviteToken.criado_em.desc()).all()
    else:
        convites = ConviteToken.query.filter_by(criado_por_id=admin_id)\
                                      .order_by(ConviteToken.criado_em.desc()).all()

    import os
    base_url = os.environ.get('FRONTEND_URL', 'https://kaizen-compras.up.railway.app')
    return [
        {
            **convite.to_dict(),
            "link": f"{base_url}/convite?token={convite.token}"
        }
        for convite in convites
    ]


def excluir_convite(admin_id: int, admin_role: str, convite_id: int):
    """Exclui um convite pelo ID (SUPER_ADMIN pode excluir qualquer, ADMIN só os seus)."""
    if admin_role == 'SUPER_ADMIN':
        convite = ConviteToken.query.filter_by(id=convite_id).first()
    else:
        convite = ConviteToken.query.filter_by(id=convite_id, criado_por_id=admin_id).first()

    if not convite:
        return {"error": "Convite não encontrado"}, 404

    db.session.delete(convite)
    db.session.commit()

    return {"message": "Convite excluído com sucesso!"}, 200


def validar_token_convite(token: str):
    """Valida se um token de convite é válido."""
    convite = ConviteToken.query.filter_by(token=token).first()

    if not convite:
        return {"valido": False, "erro": "Token não encontrado"}, 400

    if not convite.esta_valido():
        if convite.usado:
            return {"valido": False, "erro": "Token já foi utilizado"}, 400
        else:
            return {"valido": False, "erro": "Token expirado"}, 400

    return {
        "valido": True,
        "role": convite.role.value,
        "criado_em": convite.criado_em.strftime('%d/%m/%Y')
    }, 200


def register_user_com_convite(data):
    """Registra usuário usando token de convite (auto-aprovado)."""
    # Validar dados
    if not data.get('email') or not data.get('senha') or not data.get('nome'):
        return {"error": "Dados incompletos para registro."}, 400

    if not data.get('token_convite'):
        return {"error": "Token de convite não fornecido."}, 400

    # Verificar se email já existe
    if Usuario.query.filter_by(email=data['email']).first():
        return {"error": "E-mail já cadastrado."}, 409

    # Buscar token de convite
    convite = ConviteToken.query.filter_by(token=data['token_convite']).first()

    if not convite:
        return {"error": "Token de convite inválido."}, 400

    if not convite.esta_valido():
        if convite.usado:
            return {"error": "Este convite já foi utilizado."}, 400
        else:
            return {"error": "Este convite expirou."}, 400

    # Hash da senha
    hashed_password = generate_password_hash(data['senha'])

    # Criar usuário com role do convite
    new_user = Usuario(
        nome=data['nome'],
        email=data['email'],
        senha_hash=hashed_password,
        role=convite.role,  # Role vem do convite
        aprovado=True,  # Auto-aprovado
        restaurante_id=convite.restaurante_id or _get_restaurante_padrao_id()
    )

    db.session.add(new_user)
    db.session.flush()  # Flush para obter o ID do novo usuário
    log_event(
        acao="create",
        entidade="usuario",
        entidade_id=new_user.id,
        mensagem="Cadastro de usuario via convite",
        restaurante_id=new_user.restaurante_id,
        usuario_id=new_user.id
    )

    # Marcar convite como usado
    convite.usado = True
    convite.usado_em = brasilia_now()
    convite.usado_por_id = new_user.id

    db.session.commit()

    return {
        "message": f"Cadastro realizado com sucesso! Você foi registrado como {convite.role.value}.",
        "aprovado": True
    }, 201


def criar_convite_restaurante(admin_id: int, limite_usos: int = 1):
    """Cria um novo token de convite para cadastro de restaurante."""
    import os

    admin = Usuario.query.get(admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode criar convites de restaurante."}, 403

    # Valida limite_usos (1-100)
    if not isinstance(limite_usos, int) or limite_usos < 1 or limite_usos > 100:
        return {"error": "Limite de usos deve ser um número entre 1 e 100."}, 400

    token = ConviteRestaurante.gerar_token()
    expira_em = brasilia_now() + timedelta(hours=72)

    convite = ConviteRestaurante(
        token=token,
        criado_por_id=admin_id,
        expira_em=expira_em,
        limite_usos=limite_usos,
        quantidade_usos=0
    )

    db.session.add(convite)
    db.session.commit()

    base_url = os.environ.get('FRONTEND_URL', 'https://kaizen-compras.up.railway.app')
    link_convite = f"{base_url}/register-restaurant?token={token}"

    return {
        "id": convite.id,
        "token": token,
        "link": link_convite,
        "criado_em": convite.criado_em.strftime('%d/%m/%Y %H:%M'),
        "expira_em": convite.expira_em.strftime('%d/%m/%Y %H:%M') if convite.expira_em else None,
        "limite_usos": limite_usos,
        "message": "Convite de restaurante criado com sucesso!"
    }, 201


def listar_convites_restaurante(admin_id: int):
    """Lista convites de restaurante (apenas SUPER_ADMIN)."""
    import os

    admin = Usuario.query.get(admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode listar convites de restaurante."}, 403

    convites = ConviteRestaurante.query.order_by(ConviteRestaurante.criado_em.desc()).all()
    base_url = os.environ.get('FRONTEND_URL', 'https://kaizen-compras.up.railway.app')

    return {
        "convites": [
            {
                **convite.to_dict(),
                "link": f"{base_url}/register-restaurant?token={convite.token}"
            }
            for convite in convites
        ],
        "total": len(convites)
    }, 200


def excluir_convite_restaurante(admin_id: int, convite_id: int):
    """Exclui um convite de restaurante (apenas SUPER_ADMIN)."""
    admin = Usuario.query.get(admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode excluir convites de restaurante."}, 403

    convite = ConviteRestaurante.query.filter_by(id=convite_id).first()
    if not convite:
        return {"error": "Convite não encontrado."}, 404

    db.session.delete(convite)
    db.session.commit()

    return {"message": "Convite excluído com sucesso!"}, 200


def editar_convite_restaurante(admin_id: int, convite_id: int, data: dict):
    """Edita um convite de restaurante (limite_usos e ativo)."""
    import os

    admin = Usuario.query.get(admin_id)
    if not admin or admin.role != UserRoles.SUPER_ADMIN:
        return {"error": "Apenas SUPER_ADMIN pode editar convites de restaurante."}, 403

    convite = ConviteRestaurante.query.filter_by(id=convite_id).first()
    if not convite:
        return {"error": "Convite não encontrado."}, 404

    # Atualiza limite_usos se fornecido
    if 'limite_usos' in data:
        novo_limite = data['limite_usos']
        try:
            novo_limite = int(novo_limite)
        except (ValueError, TypeError):
            return {"error": "Limite de usos deve ser um número."}, 400

        if novo_limite < 1 or novo_limite > 100:
            return {"error": "Limite de usos deve ser entre 1 e 100."}, 400

        # Não permite diminuir abaixo da quantidade já usada
        if novo_limite < convite.quantidade_usos:
            return {"error": f"Limite não pode ser menor que a quantidade já usada ({convite.quantidade_usos})."}, 400

        convite.limite_usos = novo_limite

    # Atualiza ativo se fornecido
    if 'ativo' in data:
        convite.ativo = bool(data['ativo'])

    db.session.commit()

    base_url = os.environ.get('FRONTEND_URL', 'https://kaizen-compras.up.railway.app')

    return {
        "message": "Convite atualizado com sucesso!",
        "convite": {
            **convite.to_dict(),
            "link": f"{base_url}/register-restaurant?token={convite.token}"
        }
    }, 200


def validar_convite_restaurante(token: str):
    """Valida se um token de convite de restaurante é válido."""
    convite = ConviteRestaurante.query.filter_by(token=token).first()

    if not convite:
        return {"valido": False, "erro": "Token não encontrado"}, 400

    if not convite.esta_valido():
        if convite.usado:
            return {"valido": False, "erro": "Token já foi utilizado"}, 400
        return {"valido": False, "erro": "Token expirado"}, 400

    return {
        "valido": True,
        "criado_em": convite.criado_em.strftime('%d/%m/%Y'),
        "expira_em": convite.expira_em.strftime('%d/%m/%Y %H:%M') if convite.expira_em else None
    }, 200


def registrar_restaurante_com_convite(data):
    """Registra restaurante usando token de convite (cadastro imediato)."""
    from .models import Restaurante
    from slugify import slugify
    import secrets
    import string

    required_fields = [
        'nome_restaurante', 'endereco_restaurante', 'telefone_restaurante',
        'email_restaurante', 'nome_responsavel', 'email_responsavel',
        'telefone_responsavel', 'token_convite'
    ]

    missing_fields = [f for f in required_fields if not data.get(f)]
    if missing_fields:
        return {
            "error": f"Dados incompletos. Campos obrigatórios ausentes: {', '.join(missing_fields)}"
        }, 400

    email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    if not re.match(email_regex, data['email_restaurante'].strip()):
        return {"error": "Email do restaurante inválido."}, 400
    if not re.match(email_regex, data['email_responsavel'].strip()):
        return {"error": "Email do responsável inválido."}, 400

    telefone_restaurante_limpo = re.sub(r'\D', '', data['telefone_restaurante'])
    telefone_responsavel_limpo = re.sub(r'\D', '', data['telefone_responsavel'])
    if len(telefone_restaurante_limpo) < 10:
        return {"error": "Telefone do restaurante deve ter pelo menos 10 dígitos."}, 400
    if len(telefone_responsavel_limpo) < 10:
        return {"error": "Telefone do responsável deve ter pelo menos 10 dígitos."}, 400

    cnpj = (data.get('cnpj') or '').strip()
    if cnpj and len(re.sub(r'\D', '', cnpj)) != 14:
        return {"error": "CNPJ deve conter 14 dígitos."}, 400

    convite = ConviteRestaurante.query.filter_by(token=data['token_convite']).first()
    if not convite:
        return {"error": "Token de convite inválido."}, 400
    if not convite.esta_valido():
        if convite.usado:
            return {"error": "Este convite já foi utilizado."}, 400
        return {"error": "Este convite expirou."}, 400

    # Sanitiza strings removendo caracteres Unicode invisíveis (comuns em teclados mobile)
    def sanitizar_texto(texto):
        if not texto:
            return ''
        # Remove caracteres de controle e zero-width characters
        texto = ''.join(c for c in texto if unicodedata.category(c) not in ('Cc', 'Cf') or c in '\n\r\t')
        return texto.strip()

    nome_restaurante = sanitizar_texto(data['nome_restaurante'])
    if not nome_restaurante:
        return {"error": "Nome do restaurante é obrigatório."}, 400

    restaurante_existente = Restaurante.query.filter(
        func.lower(Restaurante.nome) == func.lower(nome_restaurante)
    ).first()
    if restaurante_existente:
        return {"error": f"Já existe um restaurante com o nome '{nome_restaurante}'."}, 409

    email_responsavel = sanitizar_texto(data['email_responsavel']).lower()
    usuario_existente = Usuario.query.filter_by(email=email_responsavel).first()
    if usuario_existente:
        return {"error": f"Já existe um usuário com o email '{email_responsavel}'."}, 409

    try:
        slug_base = slugify(nome_restaurante)

        # Se slugify retornar vazio (nome só com emojis/caracteres especiais), gera slug único
        if not slug_base:
            slug_base = f"restaurante-{secrets.token_hex(4)}"

        slug = slug_base
        contador = 1

        while Restaurante.query.filter_by(slug=slug).first():
            slug = f"{slug_base}-{contador}"
            contador += 1

        restaurante = Restaurante(
            nome=nome_restaurante[:200],
            slug=slug,
            ativo=True
        )
        db.session.add(restaurante)
        db.session.flush()

        caracteres = string.ascii_letters + string.digits
        senha_gerada = ''.join(secrets.choice(caracteres) for _ in range(12))

        nome_responsavel = sanitizar_texto(data['nome_responsavel'])[:100]
        if not nome_responsavel:
            return {"error": "Nome do responsável é obrigatório."}, 400

        usuario_admin = Usuario(
            nome=nome_responsavel,
            email=email_responsavel,
            senha_hash=generate_password_hash(senha_gerada),
            senha_texto_puro=senha_gerada,
            role=UserRoles.ADMIN,
            restaurante_id=restaurante.id,
            aprovado=True,
            ativo=True
        )
        db.session.add(usuario_admin)
        db.session.flush()

        # Incrementa contador de usos
        convite.quantidade_usos += 1
        convite.usado = convite.quantidade_usos >= convite.limite_usos  # Marca como usado se atingiu limite
        convite.usado_em = brasilia_now()  # Último uso
        convite.usado_por_id = usuario_admin.id  # Último usuário que usou
        convite.restaurante_id = restaurante.id  # Último restaurante criado

        db.session.commit()

        return {
            "message": "Cadastro de restaurante realizado com sucesso!",
            "restaurante": restaurante.to_dict(),
            "usuario_admin": {
                "id": usuario_admin.id,
                "nome": usuario_admin.nome,
                "email": usuario_admin.email,
                "senha_gerada": senha_gerada
            },
            "observacao": "IMPORTANTE: Copie a senha gerada e guarde em local seguro."
        }, 201
    except Exception as e:
        db.session.rollback()
        print(f"[ERRO] registrar_restaurante_com_convite: {str(e)}")
        return {"error": f"Erro ao registrar restaurante: {str(e)}"}, 500


def register_user(data):
    """Cria um novo usuário no sistema (sem convite - requer aprovação manual)."""
    if Usuario.query.filter_by(email=data['email']).first():
        return {"error": "E-mail já cadastrado."}, 409

    # Verifica se username já existe (se fornecido)
    if data.get('username') and Usuario.query.filter_by(username=data['username']).first():
        return {"error": "Nome de usuário já cadastrado."}, 409

    hashed_password = generate_password_hash(data['senha'])

    # Suporte legado para criação de ADMIN via token
    TOKEN_ADMIN = "Kaiser@210891"
    token_fornecido = data.get('token_admin')

    if token_fornecido and token_fornecido == TOKEN_ADMIN:
        role = UserRoles.ADMIN
        aprovado = True
        mensagem = "Administrador criado e aprovado automaticamente!"
    else:
        role = UserRoles.COLLABORATOR
        aprovado = False
        mensagem = "Solicitação de cadastro enviada com sucesso. Aguardando aprovação do administrador."

    new_user = Usuario(
        nome=data['nome'],
        username=data.get('username'),
        email=data['email'],
        senha_hash=hashed_password,
        role=role,
        restaurante_id=_get_restaurante_padrao_id(),
        aprovado=aprovado
    )

    db.session.add(new_user)
    db.session.flush()
    log_event(
        acao="create",
        entidade="usuario",
        entidade_id=new_user.id,
        mensagem="Cadastro de usuario criado",
        restaurante_id=new_user.restaurante_id,
        usuario_id=new_user.id
    )
    db.session.commit()

    return {"message": mensagem}, 201


# ===== SOLICITAÇÕES DE RESTAURANTE =====

def criar_solicitacao_restaurante(data):
    """Cria uma solicitação de cadastro de restaurante e notifica SUPER_ADMINs."""
    from .models import SolicitacaoRestaurante, Usuario, UserRoles, TipoNotificacao
    import re

    # Validações básicas
    campos_obrigatorios = [
        'nome_restaurante', 'endereco_restaurante', 'telefone_restaurante',
        'email_restaurante', 'nome_responsavel', 'email_responsavel', 'telefone_responsavel'
    ]

    for campo in campos_obrigatorios:
        if not data.get(campo) or not data.get(campo).strip():
            return {"error": f"Campo '{campo}' é obrigatório."}, 400

    # Validar formato de email
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_regex, data['email_restaurante'].strip()):
        return {"error": "Email do restaurante inválido."}, 400
    if not re.match(email_regex, data['email_responsavel'].strip()):
        return {"error": "Email do responsável inválido."}, 400

    # Validar telefones (mínimo 10 dígitos)
    telefone_restaurante_limpo = re.sub(r'\D', '', data['telefone_restaurante'])
    telefone_responsavel_limpo = re.sub(r'\D', '', data['telefone_responsavel'])

    if len(telefone_restaurante_limpo) < 10:
        return {"error": "Telefone do restaurante deve ter pelo menos 10 dígitos."}, 400
    if len(telefone_responsavel_limpo) < 10:
        return {"error": "Telefone do responsável deve ter pelo menos 10 dígitos."}, 400

    # Validar CNPJ se fornecido
    cnpj = data.get('cnpj', '').strip()
    if cnpj:
        cnpj_limpo = re.sub(r'\D', '', cnpj)
        if len(cnpj_limpo) != 14:
            return {"error": "CNPJ deve conter 14 dígitos."}, 400

    # Verificar duplicata (email responsável pendente)
    from .models import StatusSolicitacaoRestaurante
    solicitacao_existente = SolicitacaoRestaurante.query.filter_by(
        email_responsavel=data['email_responsavel'].strip().lower(),
        status=StatusSolicitacaoRestaurante.PENDENTE
    ).first()

    if solicitacao_existente:
        return {"error": "Já existe uma solicitação pendente com este email de responsável."}, 409

    try:
        # Criar solicitação
        solicitacao = SolicitacaoRestaurante(
            nome_restaurante=data['nome_restaurante'].strip()[:200],
            endereco_restaurante=data['endereco_restaurante'].strip()[:400],
            telefone_restaurante=data['telefone_restaurante'].strip()[:20],
            email_restaurante=data['email_restaurante'].strip().lower()[:120],
            cnpj=cnpj[:18] if cnpj else None,
            razao_social=data.get('razao_social', '').strip()[:200] if data.get('razao_social') else None,
            nome_responsavel=data['nome_responsavel'].strip()[:100],
            email_responsavel=data['email_responsavel'].strip().lower()[:120],
            telefone_responsavel=data['telefone_responsavel'].strip()[:20],
            status=StatusSolicitacaoRestaurante.PENDENTE
        )

        db.session.add(solicitacao)
        db.session.commit()

        # Notificar todos os SUPER_ADMINs (não deve quebrar se falhar)
        try:
            print(f"[DEBUG] Procurando SUPER_ADMINs...")
            super_admins = Usuario.query.filter_by(role=UserRoles.SUPER_ADMIN, ativo=True).all()
            print(f"[DEBUG] Encontrados {len(super_admins)} SUPER_ADMINs ativos")

            for admin in super_admins:
                print(f"[DEBUG] Criando notificação para admin {admin.id} ({admin.email})")
                criar_notificacao(
                    usuario_id=admin.id,
                    titulo=f"Nova Solicitação de Restaurante: {solicitacao.nome_restaurante}",
                    tipo=TipoNotificacao.SOLICITACAO_RESTAURANTE,
                    mensagem=f"Responsável: {solicitacao.nome_responsavel} ({solicitacao.email_responsavel})",
                    restaurante_id=None  # Notificação global para SUPER_ADMIN
                )
                print(f"[DEBUG] Notificação criada com sucesso para {admin.email}")
        except Exception as notif_error:
            import traceback
            print(f"[AVISO] Erro ao criar notificações: {str(notif_error)}")
            traceback.print_exc()
            # Continua mesmo se notificação falhar

        return {
            "message": "Solicitação enviada com sucesso! Você receberá um email quando for processada.",
            "solicitacao_id": solicitacao.id
        }, 201

    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"[ERRO] criar_solicitacao_restaurante: {str(e)}")
        traceback.print_exc()
        return {"error": "Erro ao processar solicitação."}, 500


def listar_solicitacoes_restaurante(filtro_status=None):
    """Lista solicitações de restaurante com filtro opcional de status."""
    from .models import SolicitacaoRestaurante, StatusSolicitacaoRestaurante

    query = SolicitacaoRestaurante.query

    if filtro_status:
        try:
            status_enum = StatusSolicitacaoRestaurante(filtro_status)
            query = query.filter_by(status=status_enum)
        except ValueError:
            return {"error": f"Status inválido: {filtro_status}"}, 400

    solicitacoes = query.order_by(SolicitacaoRestaurante.criado_em.desc()).all()

    return {
        "solicitacoes": [s.to_dict() for s in solicitacoes],
        "total": len(solicitacoes)
    }, 200


def obter_solicitacao_restaurante(solicitacao_id):
    """Obtém detalhes completos de uma solicitação."""
    from .models import SolicitacaoRestaurante

    solicitacao = SolicitacaoRestaurante.query.get(solicitacao_id)

    if not solicitacao:
        return {"error": "Solicitação não encontrada."}, 404

    return {"solicitacao": solicitacao.to_dict()}, 200


def aprovar_solicitacao_restaurante(solicitacao_id, admin_id):
    """Aprova solicitação: cria restaurante + usuário ADMIN + gera credenciais."""
    from .models import SolicitacaoRestaurante, StatusSolicitacaoRestaurante, Restaurante, Usuario, UserRoles
    from slugify import slugify
    from sqlalchemy import func
    import secrets
    import string

    solicitacao = SolicitacaoRestaurante.query.get(solicitacao_id)

    if not solicitacao:
        return {"error": "Solicitação não encontrada."}, 404

    if solicitacao.status != StatusSolicitacaoRestaurante.PENDENTE:
        return {"error": f"Solicitação já foi processada (status: {solicitacao.status.value})."}, 400

    try:
        # 1. CRIAR RESTAURANTE
        nome_restaurante = solicitacao.nome_restaurante.strip()

        # Verificar se existe restaurante com este nome
        restaurante_existente = Restaurante.query.filter(
            func.lower(Restaurante.nome) == func.lower(nome_restaurante)
        ).first()

        if restaurante_existente:
            return {"error": f"Já existe um restaurante com o nome '{nome_restaurante}'."}, 409

        # Gerar slug único
        slug_base = slugify(nome_restaurante)
        slug = slug_base
        contador = 1

        while Restaurante.query.filter_by(slug=slug).first():
            slug = f"{slug_base}-{contador}"
            contador += 1

        restaurante = Restaurante(
            nome=nome_restaurante[:200],
            slug=slug,
            ativo=True
        )
        db.session.add(restaurante)
        db.session.flush()  # Gera o ID do restaurante

        # 2. CRIAR USUÁRIO ADMIN
        email_responsavel = solicitacao.email_responsavel.strip().lower()

        # Verificar duplicata de email
        usuario_existente = Usuario.query.filter_by(email=email_responsavel).first()

        if usuario_existente:
            db.session.rollback()
            return {"error": f"Já existe um usuário com o email '{email_responsavel}'."}, 409

        # Gerar senha aleatória (12 caracteres: letras + números)
        caracteres = string.ascii_letters + string.digits
        senha_gerada = ''.join(secrets.choice(caracteres) for _ in range(12))

        usuario_admin = Usuario(
            nome=solicitacao.nome_responsavel.strip()[:100],
            email=email_responsavel,
            senha_hash=generate_password_hash(senha_gerada),
            senha_texto_puro=senha_gerada,  # Armazena senha em texto puro
            role=UserRoles.ADMIN,
            restaurante_id=restaurante.id,
            aprovado=True,
            ativo=True
        )
        db.session.add(usuario_admin)
        db.session.flush()

        # 3. ATUALIZAR SOLICITAÇÃO
        solicitacao.status = StatusSolicitacaoRestaurante.APROVADO
        solicitacao.processado_em = brasilia_now()
        solicitacao.processado_por_id = admin_id
        solicitacao.restaurante_criado_id = restaurante.id
        solicitacao.usuario_admin_criado_id = usuario_admin.id
        solicitacao.senha_gerada = senha_gerada  # Armazenar temporariamente

        db.session.commit()

        return {
            "message": "Solicitação aprovada com sucesso!",
            "restaurante": restaurante.to_dict(),
            "usuario_admin": {
                "id": usuario_admin.id,
                "nome": usuario_admin.nome,
                "email": usuario_admin.email,
                "senha_gerada": senha_gerada
            },
            "observacao": "IMPORTANTE: Copie a senha gerada e envie ao responsável. Ela não será exibida novamente."
        }, 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERRO] aprovar_solicitacao_restaurante: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": f"Erro ao aprovar solicitação: {str(e)}"}, 500


def rejeitar_solicitacao_restaurante(solicitacao_id, admin_id, motivo):
    """Rejeita solicitação de restaurante com motivo."""
    from .models import SolicitacaoRestaurante, StatusSolicitacaoRestaurante

    if not motivo or not motivo.strip():
        return {"error": "Motivo da rejeição é obrigatório."}, 400

    solicitacao = SolicitacaoRestaurante.query.get(solicitacao_id)

    if not solicitacao:
        return {"error": "Solicitação não encontrada."}, 404

    if solicitacao.status != StatusSolicitacaoRestaurante.PENDENTE:
        return {"error": f"Solicitação já foi processada (status: {solicitacao.status.value})."}, 400

    try:
        solicitacao.status = StatusSolicitacaoRestaurante.REJEITADO
        solicitacao.processado_em = brasilia_now()
        solicitacao.processado_por_id = admin_id
        solicitacao.motivo_rejeicao = motivo.strip()[:500]

        db.session.commit()

        return {
            "message": "Solicitação rejeitada com sucesso.",
            "solicitacao": solicitacao.to_dict()
        }, 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERRO] rejeitar_solicitacao_restaurante: {str(e)}")
        return {"error": "Erro ao rejeitar solicitação."}, 500


# ===== SUPER ADMIN DASHBOARD =====

def get_super_dashboard_data(restaurante_id=None, period=30):
    """
    Retorna dados completos para o Dashboard Global do Super Admin.

    Args:
        restaurante_id: Filtrar por restaurante específico (None = todos)
        period: Número de dias para dados temporais (default 30)

    Returns:
        Dicionário com todos os dados do dashboard
    """
    cache_key = _get_super_dashboard_cache_key(restaurante_id, period)
    cached = _SUPER_DASHBOARD_CACHE.get(cache_key)
    if cached and cached.get("expires_at", 0) > time.time():
        return cached["data"], 200

    today = brasilia_now().date()
    max_days = max(period, 35)
    max_period_start = today - timedelta(days=max_days)
    max_period_start_dt = datetime.combine(max_period_start, datetime.min.time())
    period_end_dt = datetime.combine(today + timedelta(days=1), datetime.min.time())

    # ===== SUMMARY =====
    # Total de restaurantes
    if restaurante_id:
        total_restaurantes = Restaurante.query.filter_by(
            id=restaurante_id,
            deletado=False,
            ativo=True
        ).count()
    else:
        total_restaurantes = Restaurante.query.filter_by(deletado=False, ativo=True).count()

    # Usuários
    users_query = Usuario.query.filter(Usuario.ativo == True)
    if restaurante_id:
        users_query = users_query.filter_by(restaurante_id=restaurante_id)

    total_users = users_query.count()

    # Usuários por role
    users_by_role = {
        'super_admin': users_query.filter_by(role=UserRoles.SUPER_ADMIN).count(),
        'admin': users_query.filter_by(role=UserRoles.ADMIN).count(),
        'collaborator': users_query.filter_by(role=UserRoles.COLLABORATOR).count()
    }

    # Usuários pendentes de aprovação
    pending_approvals = users_query.filter_by(aprovado=False).count()

    # Listas
    listas_query = Lista.query.filter_by(deletado=False)
    if restaurante_id:
        listas_query = listas_query.filter_by(restaurante_id=restaurante_id)
    total_listas = listas_query.count()

    # Itens no catálogo
    itens_query = ListaMaeItem.query
    if restaurante_id:
        itens_query = itens_query.filter_by(restaurante_id=restaurante_id)
    total_itens = itens_query.count()

    # Submissões (pedidos agregados)
    submissoes_query = Submissao.query.filter(Submissao.arquivada.is_(False))
    if restaurante_id:
        submissoes_query = submissoes_query.join(Usuario, Submissao.usuario_id == Usuario.id).filter(
            Usuario.restaurante_id == restaurante_id
        )
    total_submissoes = submissoes_query.count()

    # Cotações
    cotacoes_query = Cotacao.query
    if restaurante_id:
        cotacoes_query = cotacoes_query.join(Fornecedor).filter(
            Fornecedor.restaurante_id == restaurante_id
        )
    pending_cotacoes = cotacoes_query.filter(Cotacao.status == CotacaoStatus.PENDENTE).count()
    completed_cotacoes = cotacoes_query.filter(Cotacao.status == CotacaoStatus.CONCLUIDA).count()

    # ===== TEMPORAL DATA =====
    # Submissões por dia
    dates = [today - timedelta(days=i) for i in range(period - 1, -1, -1)]
    extended_dates = [today - timedelta(days=i) for i in range(max_days - 1, -1, -1)]
    submissions_day_query = db.session.query(
        func.date(Submissao.data_submissao).label('day'),
        func.count(Submissao.id)
    ).filter(
        Submissao.arquivada.is_(False),
        Submissao.data_submissao >= max_period_start_dt,
        Submissao.data_submissao < period_end_dt
    )
    if restaurante_id:
        submissions_day_query = submissions_day_query.join(Usuario, Submissao.usuario_id == Usuario.id).filter(
            Usuario.restaurante_id == restaurante_id
        )
    submissions_day_rows = submissions_day_query.group_by(func.date(Submissao.data_submissao)).all()
    submissions_by_day = {str(day): count for day, count in submissions_day_rows}
    submissions_per_day_data = [submissions_by_day.get(d.isoformat(), 0) for d in dates]
    submissoes_hoje = submissions_by_day.get(today.isoformat(), 0)

    summary = {
        'total_restaurantes': total_restaurantes,
        'total_users': total_users,
        'users_by_role': users_by_role,
        'pending_approvals': pending_approvals,
        'total_listas': total_listas,
        'total_itens': total_itens,
        'total_submissoes': total_submissoes,
        'submissoes_hoje': submissoes_hoje,
        'pending_cotacoes': pending_cotacoes,
        'completed_cotacoes': completed_cotacoes
    }
    submissions_per_day = {
        'labels': [d.strftime('%d/%m') for d in dates],
        'data': submissions_per_day_data
    }

    # Submissões por semana (últimas 5 semanas)
    weeks_data = []
    weeks_labels = []
    for i in range(4, -1, -1):
        week_end = today - timedelta(days=i * 7)
        week_start = week_end - timedelta(days=6)
        count = 0
        for d in extended_dates:
            if week_start <= d <= week_end:
                count += submissions_by_day.get(d.isoformat(), 0)
        weeks_data.append(count)
        weeks_labels.append(f'Sem {5-i}')
    submissions_per_week = {
        'labels': weeks_labels,
        'data': weeks_data
    }

    # Usuários criados/ativos por mês (últimos 6 meses)
    months_data = []
    active_users_data = []
    months_labels = []
    month_starts = []
    for i in range(5, -1, -1):
        month_date = today.replace(day=1) - timedelta(days=i * 30)
        month_start = month_date.replace(day=1)
        month_starts.append(month_start)
        months_labels.append(month_start.strftime('%b'))

    months_range_start = month_starts[0]
    months_range_end = (month_starts[-1].replace(day=28) + timedelta(days=4)).replace(day=1)

    engine_name = db.engine.name
    if engine_name in ("postgresql", "postgres"):
        month_expr_users = func.to_char(Usuario.criado_em, 'YYYY-MM')
        month_expr_logs = func.to_char(AppLog.criado_em, 'YYYY-MM')
    else:
        month_expr_users = func.strftime('%Y-%m', Usuario.criado_em)
        month_expr_logs = func.strftime('%Y-%m', AppLog.criado_em)

    created_query = db.session.query(
        month_expr_users.label('month'),
        func.count(Usuario.id)
    ).filter(
        Usuario.ativo == True,
        Usuario.criado_em >= months_range_start,
        Usuario.criado_em < months_range_end
    )
    if restaurante_id:
        created_query = created_query.filter(Usuario.restaurante_id == restaurante_id)
    created_rows = created_query.group_by(month_expr_users).all()
    created_by_month = {month: count for month, count in created_rows}

    active_query = db.session.query(
        month_expr_logs.label('month'),
        func.count(func.distinct(AppLog.usuario_id))
    ).filter(
        AppLog.acao == "login",
        AppLog.usuario_id.isnot(None),
        AppLog.criado_em >= months_range_start,
        AppLog.criado_em < months_range_end
    )
    if restaurante_id:
        active_query = active_query.filter(AppLog.restaurante_id == restaurante_id)
    active_rows = active_query.group_by(month_expr_logs).all()
    active_by_month = {month: count for month, count in active_rows}

    for month_start in month_starts:
        month_key = month_start.strftime('%Y-%m')
        months_data.append(created_by_month.get(month_key, 0))
        active_users_data.append(active_by_month.get(month_key, 0))

    users_created_per_month = {
        'labels': months_labels,
        'data': months_data
    }
    users_active_per_month = {
        'labels': months_labels,
        'data': active_users_data
    }

    # Distribuição de status de submissões
    status_query = db.session.query(
        Submissao.status, func.count(Submissao.id)
    ).filter(Submissao.arquivada.is_(False))
    if restaurante_id:
        status_query = status_query.join(Usuario, Submissao.usuario_id == Usuario.id).filter(
            Usuario.restaurante_id == restaurante_id
        )
    status_counts = status_query.group_by(Submissao.status).all()

    submission_status_distribution = {
        'labels': ['Pendente', 'Parcial', 'Aprovado', 'Rejeitado'],
        'data': [0, 0, 0, 0]
    }
    for status, count in status_counts:
        if status == SubmissaoStatus.PENDENTE:
            submission_status_distribution['data'][0] = count
        elif status == SubmissaoStatus.PARCIALMENTE_APROVADO:
            submission_status_distribution['data'][1] = count
        elif status == SubmissaoStatus.APROVADO:
            submission_status_distribution['data'][2] = count
        elif status == SubmissaoStatus.REJEITADO:
            submission_status_distribution['data'][3] = count

    temporal = {
        'submissions_per_day': submissions_per_day,
        'submissions_per_week': submissions_per_week,
        'users_created_per_month': users_created_per_month,
        'users_active_per_month': users_active_per_month,
        'submission_status_distribution': submission_status_distribution
    }

    # ===== LISTS DATA =====
    # Status das listas semanais
    weekly_status = []
    listas_status_query = Lista.query.filter_by(deletado=False)
    if restaurante_id:
        listas_status_query = listas_status_query.filter_by(restaurante_id=restaurante_id)
    listas = listas_status_query.limit(10).all()
    for lista in listas:
        submissoes_semana_query = Submissao.query.filter(
            Submissao.arquivada.is_(False),
            Submissao.lista_id == lista.id,
            func.date(Submissao.data_submissao) >= today - timedelta(days=7)
        )
        submissoes_semana = submissoes_semana_query.count()
        ultima_submissao = Submissao.query.filter_by(lista_id=lista.id).order_by(
            Submissao.data_submissao.desc()
        ).first()
        weekly_status.append({
            'id': lista.id,
            'nome': lista.nome,
            'submissoes_semana': submissoes_semana,
            'ultima_submissao': ultima_submissao.data_submissao.isoformat() if ultima_submissao and ultima_submissao.data_submissao else None,
            'status': 'ativo'
        })

    # Listas rápidas por prioridade
    quick_lists_query = ListaRapida.query.filter(
        ListaRapida.status == StatusListaRapida.PENDENTE,
        ListaRapida.deletado == False
    )
    if restaurante_id:
        quick_lists_query = quick_lists_query.join(
            Usuario, ListaRapida.usuario_id == Usuario.id
        ).filter(Usuario.restaurante_id == restaurante_id)
    quick_lists_pending = quick_lists_query.count()

    # Contar itens por prioridade em listas rápidas pendentes
    priority_query = db.session.query(
        ListaRapidaItem.prioridade, func.count(ListaRapidaItem.id)
    ).join(ListaRapida).filter(
        ListaRapida.status == StatusListaRapida.PENDENTE,
        ListaRapida.deletado == False,
        ListaRapidaItem.descartado == False
    )
    if restaurante_id:
        priority_query = priority_query.join(
            Usuario, ListaRapida.usuario_id == Usuario.id
        ).filter(Usuario.restaurante_id == restaurante_id)
    priority_counts = priority_query.group_by(ListaRapidaItem.prioridade).all()

    quick_lists_by_priority = {
        'prevencao': 0,
        'precisa_comprar': 0,
        'urgente': 0
    }
    for priority, count in priority_counts:
        if priority == PrioridadeItem.PREVENCAO:
            quick_lists_by_priority['prevencao'] = count
        elif priority == PrioridadeItem.PRECISA_COMPRAR:
            quick_lists_by_priority['precisa_comprar'] = count
        elif priority == PrioridadeItem.URGENTE:
            quick_lists_by_priority['urgente'] = count

    # Taxa de aprovação de listas rápidas
    total_rapidas = ListaRapida.query.filter(
        ListaRapida.status.in_([StatusListaRapida.APROVADA, StatusListaRapida.REJEITADA])
    ).count()
    aprovadas = ListaRapida.query.filter_by(status=StatusListaRapida.APROVADA).count()
    approval_rate = (aprovadas / total_rapidas * 100) if total_rapidas > 0 else 0

    lists_data = {
        'weekly_status': weekly_status,
        'quick_lists_by_priority': quick_lists_by_priority,
        'quick_lists_pending': quick_lists_pending,
        'approval_rate': round(approval_rate, 1)
    }

    # ===== CHECKLISTS DATA =====
    active_checklists = Checklist.query.filter_by(status=ChecklistStatus.ATIVO).count()
    completed_checklists = Checklist.query.filter_by(status=ChecklistStatus.FINALIZADO).count()

    # Taxa de conclusão
    total_checklist_items = ChecklistItem.query.count()
    marcados = ChecklistItem.query.filter_by(marcado=True).count()
    completion_rate = (marcados / total_checklist_items * 100) if total_checklist_items > 0 else 0

    # Tempo médio para finalizar (em horas)
    completed_with_time = Checklist.query.filter(
        Checklist.status == ChecklistStatus.FINALIZADO,
        Checklist.finalizado_em.isnot(None)
    ).all()

    avg_completion_time = None
    if completed_with_time:
        total_hours = 0
        count = 0
        for cl in completed_with_time:
            if cl.criado_em and cl.finalizado_em:
                diff = (cl.finalizado_em - cl.criado_em).total_seconds() / 3600
                total_hours += diff
                count += 1
        if count > 0:
            avg_completion_time = round(total_hours / count, 1)

    # Últimos checklists
    recent_checklists = []
    recent_cls = Checklist.query.order_by(Checklist.criado_em.desc()).limit(5).all()
    for cl in recent_cls:
        total_itens = cl.itens.count()
        itens_marcados = cl.itens.filter_by(marcado=True).count()
        recent_checklists.append({
            'id': cl.id,
            'nome': cl.nome,
            'status': cl.status.value,
            'criado_em': cl.criado_em.isoformat() if cl.criado_em else None,
            'total_itens': total_itens,
            'itens_marcados': itens_marcados
        })

    checklists_data = {
        'active_count': active_checklists,
        'completed_count': completed_checklists,
        'completion_rate': round(completion_rate, 1),
        'avg_completion_time_hours': avg_completion_time,
        'recent_checklists': recent_checklists
    }

    # ===== USERS DATA =====
    # Timeline de novos usuários (últimos 30 dias)
    new_users_data = []
    for d in dates[-30:]:
        new_users_query = Usuario.query.filter(
            func.date(Usuario.criado_em) == d,
            Usuario.ativo == True
        )
        if restaurante_id:
            new_users_query = new_users_query.filter(Usuario.restaurante_id == restaurante_id)
        new_users_data.append(new_users_query.count())
    new_users_timeline = {
        'labels': [d.strftime('%d/%m') for d in dates[-30:]],
        'data': new_users_data
    }

    # Usuários mais ativos (por submissões)
    most_active = db.session.query(
        Usuario.id, Usuario.nome, func.count(Submissao.id).label('submissoes'),
        Restaurante.nome.label('restaurante_nome')
    ).join(Submissao, Usuario.id == Submissao.usuario_id).outerjoin(
        Restaurante, Usuario.restaurante_id == Restaurante.id
    ).group_by(Usuario.id, Usuario.nome, Restaurante.nome).order_by(
        func.count(Submissao.id).desc()
    ).limit(5).all()

    most_active_users = [{
        'id': u.id,
        'nome': u.nome,
        'submissoes': u.submissoes,
        'restaurante': u.restaurante_nome or 'N/A'
    } for u in most_active]

    # Tempo médio de resposta dos admins (para sugestões)
    responded_suggestions = SugestaoItem.query.filter(
        SugestaoItem.status != SugestaoStatus.PENDENTE,
        SugestaoItem.respondido_em.isnot(None)
    ).all()

    admin_response_time = None
    if responded_suggestions:
        total_hours = 0
        count = 0
        for sug in responded_suggestions:
            if sug.criado_em and sug.respondido_em:
                diff = (sug.respondido_em - sug.criado_em).total_seconds() / 3600
                total_hours += diff
                count += 1
        if count > 0:
            admin_response_time = round(total_hours / count, 1)

    users_data = {
        'new_users_timeline': new_users_timeline,
        'most_active_users': most_active_users,
        'admin_avg_response_time_hours': admin_response_time
    }

    # ===== SUPPLIERS DATA =====
    total_fornecedores = Fornecedor.query.count()

    # Fornecedores por restaurante
    suppliers_by_rest = db.session.query(
        Restaurante.nome, Restaurante.id, func.count(Fornecedor.id).label('total')
    ).outerjoin(Fornecedor, Restaurante.id == Fornecedor.restaurante_id).filter(
        Restaurante.deletado == False
    ).group_by(Restaurante.nome, Restaurante.id).all()

    by_restaurant = [{
        'restaurante': r.nome,
        'restaurante_id': r.id,
        'total_fornecedores': r.total
    } for r in suppliers_by_rest]

    # Status de cotações
    quotation_status = {
        'pendente': pending_cotacoes,
        'concluida': completed_cotacoes
    }

    # Valor médio de cotações
    avg_value_result = db.session.query(
        func.avg(CotacaoItem.preco_unitario * CotacaoItem.quantidade)
    ).scalar()
    avg_quotation_value = round(float(avg_value_result), 2) if avg_value_result else None

    suppliers_data = {
        'total_fornecedores': total_fornecedores,
        'by_restaurant': by_restaurant,
        'quotation_status': quotation_status,
        'avg_quotation_value': avg_quotation_value
    }

    # ===== SUGGESTIONS DATA =====
    pending_suggestions = SugestaoItem.query.filter_by(status=SugestaoStatus.PENDENTE).count()
    approved_suggestions = SugestaoItem.query.filter_by(status=SugestaoStatus.APROVADA).count()
    rejected_suggestions = SugestaoItem.query.filter_by(status=SugestaoStatus.REJEITADA).count()

    total_processed = approved_suggestions + rejected_suggestions
    suggestion_approval_rate = (approved_suggestions / total_processed * 100) if total_processed > 0 else 0

    # Itens mais sugeridos
    most_suggested = db.session.query(
        SugestaoItem.nome_item, func.count(SugestaoItem.id).label('count')
    ).group_by(SugestaoItem.nome_item).order_by(
        func.count(SugestaoItem.id).desc()
    ).limit(5).all()

    most_suggested_items = [{
        'nome': s.nome_item,
        'count': s.count
    } for s in most_suggested]

    suggestions_data = {
        'pending_count': pending_suggestions,
        'approved_count': approved_suggestions,
        'rejected_count': rejected_suggestions,
        'approval_rate': round(suggestion_approval_rate, 1),
        'most_suggested_items': most_suggested_items
    }

    # ===== NOTIFICATIONS DATA =====
    total_notifications = Notificacao.query.count()
    unread_notifications = Notificacao.query.filter_by(lida=False).count()
    read_notifications = total_notifications - unread_notifications
    read_rate = (read_notifications / total_notifications * 100) if total_notifications > 0 else 0

    # Distribuição por tipo
    types_dist = db.session.query(
        Notificacao.tipo, func.count(Notificacao.id)
    ).group_by(Notificacao.tipo).all()

    types_distribution = {}
    for tipo, count in types_dist:
        types_distribution[tipo.value] = count

    notifications_data = {
        'total_count': total_notifications,
        'unread_count': unread_notifications,
        'read_rate': round(read_rate, 1),
        'types_distribution': types_distribution
    }

    # ===== RECENT ACTIVITIES =====
    recent_activities = []

    # Últimas submissões
    recent_subs_query = Submissao.query.filter(Submissao.arquivada.is_(False))
    if restaurante_id:
        recent_subs_query = recent_subs_query.join(Usuario, Submissao.usuario_id == Usuario.id).filter(
            Usuario.restaurante_id == restaurante_id
        )
    recent_subs = recent_subs_query.order_by(Submissao.data_submissao.desc()).limit(5).all()
    for sub in recent_subs:
        lista_nome = sub.lista.nome if sub.lista else 'Lista'
        usuario_nome = sub.usuario.nome if sub.usuario else 'Usuário'
        rest_nome = sub.usuario.restaurante.nome if sub.usuario and sub.usuario.restaurante else None
        recent_activities.append({
            'type': 'submissao',
            'description': f'Lista "{lista_nome}" submetida por {usuario_nome}',
            'timestamp': sub.data_submissao.isoformat() if sub.data_submissao else None,
            'restaurante': rest_nome,
            'id': sub.id
        })

    # Últimas cotações
    recent_cots_query = Cotacao.query
    if restaurante_id:
        recent_cots_query = recent_cots_query.join(Fornecedor).filter(
            Fornecedor.restaurante_id == restaurante_id
        )
    recent_cots = recent_cots_query.order_by(Cotacao.data_cotacao.desc()).limit(3).all()
    for cot in recent_cots:
        forn_nome = cot.fornecedor.nome if cot.fornecedor else 'Fornecedor'
        rest_nome = cot.fornecedor.restaurante.nome if cot.fornecedor and cot.fornecedor.restaurante else None
        recent_activities.append({
            'type': 'cotacao',
            'description': f'Cotação #{cot.id} criada para {forn_nome}',
            'timestamp': cot.data_cotacao.isoformat() if cot.data_cotacao else None,
            'restaurante': rest_nome,
            'id': cot.id
        })

    # Novos usuários
    recent_users_query = Usuario.query
    if restaurante_id:
        recent_users_query = recent_users_query.filter(Usuario.restaurante_id == restaurante_id)
    recent_users = recent_users_query.order_by(Usuario.criado_em.desc()).limit(3).all()
    for usr in recent_users:
        rest_nome = usr.restaurante.nome if usr.restaurante else None
        recent_activities.append({
            'type': 'usuario',
            'description': f'Novo usuário: {usr.nome} ({usr.role.value})',
            'timestamp': usr.criado_em.isoformat() if usr.criado_em else None,
            'restaurante': rest_nome,
            'id': usr.id
        })

    # Ordenar por timestamp
    recent_activities.sort(key=lambda x: x['timestamp'] or '', reverse=True)
    recent_activities = recent_activities[:10]

    # ===== META =====
    meta = {
        'generated_at': brasilia_now().isoformat(),
        'period_days': period,
        'filtered_restaurante_id': restaurante_id
    }

    response = {
        'summary': summary,
        'temporal': temporal,
        'lists': lists_data,
        'checklists': checklists_data,
        'users': users_data,
        'suppliers': suppliers_data,
        'suggestions': suggestions_data,
        'notifications': notifications_data,
        'recent_activities': recent_activities,
        'meta': meta
    }

    _SUPER_DASHBOARD_CACHE[cache_key] = {
        "expires_at": time.time() + _SUPER_DASHBOARD_CACHE_TTL_SECONDS,
        "data": response
    }

    return response, 200


def generate_super_dashboard_pdf(restaurante_id=None, period=30):
    """
    Gera PDF do dashboard do super admin.
    """
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    data, _ = get_super_dashboard_data(restaurante_id, period)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                           leftMargin=1*cm, rightMargin=1*cm,
                           topMargin=1*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        textColor=colors.HexColor('#2c3e50')
    )

    elements = []

    # Título
    elements.append(Paragraph("Dashboard Global - Super Admin", title_style))
    elements.append(Paragraph(f"Gerado em: {data['meta']['generated_at']}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Resumo
    elements.append(Paragraph("Resumo Geral", styles['Heading2']))
    summary = data['summary']
    summary_data = [
        ['Métrica', 'Valor'],
        ['Total de Restaurantes', str(summary['total_restaurantes'])],
        ['Total de Usuários', str(summary['total_users'])],
        ['Usuários Pendentes', str(summary['pending_approvals'])],
        ['Total de Listas', str(summary['total_listas'])],
        ['Total de Itens', str(summary['total_itens'])],
        ['Total de Submissões', str(summary['total_submissoes'])],
        ['Submissões Hoje', str(summary['submissoes_hoje'])],
        ['Cotações Pendentes', str(summary['pending_cotacoes'])],
        ['Cotações Concluídas', str(summary['completed_cotacoes'])],
    ]

    t = Table(summary_data, colWidths=[8*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6'))
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # Checklists
    elements.append(Paragraph("Checklists", styles['Heading2']))
    checklists = data['checklists']
    checklist_data = [
        ['Métrica', 'Valor'],
        ['Ativos', str(checklists['active_count'])],
        ['Finalizados', str(checklists['completed_count'])],
        ['Taxa de Conclusão', f"{checklists['completion_rate']}%"],
    ]

    t2 = Table(checklist_data, colWidths=[8*cm, 4*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2eb85c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6'))
    ]))
    elements.append(t2)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_super_dashboard_excel(restaurante_id=None, period=30):
    """
    Gera Excel com dados do dashboard do super admin.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data, _ = get_super_dashboard_data(restaurante_id, period)

    wb = Workbook()

    # Estilos
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== Sheet 1: Resumo =====
    ws1 = wb.active
    ws1.title = "Resumo"

    summary = data['summary']
    ws1['A1'] = 'Dashboard Global - Super Admin'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f"Gerado em: {data['meta']['generated_at']}"

    headers = ['Métrica', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    summary_rows = [
        ('Total de Restaurantes', summary['total_restaurantes']),
        ('Total de Usuários', summary['total_users']),
        ('Super Admins', summary['users_by_role']['super_admin']),
        ('Admins', summary['users_by_role']['admin']),
        ('Colaboradores', summary['users_by_role']['collaborator']),
        ('Pendentes Aprovação', summary['pending_approvals']),
        ('Total de Listas', summary['total_listas']),
        ('Total de Itens', summary['total_itens']),
        ('Total de Submissões', summary['total_submissoes']),
        ('Submissões Hoje', summary['submissoes_hoje']),
        ('Cotações Pendentes', summary['pending_cotacoes']),
        ('Cotações Concluídas', summary['completed_cotacoes']),
    ]

    for row, (metric, value) in enumerate(summary_rows, 5):
        ws1.cell(row=row, column=1, value=metric).border = thin_border
        ws1.cell(row=row, column=2, value=value).border = thin_border

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 15

    # ===== Sheet 2: Submissões por Dia =====
    ws2 = wb.create_sheet("Submissões por Dia")
    temporal = data['temporal']['submissions_per_day']

    ws2.cell(row=1, column=1, value='Data').fill = header_fill
    ws2.cell(row=1, column=1).font = header_font
    ws2.cell(row=1, column=2, value='Quantidade').fill = header_fill
    ws2.cell(row=1, column=2).font = header_font

    for i, (label, value) in enumerate(zip(temporal['labels'], temporal['data']), 2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=value)

    # ===== Sheet 3: Usuários Ativos =====
    ws3 = wb.create_sheet("Usuários Ativos")
    users = data['users']['most_active_users']

    headers = ['Nome', 'Submissões', 'Restaurante']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row, user in enumerate(users, 2):
        ws3.cell(row=row, column=1, value=user['nome'])
        ws3.cell(row=row, column=2, value=user['submissoes'])
        ws3.cell(row=row, column=3, value=user['restaurante'])

    # ===== Sheet 4: Fornecedores =====
    ws4 = wb.create_sheet("Fornecedores")
    suppliers = data['suppliers']['by_restaurant']

    headers = ['Restaurante', 'Total Fornecedores']
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row, sup in enumerate(suppliers, 2):
        ws4.cell(row=row, column=1, value=sup['restaurante'])
        ws4.cell(row=row, column=2, value=sup['total_fornecedores'])

    # ===== Sheet 5: Sugestões =====
    ws5 = wb.create_sheet("Sugestões")
    suggestions = data['suggestions']

    ws5['A1'] = 'Sugestões de Itens'
    ws5['A1'].font = Font(bold=True, size=12)

    ws5['A3'] = 'Pendentes'
    ws5['B3'] = suggestions['pending_count']
    ws5['A4'] = 'Aprovadas'
    ws5['B4'] = suggestions['approved_count']
    ws5['A5'] = 'Rejeitadas'
    ws5['B5'] = suggestions['rejected_count']
    ws5['A6'] = 'Taxa de Aprovação'
    ws5['B6'] = f"{suggestions['approval_rate']}%"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =====================================================
# POPs DIARIOS - CONFIGURACOES, LISTAS, EXECUCOES
# =====================================================

def _resolve_restaurante_id(data, restaurante_id):
    if restaurante_id is not None:
        return restaurante_id, None
    payload_id = data.get('restaurante_id') if isinstance(data, dict) else None
    if not payload_id:
        return None, {"error": "restaurante_id é obrigatório para SUPER_ADMIN."}
    return payload_id, None


def _get_pop_config(restaurante_id):
    config = POPConfiguracao.query.filter_by(restaurante_id=restaurante_id).first()
    if not config:
        config = POPConfiguracao(restaurante_id=restaurante_id)
        db.session.add(config)
        db.session.commit()
    return config


def _user_can_access_pop_lista(usuario, lista):
    if usuario.role == UserRoles.SUPER_ADMIN:
        return True
    if lista.restaurante_id != usuario.restaurante_id:
        return False
    if _is_admin_or_super_admin(usuario):
        return True
    if lista.publico:
        return True
    return lista.colaboradores.filter_by(id=usuario.id).first() is not None


def _apply_pop_auto_archive(restaurante_id, executed_by_id=None):
    config = _get_pop_config(restaurante_id)
    if not config.auto_arquivar:
        return 0

    now = brasilia_now()
    if config.ultimo_auto_arquivamento_em:
        delta = now - config.ultimo_auto_arquivamento_em
        if delta.days < 7:
            return 0

    limite = now.date() - timedelta(days=config.periodo_arquivamento_dias)
    execucoes = POPExecucao.query.filter(
        POPExecucao.restaurante_id == restaurante_id,
        POPExecucao.arquivado.is_(False),
        POPExecucao.data_referencia <= limite
    ).all()

    for execucao in execucoes:
        execucao.arquivado = True
        execucao.arquivado_em = now
        execucao.arquivado_por_id = executed_by_id

    config.ultimo_auto_arquivamento_em = now
    db.session.commit()
    return len(execucoes)


# ----- Configuracoes POP -----

def get_pop_config(restaurante_id):
    if not restaurante_id:
        return {"error": "Restaurante é obrigatório."}, 400
    config = _get_pop_config(restaurante_id)
    return config.to_dict(), 200


def update_pop_config(data, restaurante_id):
    rest_id, error = _resolve_restaurante_id(data, restaurante_id)
    if error:
        return error, 400
    config = _get_pop_config(rest_id)

    if isinstance(data, dict):
        if 'auto_arquivar' in data:
            config.auto_arquivar = bool(data.get('auto_arquivar'))
        if 'periodo_arquivamento_dias' in data:
            try:
                periodo = int(data.get('periodo_arquivamento_dias') or 7)
            except (TypeError, ValueError):
                return {"error": "periodo_arquivamento_dias inválido."}, 400
            if periodo < 1:
                return {"error": "periodo_arquivamento_dias deve ser maior que 0."}, 400
            config.periodo_arquivamento_dias = periodo
        if 'hora_execucao_arquivamento' in data:
            hora = data.get('hora_execucao_arquivamento')
            if hora:
                try:
                    config.hora_execucao_arquivamento = datetime.strptime(hora, "%H:%M").time()
                except ValueError:
                    return {"error": "hora_execucao_arquivamento inválida. Use HH:MM."}, 400
            else:
                config.hora_execucao_arquivamento = None

    db.session.commit()
    return config.to_dict(), 200


# ----- Categorias -----

def list_pop_categorias(restaurante_id):
    query = POPCategoria.query
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    categorias = query.order_by(POPCategoria.ordem.asc(), POPCategoria.nome.asc()).all()
    return [cat.to_dict() for cat in categorias], 200


def create_pop_categoria(data, restaurante_id, usuario_id=None):
    rest_id, error = _resolve_restaurante_id(data, restaurante_id)
    if error:
        return error, 400
    if not data or not data.get('nome'):
        return {"error": "Nome é obrigatório."}, 400

    categoria = POPCategoria(
        restaurante_id=rest_id,
        nome=data['nome'].strip(),
        descricao=data.get('descricao'),
        icone=data.get('icone'),
        cor=data.get('cor'),
        ordem=int(data.get('ordem') or 0),
        ativo=bool(data.get('ativo', True))
    )
    db.session.add(categoria)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Categoria já existe."}, 409

    return categoria.to_dict(), 201


def update_pop_categoria(categoria_id, data, restaurante_id):
    categoria = POPCategoria.query.get(categoria_id)
    if not categoria:
        return {"error": "Categoria não encontrada."}, 404
    if restaurante_id is not None and categoria.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403

    if 'nome' in data and data['nome']:
        categoria.nome = data['nome'].strip()
    if 'descricao' in data:
        categoria.descricao = data.get('descricao')
    if 'icone' in data:
        categoria.icone = data.get('icone')
    if 'cor' in data:
        categoria.cor = data.get('cor')
    if 'ordem' in data:
        categoria.ordem = int(data.get('ordem') or 0)
    if 'ativo' in data:
        categoria.ativo = bool(data.get('ativo'))

    db.session.commit()
    return categoria.to_dict(), 200


def delete_pop_categoria(categoria_id, restaurante_id):
    categoria = POPCategoria.query.get(categoria_id)
    if not categoria:
        return {"error": "Categoria não encontrada."}, 404
    if restaurante_id is not None and categoria.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    db.session.delete(categoria)
    db.session.commit()
    return {"message": "Categoria removida."}, 200


# ----- Templates -----

def list_pop_templates(restaurante_id, filtros=None):
    query = POPTemplate.query
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    filtros = filtros or {}
    if filtros.get('categoria_id'):
        try:
            query = query.filter_by(categoria_id=int(filtros.get('categoria_id')))
        except (TypeError, ValueError):
            return {"error": "categoria_id inválido."}, 400
    if filtros.get('area_id'):
        try:
            query = query.filter_by(area_id=int(filtros.get('area_id')))
        except (TypeError, ValueError):
            return {"error": "area_id inválido."}, 400
    if filtros.get('tipo_verificacao'):
        try:
            tipo = TipoVerificacao(filtros.get('tipo_verificacao'))
            query = query.filter_by(tipo_verificacao=tipo)
        except ValueError:
            return {"error": "tipo_verificacao inválido."}, 400
    if 'rapida' in filtros:
        query = query.filter_by(rapida=bool(filtros.get('rapida')))
    if 'ativo' in filtros:
        query = query.filter_by(ativo=bool(filtros.get('ativo')))

    templates = query.order_by(POPTemplate.ordem.asc(), POPTemplate.titulo.asc()).all()
    return [tpl.to_dict() for tpl in templates], 200


def create_pop_template(data, restaurante_id, usuario_id=None):
    rest_id, error = _resolve_restaurante_id(data, restaurante_id)
    if error:
        return error, 400
    if not data or not data.get('titulo'):
        return {"error": "Título é obrigatório."}, 400

    tipo = data.get('tipo_verificacao', TipoVerificacao.CHECKBOX.value)
    try:
        tipo_enum = TipoVerificacao(tipo)
    except ValueError:
        return {"error": "Tipo de verificação inválido."}, 400

    criticidade = data.get('criticidade', CriticidadeTarefa.NORMAL.value)
    try:
        criticidade_enum = CriticidadeTarefa(criticidade)
    except ValueError:
        return {"error": "Criticidade inválida."}, 400

    if tipo_enum in (TipoVerificacao.MEDICAO, TipoVerificacao.TEMPERATURA) and not data.get('unidade_medicao'):
        return {"error": "unidade_medicao é obrigatória para medições."}, 400

    if data.get('valor_minimo') is not None and data.get('valor_maximo') is not None:
        try:
            if float(data.get('valor_maximo')) <= float(data.get('valor_minimo')):
                return {"error": "valor_maximo deve ser maior que valor_minimo."}, 400
        except (TypeError, ValueError):
            return {"error": "valor_minimo/valor_maximo inválidos."}, 400

    template = POPTemplate(
        restaurante_id=rest_id,
        categoria_id=data.get('categoria_id'),
        area_id=data.get('area_id'),
        titulo=data['titulo'].strip(),
        descricao=data.get('descricao'),
        instrucoes=data.get('instrucoes'),
        tipo_verificacao=tipo_enum,
        requer_foto=bool(data.get('requer_foto', False)),
        requer_medicao=bool(data.get('requer_medicao', False)),
        rapida=bool(data.get('rapida', False)),
        unidade_medicao=data.get('unidade_medicao'),
        valor_minimo=data.get('valor_minimo'),
        valor_maximo=data.get('valor_maximo'),
        criticidade=criticidade_enum,
        tempo_estimado=data.get('tempo_estimado'),
        ordem=int(data.get('ordem') or 0),
        ativo=bool(data.get('ativo', True)),
        criado_por_id=usuario_id
    )
    db.session.add(template)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Template já existe."}, 409

    return template.to_dict(), 201


def update_pop_template(template_id, data, restaurante_id):
    template = POPTemplate.query.get(template_id)
    if not template:
        return {"error": "Template não encontrado."}, 404
    if restaurante_id is not None and template.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403

    if 'titulo' in data and data['titulo']:
        template.titulo = data['titulo'].strip()
    if 'descricao' in data:
        template.descricao = data.get('descricao')
    if 'instrucoes' in data:
        template.instrucoes = data.get('instrucoes')
    if 'categoria_id' in data:
        template.categoria_id = data.get('categoria_id')
    if 'area_id' in data:
        template.area_id = data.get('area_id')
    if 'tipo_verificacao' in data:
        try:
            template.tipo_verificacao = TipoVerificacao(data.get('tipo_verificacao'))
        except ValueError:
            return {"error": "Tipo de verificação inválido."}, 400
    if 'requer_foto' in data:
        template.requer_foto = bool(data.get('requer_foto'))
    if 'requer_medicao' in data:
        template.requer_medicao = bool(data.get('requer_medicao'))
    if 'rapida' in data:
        template.rapida = bool(data.get('rapida'))
    if 'unidade_medicao' in data:
        template.unidade_medicao = data.get('unidade_medicao')
    if 'valor_minimo' in data:
        template.valor_minimo = data.get('valor_minimo')
    if 'valor_maximo' in data:
        template.valor_maximo = data.get('valor_maximo')
    if 'criticidade' in data:
        try:
            template.criticidade = CriticidadeTarefa(data.get('criticidade'))
        except ValueError:
            return {"error": "Criticidade inválida."}, 400
    if 'tempo_estimado' in data:
        template.tempo_estimado = data.get('tempo_estimado')
    if 'ordem' in data:
        template.ordem = int(data.get('ordem') or 0)
    if 'ativo' in data:
        template.ativo = bool(data.get('ativo'))

    if template.tipo_verificacao in (TipoVerificacao.MEDICAO, TipoVerificacao.TEMPERATURA) and not template.unidade_medicao:
        return {"error": "unidade_medicao é obrigatória para medições."}, 400
    if template.valor_minimo is not None and template.valor_maximo is not None:
        try:
            if float(template.valor_maximo) <= float(template.valor_minimo):
                return {"error": "valor_maximo deve ser maior que valor_minimo."}, 400
        except (TypeError, ValueError):
            return {"error": "valor_minimo/valor_maximo inválidos."}, 400

    db.session.commit()
    return template.to_dict(), 200


def delete_pop_template(template_id, restaurante_id):
    template = POPTemplate.query.get(template_id)
    if not template:
        return {"error": "Template não encontrado."}, 404
    if restaurante_id is not None and template.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    db.session.delete(template)
    db.session.commit()
    return {"message": "Template removido."}, 200


def toggle_pop_template_ativo(template_id, restaurante_id):
    template = POPTemplate.query.get(template_id)
    if not template:
        return {"error": "Template não encontrado."}, 404
    if restaurante_id is not None and template.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    template.ativo = not template.ativo
    db.session.commit()
    return template.to_dict(), 200


# ----- Listas -----

def list_pop_listas(restaurante_id):
    query = POPLista.query.filter_by(deletado=False)
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)
    listas = query.order_by(POPLista.nome.asc()).all()
    return [lista.to_dict() for lista in listas], 200


def get_pop_lista(lista_id, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    return lista.to_dict(), 200


def create_pop_lista(data, restaurante_id, usuario_id=None):
    rest_id, error = _resolve_restaurante_id(data, restaurante_id)
    if error:
        return error, 400
    if not data or not data.get('nome'):
        return {"error": "Nome é obrigatório."}, 400

    recorrencia_value = data.get('recorrencia', RecorrenciaLista.DIARIA.value)
    try:
        recorrencia_enum = RecorrenciaLista(recorrencia_value)
    except ValueError:
        return {"error": "Recorrência inválida."}, 400

    horario = None
    if data.get('horario_sugerido'):
        try:
            horario = datetime.strptime(data.get('horario_sugerido'), "%H:%M").time()
        except ValueError:
            return {"error": "horario_sugerido inválido. Use HH:MM."}, 400

    lista = POPLista(
        restaurante_id=rest_id,
        area_id=data.get('area_id'),
        categoria_id=data.get('categoria_id'),
        nome=data['nome'].strip(),
        descricao=data.get('descricao'),
        recorrencia=recorrencia_enum,
        dias_semana=data.get('dias_semana'),
        horario_sugerido=horario,
        publico=bool(data.get('publico', False)),
        criado_por_id=usuario_id
    )
    db.session.add(lista)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Lista já existe."}, 409

    template_ids = data.get('template_ids') or []
    if template_ids:
        for ordem, template_id in enumerate(template_ids):
            db.session.add(POPListaTarefa(lista_id=lista.id, template_id=template_id, ordem=ordem))
        db.session.commit()

    return lista.to_dict(), 201


def update_pop_lista(lista_id, data, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403

    if 'nome' in data and data['nome']:
        lista.nome = data['nome'].strip()
    if 'descricao' in data:
        lista.descricao = data.get('descricao')
    if 'area_id' in data:
        lista.area_id = data.get('area_id')
    if 'categoria_id' in data:
        lista.categoria_id = data.get('categoria_id')
    if 'recorrencia' in data:
        try:
            lista.recorrencia = RecorrenciaLista(data.get('recorrencia'))
        except ValueError:
            return {"error": "Recorrência inválida."}, 400
    if 'dias_semana' in data:
        lista.dias_semana = data.get('dias_semana')
    if 'horario_sugerido' in data:
        horario = data.get('horario_sugerido')
        if horario:
            try:
                lista.horario_sugerido = datetime.strptime(horario, "%H:%M").time()
            except ValueError:
                return {"error": "horario_sugerido inválido. Use HH:MM."}, 400
        else:
            lista.horario_sugerido = None
    if 'publico' in data:
        lista.publico = bool(data.get('publico'))
    if 'ativo' in data:
        lista.ativo = bool(data.get('ativo'))

    db.session.commit()
    return lista.to_dict(), 200


def delete_pop_lista(lista_id, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    lista.deletado = True
    db.session.commit()
    return {"message": "Lista removida."}, 200


def restore_pop_lista(lista_id, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    lista.deletado = False
    db.session.commit()
    return {"message": "Lista restaurada."}, 200


def list_pop_lista_tarefas(lista_id, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    tarefas = lista.tarefas.order_by(POPListaTarefa.ordem.asc()).all()
    return [tarefa.to_dict() for tarefa in tarefas], 200


def add_pop_lista_tarefas(lista_id, template_ids, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403

    existentes = {t.template_id for t in lista.tarefas.all()}
    ordem_base = lista.tarefas.count()
    adicionados = 0
    for offset, template_id in enumerate(template_ids or []):
        if template_id in existentes:
            continue
        db.session.add(POPListaTarefa(lista_id=lista.id, template_id=template_id, ordem=ordem_base + offset))
        adicionados += 1
    db.session.commit()
    return {"message": "Tarefas adicionadas.", "adicionados": adicionados}, 201


def remove_pop_lista_tarefa(lista_id, tarefa_id, restaurante_id):
    tarefa = POPListaTarefa.query.get(tarefa_id)
    if not tarefa or tarefa.lista_id != lista_id:
        return {"error": "Tarefa não encontrada."}, 404
    if restaurante_id is not None and tarefa.lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    db.session.delete(tarefa)
    db.session.commit()
    return {"message": "Tarefa removida."}, 200


def reorder_pop_lista_tarefas(lista_id, ordem_data, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403

    for item in ordem_data or []:
        tarefa = POPListaTarefa.query.get(item.get('id'))
        if tarefa and tarefa.lista_id == lista.id:
            tarefa.ordem = int(item.get('ordem') or 0)
    db.session.commit()
    return {"message": "Ordem atualizada."}, 200


def assign_pop_lista_colaboradores(lista_id, colaborador_ids, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403

    lista.colaboradores = []
    for user_id in colaborador_ids or []:
        usuario = Usuario.query.get(user_id)
        if not usuario:
            continue
        if usuario.role == UserRoles.SUPER_ADMIN:
            continue
        if usuario.restaurante_id != lista.restaurante_id:
            continue
        lista.colaboradores.append(usuario)
    db.session.commit()
    return {"message": "Colaboradores atualizados."}, 200


def list_pop_lista_colaboradores(lista_id, restaurante_id):
    lista = POPLista.query.get(lista_id)
    if not lista or lista.deletado:
        return {"error": "Lista não encontrada."}, 404
    if restaurante_id is not None and lista.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    colaboradores = lista.colaboradores.all()
    return [{"id": u.id, "nome": u.nome, "email": u.email} for u in colaboradores], 200


# ----- Execucoes (Colaborador/Admin) -----

def list_pop_listas_colaborador(user_id):
    usuario = Usuario.query.get(user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    if usuario.role == UserRoles.SUPER_ADMIN:
        listas = POPLista.query.filter_by(deletado=False, ativo=True).all()
    elif _is_admin_or_super_admin(usuario):
        listas = POPLista.query.filter_by(restaurante_id=usuario.restaurante_id, deletado=False, ativo=True).all()
    else:
        listas = POPLista.query.filter(
            POPLista.deletado.is_(False),
            POPLista.ativo.is_(True),
            POPLista.restaurante_id == usuario.restaurante_id,
            or_(POPLista.publico.is_(True), POPLista.colaboradores.any(id=usuario.id))
        ).all()

    return {"listas": [lista.to_dict() for lista in listas]}, 200


def start_pop_execucao(user_id, data):
    usuario = Usuario.query.get(user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404
    if not data or not data.get('lista_id'):
        return {"error": "lista_id é obrigatório."}, 400

    lista = POPLista.query.get(data.get('lista_id'))
    if not lista or lista.deletado or not lista.ativo:
        return {"error": "Lista não encontrada."}, 404
    if not _user_can_access_pop_lista(usuario, lista):
        return {"error": "Acesso negado."}, 403

    data_referencia = data.get('data_referencia')
    if data_referencia:
        try:
            data_ref = datetime.fromisoformat(data_referencia).date()
        except ValueError:
            return {"error": "data_referencia inválida. Use YYYY-MM-DD."}, 400
    else:
        data_ref = brasilia_now().date()

    existente = POPExecucao.query.filter_by(
        lista_id=lista.id,
        usuario_id=usuario.id,
        data_referencia=data_ref
    ).first()
    if existente:
        return {"error": "Execução já criada para esta lista neste dia."}, 400

    execucao = POPExecucao(
        lista_id=lista.id,
        usuario_id=usuario.id,
        restaurante_id=lista.restaurante_id,
        data_referencia=data_ref,
        status=StatusExecucao.EM_ANDAMENTO
    )
    db.session.add(execucao)
    db.session.flush()

    tarefas = lista.tarefas.order_by(POPListaTarefa.ordem.asc()).all()
    itens = []
    for ordem, tarefa in enumerate(tarefas):
        template = tarefa.template
        if not template or not template.ativo:
            continue
        item = POPExecucaoItem(
            execucao_id=execucao.id,
            template_id=template.id,
            lista_tarefa_id=tarefa.id,
            titulo=template.titulo,
            descricao=template.descricao,
            tipo_verificacao=template.tipo_verificacao,
            unidade_medicao=template.unidade_medicao,
            ordem=ordem
        )
        itens.append(item)
        db.session.add(item)

    if not itens:
        db.session.rollback()
        return {"error": "Lista não possui tarefas ativas."}, 400

    execucao.total_tarefas = len(itens)
    db.session.commit()
    return execucao.to_dict(include_itens=True), 201


def get_pop_execucao(user_id, execucao_id):
    execucao = POPExecucao.query.get(execucao_id)
    if not execucao:
        return {"error": "Execução não encontrada."}, 404
    usuario = Usuario.query.get(user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404
    if not _is_admin_or_super_admin(usuario) and execucao.usuario_id != usuario.id:
        return {"error": "Acesso negado."}, 403
    return execucao.to_dict(include_itens=True), 200


def list_pop_execucoes(user_id, filtros=None):
    usuario = Usuario.query.get(user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404

    filtros = filtros or {}
    query = POPExecucao.query

    if usuario.role == UserRoles.SUPER_ADMIN:
        pass
    elif _is_admin_or_super_admin(usuario):
        query = query.filter_by(restaurante_id=usuario.restaurante_id)
    else:
        query = query.filter_by(usuario_id=usuario.id)

    if filtros.get('data_inicio'):
        data_inicio_raw = filtros.get('data_inicio')
        if isinstance(data_inicio_raw, str):
            try:
                data_inicio = datetime.fromisoformat(data_inicio_raw).date()
            except ValueError:
                return {"error": "data_inicio inválida. Use YYYY-MM-DD."}, 400
        else:
            data_inicio = data_inicio_raw
        query = query.filter(POPExecucao.data_referencia >= data_inicio)
    if filtros.get('data_fim'):
        data_fim_raw = filtros.get('data_fim')
        if isinstance(data_fim_raw, str):
            try:
                data_fim = datetime.fromisoformat(data_fim_raw).date()
            except ValueError:
                return {"error": "data_fim inválida. Use YYYY-MM-DD."}, 400
        else:
            data_fim = data_fim_raw
        query = query.filter(POPExecucao.data_referencia <= data_fim)
    if filtros.get('status'):
        query = query.filter_by(status=filtros.get('status'))
    if not filtros.get('include_arquivados'):
        query = query.filter_by(arquivado=False)

    execucoes = query.order_by(POPExecucao.data_referencia.desc()).all()
    return {"execucoes": [e.to_dict() for e in execucoes]}, 200


def list_pop_execucoes_hoje(user_id):
    hoje = brasilia_now().date()
    return list_pop_execucoes(user_id, {"data_inicio": hoje, "data_fim": hoje})


def update_pop_execucao_item(user_id, execucao_id, item_id, data):
    usuario = Usuario.query.get(user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404
    execucao = POPExecucao.query.get(execucao_id)
    if not execucao:
        return {"error": "Execução não encontrada."}, 404
    if not _is_admin_or_super_admin(usuario) and execucao.usuario_id != usuario.id:
        return {"error": "Acesso negado."}, 403

    item = POPExecucaoItem.query.get(item_id)
    if not item or item.execucao_id != execucao.id:
        return {"error": "Item não encontrado."}, 404

    if 'valor_medido' in data:
        item.valor_medido = data.get('valor_medido')
        template = POPTemplate.query.get(item.template_id)
        if template and template.valor_minimo is not None and template.valor_maximo is not None:
            try:
                medido = float(item.valor_medido)
                min_val = float(template.valor_minimo)
                max_val = float(template.valor_maximo)
                item.dentro_padrao = min_val <= medido <= max_val
            except (TypeError, ValueError):
                item.dentro_padrao = None
        else:
            item.dentro_padrao = None

    if 'observacoes' in data:
        item.observacoes = data.get('observacoes')
    if 'tem_desvio' in data:
        item.tem_desvio = bool(data.get('tem_desvio'))
    if 'descricao_desvio' in data:
        item.descricao_desvio = data.get('descricao_desvio')
    if 'acao_corretiva' in data:
        item.acao_corretiva = data.get('acao_corretiva')

    if 'concluido' in data:
        item.concluido = bool(data.get('concluido'))
        item.concluido_em = brasilia_now() if item.concluido else None

    total = execucao.total_tarefas or execucao.itens.count()
    concluidas = execucao.itens.filter_by(concluido=True).count()
    desvios = execucao.itens.filter_by(tem_desvio=True).count()
    execucao.total_tarefas = total
    execucao.tarefas_concluidas = concluidas
    execucao.tarefas_com_desvio = desvios
    execucao.progresso = int((concluidas / total) * 100) if total else 0

    db.session.commit()
    return item.to_dict(), 200


def finalizar_pop_execucao(user_id, execucao_id, data):
    usuario = Usuario.query.get(user_id)
    if not usuario:
        return {"error": "Usuário não encontrado."}, 404
    execucao = POPExecucao.query.get(execucao_id)
    if not execucao:
        return {"error": "Execução não encontrada."}, 404
    if not _is_admin_or_super_admin(usuario) and execucao.usuario_id != usuario.id:
        return {"error": "Acesso negado."}, 403

    obrigatorias = POPListaTarefa.query.filter_by(lista_id=execucao.lista_id, obrigatoria=True).all()
    obrigatorias_ids = {t.id for t in obrigatorias}
    concluidas_ids = {item.lista_tarefa_id for item in execucao.itens.filter_by(concluido=True).all()}

    if obrigatorias_ids and not obrigatorias_ids.issubset(concluidas_ids):
        return {"error": "Nem todas as tarefas obrigatórias foram concluídas."}, 400

    execucao.status = StatusExecucao.CONCLUIDO
    execucao.finalizado_em = brasilia_now()
    if data:
        execucao.observacoes = data.get('observacoes')
        execucao.assinatura_digital = data.get('assinatura_digital')
    db.session.commit()
    return execucao.to_dict(include_itens=True), 200


# ----- Auditoria/Admin -----

def list_pop_execucoes_admin(restaurante_id, filtros=None, executor_id=None):
    if restaurante_id:
        _apply_pop_auto_archive(restaurante_id, executor_id)

    filtros = filtros or {}
    query = POPExecucao.query
    if restaurante_id is not None:
        query = query.filter_by(restaurante_id=restaurante_id)

    if filtros.get('status'):
        query = query.filter_by(status=filtros.get('status'))
    if filtros.get('usuario_id'):
        try:
            query = query.filter_by(usuario_id=int(filtros.get('usuario_id')))
        except (TypeError, ValueError):
            return {"error": "usuario_id inválido."}, 400
    if filtros.get('lista_id'):
        try:
            query = query.filter_by(lista_id=int(filtros.get('lista_id')))
        except (TypeError, ValueError):
            return {"error": "lista_id inválido."}, 400
    if filtros.get('data_inicio'):
        data_inicio_raw = filtros.get('data_inicio')
        if isinstance(data_inicio_raw, str):
            try:
                data_inicio = datetime.fromisoformat(data_inicio_raw).date()
            except ValueError:
                return {"error": "data_inicio inválida. Use YYYY-MM-DD."}, 400
        else:
            data_inicio = data_inicio_raw
        query = query.filter(POPExecucao.data_referencia >= data_inicio)
    if filtros.get('data_fim'):
        data_fim_raw = filtros.get('data_fim')
        if isinstance(data_fim_raw, str):
            try:
                data_fim = datetime.fromisoformat(data_fim_raw).date()
            except ValueError:
                return {"error": "data_fim inválida. Use YYYY-MM-DD."}, 400
        else:
            data_fim = data_fim_raw
        query = query.filter(POPExecucao.data_referencia <= data_fim)
    if not filtros.get('include_arquivados'):
        query = query.filter_by(arquivado=False)

    execucoes = query.order_by(POPExecucao.data_referencia.desc()).all()
    return {"execucoes": [e.to_dict() for e in execucoes]}, 200


def review_pop_execucao(execucao_id, data, restaurante_id, reviewer_id=None):
    execucao = POPExecucao.query.get(execucao_id)
    if not execucao:
        return {"error": "Execução não encontrada."}, 404
    if restaurante_id is not None and execucao.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403

    execucao.revisado = bool(data.get('revisado', True))
    execucao.observacoes_revisao = data.get('observacoes_revisao')
    execucao.revisado_por_id = reviewer_id
    execucao.revisado_em = brasilia_now()
    db.session.commit()
    return execucao.to_dict(include_itens=True), 200


def archive_pop_execucao(execucao_id, restaurante_id, executor_id=None):
    execucao = POPExecucao.query.get(execucao_id)
    if not execucao:
        return {"error": "Execução não encontrada."}, 404
    if restaurante_id is not None and execucao.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    execucao.arquivado = True
    execucao.arquivado_em = brasilia_now()
    execucao.arquivado_por_id = executor_id
    db.session.commit()
    return execucao.to_dict(), 200


def unarchive_pop_execucao(execucao_id, restaurante_id):
    execucao = POPExecucao.query.get(execucao_id)
    if not execucao:
        return {"error": "Execução não encontrada."}, 404
    if restaurante_id is not None and execucao.restaurante_id != restaurante_id:
        return {"error": "Acesso negado."}, 403
    execucao.arquivado = False
    execucao.arquivado_em = None
    execucao.arquivado_por_id = None
    db.session.commit()
    return execucao.to_dict(), 200


def run_pop_auto_archive(restaurante_id, executor_id=None):
    if not restaurante_id:
        return {"error": "Restaurante é obrigatório."}, 400
    total = _apply_pop_auto_archive(restaurante_id, executor_id)
    return {"arquivados": total}, 200


# ===============================================
# SERVIÇO: Configuração de Threshold/Fardo
# ===============================================

def atualizar_config_item_ref(lista_id, item_id, quantidade_minima, quantidade_por_fardo, restaurante_id):
    """
    Atualiza a configuração de threshold/fardo para um item de uma lista.

    Args:
        lista_id: ID da lista
        item_id: ID do item (ListaMaeItem)
        quantidade_minima: Quantidade mínima (threshold) para gerar pedido
        quantidade_por_fardo: Quantidade a ser pedida quando atingir o mínimo
        restaurante_id: ID do restaurante para validação

    Returns:
        Tuple (dados, status_code)
    """
    # Verificar se a lista pertence ao restaurante
    lista = Lista.query.filter_by(id=lista_id, restaurante_id=restaurante_id, deletado=False).first()
    if not lista:
        return {"error": "Lista não encontrada."}, 404

    # Buscar o item_ref
    ref = ListaItemRef.query.filter_by(lista_id=lista_id, item_id=item_id).first()
    if not ref:
        return {"error": "Item não encontrado na lista."}, 404

    # Validar quantidade_por_fardo
    if quantidade_por_fardo is not None and quantidade_por_fardo <= 0:
        return {"error": "Quantidade por fardo deve ser maior que zero."}, 400

    # Atualizar campos
    if quantidade_minima is not None:
        ref.quantidade_minima = float(quantidade_minima)
    ref.quantidade_por_fardo = float(quantidade_por_fardo) if quantidade_por_fardo else 1.0
    ref.usa_threshold = True  # Sempre True na nova lógica simplificada

    db.session.commit()

    return ref.to_dict(), 200
